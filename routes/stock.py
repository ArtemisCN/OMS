"""备件库存管理"""
# 导入标准库与三方库
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, g, abort
from flask_login import login_required, current_user
from routes.auth import admin_required
# 导入数据模型与审计日志工具
from models import db, SparePart, StockRecord, StorageLocation, WorkOrder, log_audit
from datetime import datetime

# 创建备件库存管理蓝图，URL 前缀 /stock
stock_bp = Blueprint('stock', __name__, url_prefix='/stock')


@stock_bp.route('/')
@login_required
def index():
    """备件库存总览"""
    # 获取查询参数：分类筛选 / 仅显示低库存
    category = request.args.get('category', '').strip()
    low_only = request.args.get('low_only', '') == '1'
    q = request.args.get('q', '').strip()
    # 构建查询，按库存量升序排列
    query = SparePart.query
    if category:
        query = query.filter_by(category=category)
    if q:
        safe_q = q.replace('%', '\\%').replace('_', '\\_')
        like = f'%{safe_q}%'
        query = query.filter(db.or_(
            SparePart.name.ilike(like),
            SparePart.brand.ilike(like),
            SparePart.model_no.ilike(like),
            SparePart.location.ilike(like),
            SparePart.notes.ilike(like),
        ))
    parts = query.order_by(SparePart.stock.asc()).all()
    # 低库存过滤（仅显示库存不足项）
    if low_only:
        parts = [p for p in parts if p.is_low]

    # 统计总览数据：种类数、总库存量、低库存数量、所有分类列表
    total_types = SparePart.query.count()
    total_stock = db.session.query(db.func.sum(SparePart.stock)).scalar() or 0
    low_count = sum(1 for p in SparePart.query.all() if p.is_low)
    categories = [r[0] for r in db.session.query(SparePart.category).distinct().all() if r[0]]

    # 近日出入库动态（最近10条）
    recent_records = StockRecord.query.order_by(StockRecord.created_at.desc()).limit(10).all()

    return render_template('stock/index.html', parts=parts, categories=categories,
                           total_types=total_types, total_stock=total_stock,
                           low_count=low_count, recent_records=recent_records,
                           current_category=category, low_only=low_only, q=q,
                           now=datetime.now())


@stock_bp.route('/departments')
@login_required
def get_departments():
    """获取所有科室列表（用于一键出库）"""
    depts = [r[0] for r in db.session.query(WorkOrder.department).distinct().order_by(WorkOrder.department).all() if r[0]]
    return jsonify({'departments': depts})


@stock_bp.route('/parts-json')
@login_required
def parts_json():
    """获取备件列表JSON（用于一键出库弹窗）"""
    parts = SparePart.query.order_by(SparePart.category, SparePart.name).all()
    return jsonify({'parts': [{
        'id': p.id, 'name': p.name, 'category': p.category,
        'brand': p.brand, 'model_no': p.model_no,
        'stock': p.stock, 'unit': p.unit, 'is_low': p.is_low,
    } for p in parts]})


@stock_bp.route('/detail/<int:part_id>')
@login_required
def detail(part_id):
    """备件详情 + 出入库记录"""
    # 按 ID 查询备件，不存在则 404
    part = SparePart.query.filter(SparePart.id == part_id, SparePart.hospital_id == getattr(g, 'hospital_id', None)).first()
    if not part:
        abort(404)
    # 查询该备件所有出入库记录，按时间倒序
    records = StockRecord.query.filter_by(part_id=part_id)\
        .order_by(StockRecord.created_at.desc()).all()
    return render_template('stock/detail.html', part=part, records=records,
                           now=datetime.now())


@stock_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    # 权限校验：仅管理员可新增备件
    if not current_user.is_admin:
        flash('仅管理员可操作', 'danger')
        return redirect(url_for('stock.index'))
    # 处理表单提交（POST）
    if request.method == 'POST':
        try:
            # 从表单构建备件对象
            part = SparePart(
                name=request.form.get('name', ''),
                category=request.form.get('category', '其他'),
                brand=request.form.get('brand', ''),
                model_no=request.form.get('model_no', ''),
                unit=request.form.get('unit', '个'),
                stock=int(request.form.get('stock', 0)),
                min_stock=int(request.form.get('min_stock', 5)),
                location=request.form.get('location', ''),
                notes=request.form.get('notes', ''),
            )
            db.session.add(part)
            db.session.commit()

            # 初次入库 > 0 时自动生成入库流水
            if part.stock > 0:
                record = StockRecord(
                    part_id=part.id, type='in', quantity=part.stock,
                    balance=part.stock, operator=current_user.display_name,
                    note='初始入库'
                )
                db.session.add(record)
                db.session.commit()

            # 记录审计日志
            log_audit('create', 'spare_part', current_user.display_name,
                      target_id=part.id, target_desc=f'新增备件 {part.name}')
            flash('备件添加成功', 'success')
            return redirect(url_for('stock.index'))
        except Exception as e:
            # 发生异常时回滚数据库
            db.session.rollback()
            flash(f'添加失败: {str(e)}', 'danger')
    # GET 请求：加载表单所需数据（存放位置列表、已有型号列表）
    locations = StorageLocation.query.filter_by(is_active=True).order_by(StorageLocation.sort_order, StorageLocation.name).all()
    model_list = [r[0] for r in db.session.query(SparePart.model_no).distinct().all() if r[0]]
    return render_template('stock/form.html', part=None, locations=locations, model_list=model_list, now=datetime.now())


@stock_bp.route('/edit/<int:part_id>', methods=['GET', 'POST'])
@login_required
def edit(part_id):
    # 权限校验：仅管理员可修改
    if not current_user.is_admin:
        flash('仅管理员可操作', 'danger')
        return redirect(url_for('stock.index'))
    # 查询目标备件，不存在则 404
    part = SparePart.query.filter(SparePart.id == part_id, SparePart.hospital_id == getattr(g, 'hospital_id', None)).first()
    if not part:
        abort(404)
    # 处理表单提交（POST）
    if request.method == 'POST':
        try:
            # 更新备件各字段
            part.name = request.form.get('name', '')
            part.category = request.form.get('category', '其他')
            part.brand = request.form.get('brand', '')
            part.model_no = request.form.get('model_no', '')
            part.unit = request.form.get('unit', '个')
            part.stock = int(request.form.get('stock', 0))
            part.min_stock = int(request.form.get('min_stock', 5))
            part.location = request.form.get('location', '')
            part.notes = request.form.get('notes', '')
            db.session.commit()
            # 记录审计日志
            log_audit('update', 'spare_part', current_user.display_name,
                      target_id=part.id, target_desc=f'修改备件 {part.name}')
            flash('备件修改成功', 'success')
            return redirect(url_for('stock.index'))
        except Exception as e:
            # 异常时回滚
            db.session.rollback()
            flash(f'修改失败: {str(e)}', 'danger')
    # GET 请求：加载表单下拉数据
    locations = StorageLocation.query.filter_by(is_active=True).order_by(StorageLocation.sort_order, StorageLocation.name).all()
    model_list = [r[0] for r in db.session.query(SparePart.model_no).distinct().all() if r[0]]
    return render_template('stock/form.html', part=part, locations=locations, model_list=model_list, now=datetime.now())


@stock_bp.route('/delete/<int:part_id>', methods=['POST'])
@login_required
def delete(part_id):
    # 权限校验：仅管理员可删除
    if not current_user.is_admin:
        flash('权限不足', 'danger')
        return redirect(url_for('stock.index'))
    # 查询目标备件
    part = SparePart.query.filter(SparePart.id == part_id, SparePart.hospital_id == getattr(g, 'hospital_id', None)).first()
    if not part:
        abort(404)
    name = part.name
    # 级联删除关联的出入库记录
    StockRecord.query.filter_by(part_id=part_id).delete()
    # 记录审计日志
    log_audit('delete', 'spare_part', current_user.display_name,
              target_id=part.id, target_desc=f'删除备件 {name}')
    # 删除备件本身并提交
    db.session.delete(part)
    db.session.commit()
    flash(f'备件「{name}」已删除', 'success')
    return redirect(url_for('stock.index'))


@stock_bp.route('/out-records')
@login_required
def out_records():
    """出库总记录：汇总所有备件的出库记录"""
    # 搜索/筛选
    q = request.args.get('q', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    department = request.args.get('department', '').strip()

    query = StockRecord.query.filter_by(type='out')\
        .join(SparePart, StockRecord.part_id == SparePart.id)\
        .add_columns(
            SparePart.name.label('part_name'),
            SparePart.category.label('part_category'),
            SparePart.brand.label('part_brand'),
            SparePart.model_no.label('part_model'),
            SparePart.unit.label('part_unit'),
        )

    if q:
        query = query.filter(SparePart.name.ilike(f'%{q}%'))
    if department:
        query = query.filter(StockRecord.department.ilike(f'%{department}%'))
    if date_from:
        try:
            dt = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(StockRecord.created_at >= dt)
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, '%Y-%m-%d')
            query = query.filter(StockRecord.created_at <= dt.replace(hour=23, minute=59, second=59))
        except ValueError:
            pass

    records = query.order_by(StockRecord.created_at.desc()).all()

    # 获取所有出库科室列表（用于筛选下拉）
    departments = [r[0] for r in
                   db.session.query(StockRecord.department)
                   .filter(StockRecord.type == 'out', StockRecord.department != '')
                   .distinct().order_by(StockRecord.department).all()]

    return render_template('stock/out_records.html', records=records,
                           departments=departments, q=q,
                           date_from=date_from, date_to=date_to,
                           department=department, now=datetime.now())


@stock_bp.route('/inout', methods=['POST'])
@login_required
def inout():
    """出入库操作"""
    # 权限校验：仅管理员可执行出入库
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    # 解析表单参数：备件ID、操作类型(in/out)、数量、备注
    part_id = request.form.get('part_id', type=int)
    action = request.form.get('action')  # in/out
    qty = request.form.get('quantity', type=int, default=0)
    note = request.form.get('note', '')

    # 参数完整性校验
    if not part_id or not action or qty <= 0:
        return jsonify({'ok': False, 'msg': '参数错误'}), 400

    # 查询备件，校验出库库存是否充足
    part = SparePart.query.filter(SparePart.id == part_id, SparePart.hospital_id == getattr(g, 'hospital_id', None)).first()
    if not part:
        abort(404)
    if action == 'out' and part.stock < qty:
        return jsonify({'ok': False, 'msg': f'库存不足（当前 {part.stock}{part.unit}）'}), 400

    # 更新库存数量：入库增 / 出库减
    if action == 'in':
        part.stock += qty
    else:
        part.stock -= qty

    # 创建出入库流水记录
    record = StockRecord(
        part_id=part_id, type=action, quantity=qty,
        balance=part.stock, operator=current_user.display_name,
        work_order_id=request.form.get('work_order_id', type=int),
        note=note
    )
    db.session.add(record)
    db.session.commit()

    # 记录审计日志（中文描述）
    action_name = '入库' if action == 'in' else '出库'
    log_audit(action, 'spare_part', current_user.display_name,
              target_id=part.id, target_desc=f'备件{action_name} {part.name} ×{qty}')
    return jsonify({'ok': True, 'balance': part.stock})


@stock_bp.route('/batch-out', methods=['POST'])
@login_required
def batch_out():
    """一键出库：选择科室，批量出库"""
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    department = request.form.get('department', '').strip()
    if not department:
        return jsonify({'ok': False, 'msg': '请选择科室'}), 400
    items = request.form.getlist('items[]')  # part_id:quantity
    if not items:
        return jsonify({'ok': False, 'msg': '请选择出库物品'}), 400
    out_records = []
    for item in items:
        parts = item.split(':')
        if len(parts) != 2:
            continue
        part_id, qty = parts
        try:
            part_id, qty = int(part_id), int(qty)
        except ValueError:
            continue
        if qty <= 0:
            continue
        part = SparePart.query.filter(SparePart.id == part_id, SparePart.hospital_id == getattr(g, 'hospital_id', None)).first()
        if not part:
            continue
        if part.stock < qty:
            return jsonify({'ok': False, 'msg': f'「{part.name}」库存不足（当前 {part.stock}{part.unit}）'}), 400
        part.stock -= qty
        record = StockRecord(
            part_id=part_id, type='out', quantity=qty,
            balance=part.stock, operator=current_user.display_name,
            department=department,
            note=f'一键出库至{department}',
        )
        db.session.add(record)
        out_records.append({'part': part.name, 'qty': qty, 'unit': part.unit})
    db.session.commit()
    log_audit('batch_out', 'spare_part', current_user.display_name,
              target_desc=f'一键出库至{department}，共{len(out_records)}项')
    return jsonify({'ok': True, 'msg': f'已出库 {len(out_records)} 项物品至「{department}」', 'records': out_records})


@stock_bp.route('/batch-out-sign/init', methods=['POST'])
@login_required
def batch_out_sign_init():
    """创建出库签名请求，返回token和二维码URL"""
    from models import StockSignRequest
    import secrets, json as pyjson
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    department = request.form.get('department', '').strip()
    if not department:
        return jsonify({'ok': False, 'msg': '请选择科室'}), 400
    items = request.form.getlist('items[]')
    if not items:
        return jsonify({'ok': False, 'msg': '请选择出库物品'}), 400
    # 检查库存
    for item in items:
        parts = item.split(':')
        if len(parts) != 2: continue
        try:
            pid, qty = int(parts[0]), int(parts[1])
        except ValueError: continue
        if qty <= 0: continue
        part = SparePart.query.filter(SparePart.id == pid, SparePart.hospital_id == getattr(g, 'hospital_id', None)).first()
        if not part: continue
        if part.stock < qty:
            return jsonify({'ok': False, 'msg': f'「{part.name}」库存不足'}), 400
    token = secrets.token_hex(16)
    sign_req = StockSignRequest(
        token=token, department=department,
        items_json=pyjson.dumps(items),
        operator=current_user.display_name,
    )
    db.session.add(sign_req)
    db.session.commit()
    qr_url = f'https://demolin.cn/forms/sign-stock/{token}'
    return jsonify({'ok': True, 'token': token, 'qr_url': qr_url})


@stock_bp.route('/batch-out-sign/execute/<token>', methods=['POST'])
@login_required
def batch_out_sign_execute(token):
    """签名完成后执行出库"""
    from models import StockSignRequest
    import json as pyjson
    from datetime import datetime
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    sign_req = StockSignRequest.query.filter_by(token=token).first()
    if not sign_req:
        return jsonify({'ok': False, 'msg': '签名请求不存在'}), 404
    if sign_req.status != 'signed':
        return jsonify({'ok': False, 'msg': '尚未签名'}), 400
    if sign_req.status == 'completed':
        return jsonify({'ok': False, 'msg': '已执行'}), 400
    department = sign_req.department
    items = pyjson.loads(sign_req.items_json)
    signature = sign_req.signature
    out_records = []
    for item in items:
        parts = item.split(':')
        if len(parts) != 2: continue
        try:
            pid, qty = int(parts[0]), int(parts[1])
        except ValueError: continue
        if qty <= 0: continue
        part = SparePart.query.filter(SparePart.id == pid, SparePart.hospital_id == getattr(g, 'hospital_id', None)).first()
        if not part: continue
        if part.stock < qty:
            return jsonify({'ok': False, 'msg': f'「{part.name}」库存不足'}), 400
        part.stock -= qty
        record = StockRecord(
            part_id=pid, type='out', quantity=qty,
            balance=part.stock, operator=sign_req.operator,
            department=department, note=f'扫码签名出库至{department}',
            signature=signature,
        )
        db.session.add(record)
        out_records.append({'part': part.name, 'qty': qty, 'unit': part.unit})
    sign_req.status = 'completed'
    sign_req.completed_at = datetime.now()
    db.session.commit()
    log_audit('batch_out', 'spare_part', sign_req.operator,
              target_desc=f'扫码签名出库至{department}，共{len(out_records)}项')
    return jsonify({'ok': True, 'msg': f'已出库 {len(out_records)} 项物品至「{department}」', 'records': out_records})


@stock_bp.route('/import-excel', methods=['POST'])
@login_required
def import_stock_excel():
    """从Excel导入备件"""
    # 权限校验
    if not current_user.is_admin:
        flash('仅管理员可操作', 'danger')
        return redirect(url_for('stock.index'))
    # 检查服务端是否安装 openpyxl
    try:
        import openpyxl
    except ImportError:
        flash('服务端缺少 openpyxl 库', 'danger')
        return redirect(url_for('stock.index'))
    # 验证上传文件
    file = request.files.get('file')
    if not file:
        flash('请选择文件', 'danger')
        return redirect(url_for('stock.index'))
    try:
        from models import SparePart, StorageLocation
        # 加载工作簿并读取所有行
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash('空文件', 'danger')
            return redirect(url_for('stock.index'))
        # 解析表头：转小写、去空格，建立列名到字段的映射
        header = [str(c or '').strip().lower() for c in rows[0]]
        data_rows = rows[1:]
        # 中英文列名映射表
        col_map = {
            '备件名称': 'name', '名称': 'name', 'name': 'name',
            '分类': 'category', '类别': 'category', 'category': 'category',
            '型号': 'model_no', '规格型号': 'model_no', '规格': 'model_no', 'model': 'model_no',
            '品牌': 'brand', 'brand': 'brand',
            '库存': 'stock', '数量': 'stock', '库存量': 'stock', 'stock': 'stock',
            '最低库存': 'min_stock', '预警': 'min_stock', '预警量': 'min_stock',
            '单位': 'unit', 'unit': 'unit',
            '单价': 'price', 'price': 'price',
            '位置': 'location', '存放位置': 'location', 'location_id': 'location',
            '备注': 'notes', 'note': 'notes',
        }
        imported = 0; skipped = 0; errors = []
        # 逐行处理数据（从第2行开始）
        for idx, row in enumerate(data_rows, start=2):
            # 跳过完全空行
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            # 将行数据按列名映射为字典
            row_dict = {}
            for ci, cn in enumerate(header):
                f = col_map.get(cn)
                if f:
                    v = row[ci] if ci < len(row) else None
                    if v is not None: v = str(v).strip()
                    row_dict[f] = v
            name = row_dict.get('name', '')
            if not name:
                skipped += 1; continue
            # 按名称查找已存在的备件：存在则更新，不存在则新建
            existing = SparePart.query.filter_by(name=name).first()
            if existing:
                # 更新已有备件各字段
                if row_dict.get('category'): existing.category = row_dict['category']
                if row_dict.get('model_no'): existing.model_no = row_dict['model_no']
                if row_dict.get('brand'): existing.brand = row_dict['brand']
                if row_dict.get('stock'): existing.stock = int(float(row_dict['stock']))
                if row_dict.get('min_stock'): existing.min_stock = int(float(row_dict['min_stock']))
                if row_dict.get('unit'): existing.unit = row_dict['unit']
                if row_dict.get('price'): existing.price = float(row_dict['price'])
                if row_dict.get('location'):
                    # 根据名称查找存放位置ID
                    loc = StorageLocation.query.filter_by(name=row_dict['location']).first()
                    if loc: existing.location_id = loc.id
                if row_dict.get('notes'): existing.notes = row_dict['notes']
                imported += 1; continue
            try:
                # 创建新备件记录
                loc_id = None
                if row_dict.get('location'):
                    loc = StorageLocation.query.filter_by(name=row_dict['location']).first()
                    if loc: loc_id = loc.id
                part = SparePart(
                    name=name, category=row_dict.get('category', ''),
                    model_no=row_dict.get('model_no', ''),
                    brand=row_dict.get('brand', ''),
                    stock=int(float(row_dict['stock'])) if row_dict.get('stock') else 0,
                    min_stock=int(float(row_dict['min_stock'])) if row_dict.get('min_stock') else 5,
                    unit=row_dict.get('unit', '个'),
                    price=float(row_dict['price']) if row_dict.get('price') else 0,
                    location_id=loc_id,
                    notes=row_dict.get('notes', ''),
                )
                db.session.add(part)
                imported += 1
            except Exception as e:
                # 单行解析失败，记录错误并跳过
                errors.append(f'第{idx}行({name}): {str(e)}'); skipped += 1
        # 提交全部变更并返回导入结果
        db.session.commit()
        msg = f'导入完成: 成功 {imported} 条'
        if skipped: msg += f', 跳过 {skipped} 条'
        if errors: msg += '<br>' + '<br>'.join(errors[:10])
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'导入失败: {str(e)}', 'danger')
    return redirect(url_for('stock.index'))


@stock_bp.route('/export-template')
@login_required
def export_stock_template():
    """下载备件导入表头模板"""
    # 检查 openpyxl 可用性
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        flash('服务端缺少 openpyxl 库', 'danger')
        return redirect(url_for('stock.index'))
    # 创建工作簿并命名工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '备件导入模板'
    # 定义表头行内容
    headers = ['备件名称', '分类', '规格型号', '品牌', '库存量', '最低库存', '单位', '单价', '存放位置', '备注']
    # 设置表头样式：加粗白色字体 + 紫色填充 + 细边框
    hf = Font(bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center', vertical='center'); c.border = tb
    # 调整首列宽度便于输入
    ws.column_dimensions['A'].width = 20
    # 写入内存流并返回文件下载响应
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='备件导入模板.xlsx')

# ============ 备件领用审批 ============

@stock_bp.route('/requests')
@login_required
def request_list():
    """领用申请列表"""
    from models import PartRequest
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    query = PartRequest.query
    if status in ('pending', 'approved', 'rejected'):
        query = query.filter(PartRequest.status == status)
    pagination = query.order_by(PartRequest.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('stock/requests.html', pagination=pagination, orders=pagination.items, status=status)


@stock_bp.route('/requests/create', methods=['GET', 'POST'])
@login_required
def request_create():
    """创建领用申请"""
    from models import SparePart, PartRequest, WorkOrder
    if request.method == 'POST':
        part_id = request.form.get('part_id', type=int)
        quantity = request.form.get('quantity', 1, type=int)
        wo_id = request.form.get('work_order_id', type=int)
        reason = request.form.get('reason', '').strip()
        if not part_id:
            flash('请选择备件', 'danger')
            return redirect(url_for('stock.request_create'))
        part = db.session.get(SparePart, part_id)
        if not part:
            flash('备件不存在', 'danger')
            return redirect(url_for('stock.request_create'))
        if quantity < 1:
            flash('数量必须大于0', 'danger')
            return redirect(url_for('stock.request_create'))
        req = PartRequest(
            part_id=part_id,
            quantity=quantity,
            requester=current_user.display_name or current_user.username,
            work_order_id=wo_id if wo_id else None,
            reason=reason,
            status='pending',
        )
        db.session.add(req)
        db.session.commit()
        flash(f'✅ 领用申请已提交（{part.name} x{quantity}），等待审批', 'success')
        return redirect(url_for('stock.request_list'))
    parts = SparePart.query.order_by(SparePart.name).all()
    work_orders = WorkOrder.query.filter(
        WorkOrder.person == (current_user.display_name or current_user.username),
        WorkOrder.status.in_(['pending', 'in_progress'])
    ).order_by(WorkOrder.created_at.desc()).all()
    return render_template('stock/request_create.html', parts=parts, work_orders=work_orders)


@stock_bp.route('/requests/<int:rid>/approve', methods=['POST'])
@login_required
def request_approve(rid):
    """审批通过：扣减库存 + 创建出库记录"""
    from models import PartRequest, StockRecord
    req = PartRequest.query.get_or_404(rid)
    if req.status != 'pending':
        flash('该申请已被处理', 'warning')
        return redirect(url_for('stock.request_list'))
    part = req.part
    if part.stock < req.quantity:
        flash(f'❌ 库存不足（当前 {part.stock}，需要 {req.quantity}）', 'danger')
        return redirect(url_for('stock.request_list'))
    # 扣减库存
    part.stock -= req.quantity
    # 记录出库流水
    record = StockRecord(
        part_id=part.id,
        type='out',
        quantity=req.quantity,
        balance=part.stock,
        operator=current_user.display_name or current_user.username,
        work_order_id=req.work_order_id,
        note=f'领用审批出库：{req.requester} - {req.reason}',
    )
    db.session.add(record)
    # 更新申请状态
    req.status = 'approved'
    req.approver = current_user.display_name or current_user.username
    req.approved_at = datetime.now()
    db.session.commit()
    flash(f'✅ 已批准 {part.name} x{req.quantity}，库存已扣减', 'success')
    return redirect(url_for('stock.request_list'))


@stock_bp.route('/requests/<int:rid>/reject', methods=['POST'])
@login_required
def request_reject(rid):
    """拒绝申请"""
    from models import PartRequest
    req = PartRequest.query.get_or_404(rid)
    if req.status != 'pending':
        flash('该申请已被处理', 'warning')
        return redirect(url_for('stock.request_list'))
    req.status = 'rejected'
    req.approver = current_user.display_name or current_user.username
    db.session.commit()
    flash(f'已拒绝 {req.part.name} x{req.quantity} 的领用申请', 'info')
    return redirect(url_for('stock.request_list'))


@stock_bp.route('/out-records/<int:rid>/delete', methods=['POST'])
@login_required
@admin_required
def delete_out_record(rid):
    """删除出库记录"""
    from models import StockRecord
    rec = StockRecord.query.get_or_404(rid)
    db.session.delete(rec)
    db.session.commit()
    flash('已删除出库记录', 'success')
    return redirect(url_for('stock.out_records'))
