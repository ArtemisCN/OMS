"""工单管理路由（薄路由版 —— 业务逻辑委托给 services/order_service）"""

from utils.csrf import csrf_protect

from utils.helpers import safe_get, safe_get_or_404
import io
import json
import os
import uuid
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from openpyxl import Workbook

from models import db, WorkOrder, SystemSetting, WorkOrderChatMessage, WorkOrderPhoto, can_access, User, ShiftHandover
from services import order_service as svc
from services.fault_matcher import match_fault
from routes.auth import admin_required
from utils.permissions import permission_required, has_permission
from utils.time_helpers import fmt_dt, now, fmt_date, resolve_team

orders_bp = Blueprint('orders', __name__, url_prefix='/orders')


# ==================== 工单列表 ====================

@orders_bp.route('/')
@permission_required('order:view')
def list_orders():
    page = request.args.get('page', 1, type=int)
    per_page = 20

    # 提取筛选条件
    filters = {
        'status': request.args.get('status', 'pending'),
        'fault_type': request.args.get('fault_type', ''),
        'person': request.args.get('person', ''),
        'keyword': request.args.get('keyword', ''),
        'date_from': request.args.get('date_from', ''),
        'date_to': request.args.get('date_to', ''),
        'building': request.args.get('building', ''),
        'department': request.args.get('department', ''),
        'floor': request.args.get('floor', ''),
        'location': request.args.get('location', ''),
        'team': resolve_team(request, current_user),
    }

    sort = request.args.get('sort', '')
    order = request.args.get('order', '')
    filters['sort'] = sort
    filters['order'] = order

    query = svc.build_order_query(filters, current_user)
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    persons, buildings, teams = svc.get_filter_data()
    stats = svc.get_order_stats()

    from datetime import datetime
    now = datetime.now()
    ball_map = {}
    for o in pagination.items:
        end_t = o.end_time or o.completed_at
        start_t = o.start_time or o.accepted_at or o.created_at
        if o.status == 'completed' and start_t and end_t:
            duration = (end_t - start_t).total_seconds() / 60
            if duration < 30:
                ball_map[o.id] = 'normal'
            elif duration < 60:
                ball_map[o.id] = 'urgent'
            else:
                ball_map[o.id] = 'emergency'
        elif o.status == 'pending' and o.created_at:
            # 未接单：按等待时间自动升级紧急程度（只升不降）
            wait_hours = (now - o.created_at).total_seconds() / 3600
            base_priority = o.priority or 'normal'
            priority_level = {'normal': 0, 'urgent': 1, 'emergency': 2}
            display_level = priority_level.get(base_priority, 0)
            if wait_hours > 8:
                display_level = max(display_level, 2)
            elif wait_hours > 4:
                display_level = max(display_level, 1)
            ball_map[o.id] = ['normal', 'urgent', 'emergency'][display_level]
        else:
            ball_map[o.id] = o.priority or 'normal'

    return render_template('orders/list.html',
                           pagination=pagination, orders=pagination.items,
                           persons=persons, buildings=buildings,
                           status=filters['status'], stats=stats,
                           fault_type=filters['fault_type'],
                           person_sel=filters['person'],
                           keyword=filters['keyword'],
                           date_from=filters['date_from'],
                           date_to=filters['date_to'],
                           building=filters['building'],
                           department=filters['department'],
                           floor_sel=filters['floor'],
                           location_sel=filters['location'],
                           team_sel=filters['team'], teams=teams,
                           ball_map=ball_map, sort=sort, order=order,
                           starred_ids=current_user.get_pref('starred_ids', []))


# ==================== 新建工单 ====================

@orders_bp.route('/create', methods=['GET', 'POST'])
@csrf_protect
@permission_required('order:create')
def create_order():
    """新建工单（Web页面）
    ---
    tags:
      - Web 页面
    summary: 新建工单
    x-web-page: true
    responses:
      200:
        description: 工单创建页面（GET）或创建成功重定向（POST）
    """
    if not can_access('work_order'):
        return "无权访问", 403
    from services.address import get_merged_addresses, get_all_buildings
    team = resolve_team(request, current_user)

    if request.method == 'POST':
        try:
            order = svc.create_order(
                request.form,
                current_user.display_name or current_user.username,
            )
            flash('工单创建成功', 'success')
            return redirect(url_for('orders.list_orders'))
        except ValueError as e:
            flash(str(e), 'danger')
            from services.data_service import get_team_options
            from services.order_service import get_create_page_data
            persons, templates = get_create_page_data(current_user)
            all_teams = get_team_options()
            return render_template('orders/create.html',
                                   addr_list=get_merged_addresses(team=team),
                                   persons=persons, templates=templates,
                                   buildings=get_all_buildings(team=team),
                                   team=team, all_teams=all_teams)

    persons, templates = svc.get_create_page_data(current_user)
    from services.data_service import get_team_options
    all_teams = get_team_options()
    return render_template('orders/create.html',
                           addr_list=get_merged_addresses(team=team),
                           persons=persons, templates=templates,
                           buildings=get_all_buildings(team=team),
                           team=team, all_teams=all_teams)


# ==================== 发布工单（匿名/登录通用） ====================

@orders_bp.route('/publish', methods=['GET', 'POST'])
@csrf_protect
@permission_required('order:create')
def publish_order():
    if request.method == 'POST':
        try:
            order = svc.publish_order(
                request.form,
                current_user.display_name or current_user.username,
            )
            # 推送通知
            try:
                from routes.api_mobile import send_new_order_notification, send_wecom_notification
                send_new_order_notification(order)
                send_wecom_notification(order)
            except Exception:
                pass
            flash('✅ 工单已发布，等待手机端接单', 'success')
            return redirect(url_for('orders.list_orders'))
        except ValueError as e:
            flash(str(e), 'danger')

    from services.data_service import get_team_options
    all_teams = get_team_options()
    team = resolve_team(request, current_user)

    # ====== 交接班数据 ======
    handovers = ShiftHandover.query.order_by(ShiftHandover.created_at.desc()).limit(20).all()
    pending_orders = WorkOrder.query.filter(
        WorkOrder.status.in_(['pending', 'in_progress'])
    ).order_by(WorkOrder.created_at.desc()).all()
    persons = User.query.filter(User.is_active == True).order_by(User.display_name).all()
    handover_person_set = set()
    for h in handovers:
        if h.handover_person: handover_person_set.add(h.handover_person)
        if h.receive_person: handover_person_set.add(h.receive_person)
    handover_records_data = [{
        'id': h.id,
        'handover_person': h.handover_person,
        'receive_person': h.receive_person,
        'content': h.content,
        'unfinished_orders': h.unfinished_orders or [],
        'notes': h.notes,
        'status': h.status,
        'created_at': h.created_at.strftime('%Y-%m-%d %H:%M') if h.created_at else '',
    } for h in handovers]
    return render_template('orders/publish.html',
        all_teams=all_teams, team=team,
        handover_records=handovers,
        handover_records_data=handover_records_data,
        handover_stats={
            'total': len(handovers),
            'today': sum(1 for h in handovers if h.created_at and h.created_at.date() == datetime.now().date()),
            'pending_orders': len(pending_orders),
            'person_count': len(handover_person_set),
        },
        handover_persons=persons,
        handover_pending_orders=pending_orders,
    )


# ==================== API 辅助 ====================

@orders_bp.route('/api/guess')
@login_required
def api_guess_fault():
    return jsonify(svc.api_guess_fault(request.args.get('title', ''), team=request.args.get('team', '')))


@orders_bp.route('/api/solution_suggest')
@login_required
def api_solution_suggest():
    return jsonify(svc.api_solution_suggest(
        request.args.get('q', '').strip(), current_user
    ))


@orders_bp.route('/api/address/all')
@login_required
def api_address_all():
    return jsonify(svc.api_address_all(team=request.args.get('team', '')))


@orders_bp.route('/api/address/options')
@login_required
def api_address_options():
    return jsonify(svc.api_address_options(
        request.args.get('building', ''),
        request.args.get('floor', ''),
        team=request.args.get('team', ''),
    ))


# ==================== 批量生成 ====================

@orders_bp.route('/batch', methods=['GET', 'POST'])
@csrf_protect
@permission_required('order:batch')
def batch_create():
    selected_team = request.args.get('team', 'all')
    persons, templates, fault_groups, fault_group_items, team_groups, teams, default_team = \
        svc.get_batch_form_data(current_user, selected_team=selected_team)

    if request.method == 'POST' and request.form.get('action') == 'preview':
        try:
            serialized, by_date, sorted_dates, total = svc.batch_preview(
                request.form, current_user
            )
            preview_json_obj = json.dumps(serialized, ensure_ascii=False)
            return render_template('orders/batch.html',
                                   persons=persons, team_groups=team_groups,
                                   teams=teams, default_team=default_team,
                                   selected_team=selected_team,
                                   preview_json=preview_json_obj,
                                   preview_orders=serialized,
                                   preview_total=total,
                                   templates=templates,
                                   fault_groups=fault_groups,
                                   fault_group_items=fault_group_items,
                                   by_date=dict(by_date),
                                   sorted_dates=sorted_dates,
                                   year=int(request.form.get('year', datetime.now().year)),
                                   month=int(request.form.get('month', datetime.now().month)),
                                   min_per_day=int(request.form.get('min_per_day', 20)),
                                   max_per_day=int(request.form.get('max_per_day', 45)),
                                   everyday=request.form.get('everyday') == 'on',
                                   selected_names=request.form.getlist('selected_names'),
                                   dates_str=request.form.get('specific_dates', '').strip())
        except ValueError as e:
            flash(str(e), 'danger')
            return render_template('orders/batch.html', persons=persons,
                                   team_groups=team_groups, teams=teams,
                                   default_team=default_team,
                                   selected_team=selected_team,
                                   templates=templates, fault_groups=fault_groups,
                                   fault_group_items=fault_group_items)
        except Exception as e:
            flash(f'生成失败：{str(e)}', 'danger')
            return render_template('orders/batch.html', persons=persons,
                                   team_groups=team_groups, teams=teams,
                                   default_team=default_team,
                                   selected_team=selected_team,
                                   templates=templates, fault_groups=fault_groups,
                                   fault_group_items=fault_group_items)

    if request.method == 'POST' and request.form.get('action') == 'confirm':
        try:
            total, batch_ids = svc.batch_confirm(
                request.form.get('preview_json', ''), current_user
            )
            session['last_batch_time'] = datetime.now().isoformat()
            session['last_batch_count'] = total
            session['last_batch_ids'] = batch_ids
            flash(f'批量生成成功！共 {total} 条工单已保存到当月工单', 'success')
            return redirect(url_for('orders.list_orders'))
        except ValueError as e:
            flash(str(e), 'danger')
        except Exception as e:
            flash(f'保存失败：{str(e)}', 'danger')

    # GET：检查可反悔批次
    can_undo = False
    undo_count = 0
    last_batch_time = session.get('last_batch_time')
    if last_batch_time and session.get('last_batch_ids'):
        try:
            bt = datetime.fromisoformat(last_batch_time)
            if (datetime.now() - bt).total_seconds() < 300:
                can_undo = True
                undo_count = session.get('last_batch_count', 0)
            else:
                session.pop('last_batch_time', None)
                session.pop('last_batch_ids', None)
                session.pop('last_batch_count', None)
        except Exception:
            pass

    return render_template('orders/batch.html', persons=persons,
                           team_groups=team_groups, teams=teams,
                           default_team=default_team,
                           selected_team=selected_team,
                           can_undo=can_undo, undo_count=undo_count,
                           templates=templates, fault_groups=fault_groups,
                           fault_group_items=fault_group_items)


@orders_bp.route('/batch/undo', methods=['POST'])
@csrf_protect
@login_required
def batch_undo():
    ids = session.get('last_batch_ids', [])
    if not ids:
        flash('没有可撤回的批次，或已超时（限5分钟内）', 'warning')
        return redirect(url_for('orders.batch_create'))
    try:
        deleted = svc.batch_undo(ids)
        session.pop('last_batch_time', None)
        session.pop('last_batch_ids', None)
        session.pop('last_batch_count', None)
        flash(f'已撤销最近一次批量生成的 {deleted} 条工单', 'success')
    except Exception as e:
        flash(f'撤销失败：{str(e)}', 'danger')
    return redirect(url_for('orders.list_orders'))


# ==================== 详情 / 编辑 / 删除 ====================

@orders_bp.route('/<int:order_id>')
@permission_required('order:view')
def detail(order_id):
    order = svc.get_order_or_404(order_id)
    photos = svc.get_order_photos(order_id)
    from models import SparePart, StockRecord
    spare_parts = SparePart.query.order_by(SparePart.name).all()
    linked_parts = StockRecord.query.filter_by(work_order_id=order_id, type='out').order_by(StockRecord.created_at.desc()).all()
    return render_template('orders/detail.html', order=order, photos=photos,
                           spare_parts=spare_parts, linked_parts=linked_parts)


@orders_bp.route('/<int:order_id>/solution', methods=['POST'])
@csrf_protect
@permission_required('order:solve')
def update_solution(order_id):
    """更新工单解决方案（AJAX）"""
    order = safe_get_or_404(WorkOrder, order_id)
    solution = request.form.get('solution', '').strip()
    order.solution = solution
    if solution and order.status == 'in_progress':
        order.status = 'completed'
        order.completed_at = datetime.now()
        order.end_time = datetime.now()
    db.session.commit()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True, 'solution': solution, 'status': order.status})
    flash('✅ 解决方案已更新', 'success')
    return redirect(url_for('orders.detail', order_id=order.id))


@orders_bp.route('/<int:order_id>/edit', methods=['GET', 'POST'])
@csrf_protect
@permission_required('order:edit')
def edit_order(order_id):
    order = svc.get_order_or_404(order_id)
    if request.method == 'POST':
        svc.update_order(order_id, request.form)
        flash('✅ 工单已更新', 'success')
        return redirect(url_for('orders.detail', order_id=order.id))
    persons = User.query.filter(User.is_active == True).all()
    from services.data_service import get_team_options
    all_teams = get_team_options()
    return render_template('orders/edit.html', order=order, persons=persons, all_teams=all_teams)


@orders_bp.route('/<int:order_id>/delete', methods=['POST'])
@csrf_protect
@permission_required('order:delete')
def delete_order(order_id):
    svc.delete_order(order_id,
                     current_user.display_name or current_user.username)
    flash('工单已删除', 'success')
    return redirect(url_for('orders.list_orders'))


@orders_bp.route('/<int:order_id>/toggle_priority', methods=['POST'])
@csrf_protect
@login_required
def toggle_priority(order_id):
    try:
        priority = svc.toggle_priority(order_id)
        return jsonify({'priority': priority})
    except ValueError as e:
        order = svc.get_order_or_404(order_id)
        return jsonify({'error': str(e), 'priority': order.priority}), 403


# ==================== 导出 Excel ====================

@orders_bp.route('/export')
@login_required
def export_excel():
    filters = {
        'status': request.args.get('status', ''),
        'fault_type': request.args.get('fault_type', ''),
        'person': request.args.get('person', ''),
        'keyword': request.args.get('keyword', ''),
        'date_from': request.args.get('date_from', ''),
        'date_to': request.args.get('date_to', ''),
    }
    query = svc.build_order_query(filters)
    orders = query.all()
    total = len(orders)

    wb = Workbook()
    ws = wb.active
    ws.title = '工单导出'
    headers = ['编号', '工单名称', '设备类型', '故障类型', '描述',
               '楼区', '楼层', '科室', '位置', '处理人',
               '开始时间', '结束时间', '状态', '解决方案', '创建人', '创建时间']
    ws.append(headers)
    for o in orders:
        ws.append([
            o.id, o.title, o.device_type, o.fault_type, o.description,
            o.building, o.floor, o.department, o.location, o.person,
            fmt_dt(o.start_time, '%Y-%m-%d %H:%M'),
            fmt_dt(o.end_time, '%Y-%m-%d %H:%M'),
            {'pending': '待接单', 'in_progress': '处理中', 'completed': '已完成'}.get(o.status, o.status),
            o.solution,
            o.created_by,
            fmt_dt(o.created_at, '%Y-%m-%d %H:%M'),
        ])

    filename = f'工单导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ==================== 发布页面（无登录报修） ====================


def _publish_render(anonymous=False):
    """渲染 publish.html 的统一入口：补全交接班上下文（匿名/登录通用）"""
    from services.data_service import get_team_options
    from utils.time_helpers import resolve_team
    all_teams = get_team_options()
    try:
        team = resolve_team(request, current_user) if current_user.is_authenticated else ''
    except Exception:
        team = ''
    handovers = ShiftHandover.query.order_by(ShiftHandover.created_at.desc()).limit(20).all()
    pending_orders = WorkOrder.query.filter(
        WorkOrder.status.in_(['pending', 'in_progress'])
    ).order_by(WorkOrder.created_at.desc()).all()
    persons = User.query.filter(User.is_active == True).order_by(User.display_name).all()
    handover_person_set = set()
    for h in handovers:
        if h.handover_person: handover_person_set.add(h.handover_person)
        if h.receive_person: handover_person_set.add(h.receive_person)
    handover_records_data = [{
        'id': h.id,
        'handover_person': h.handover_person,
        'receive_person': h.receive_person,
        'content': h.content,
        'unfinished_orders': h.unfinished_orders or [],
        'notes': h.notes,
        'status': h.status,
        'created_at': h.created_at.strftime('%Y-%m-%d %H:%M') if h.created_at else '',
    } for h in handovers]
    handover_stats = {
        'total': len(handovers),
        'today': sum(1 for h in handovers if h.created_at and h.created_at.date() == datetime.now().date()),
        'pending_orders': len(pending_orders),
        'person_count': len(handover_person_set),
    }
    return render_template('orders/publish.html',
        all_teams=all_teams, team=team,
        handover_records=handovers,
        handover_records_data=handover_records_data,
        handover_stats=handover_stats,
        handover_persons=persons,
        handover_pending_orders=pending_orders,
        anonymous=anonymous)


@orders_bp.route('/anonymous_publish', methods=['GET', 'POST'])
@csrf_protect
def anonymous_publish():
    """匿名报修（无需登录）"""
    verification = session.get('publish_verified', False)

    if request.method == 'POST':
        if not verification:
            code = request.form.get('verify_code', '').strip()
            # 验证码从系统设置读取（P1-6，默认 4567）
            _an_setting = SystemSetting.query.filter_by(key='anonymous_code').first()
            expected_code = (_an_setting.value if _an_setting and _an_setting.value else '4567')
            if code != expected_code:
                flash('验证码错误', 'danger')
                return _publish_render(anonymous=True)
            session['publish_verified'] = True
            verification = True

        # 走正常的发布逻辑，但创建人为"匿名"
        from services.address import extract_address_from_title
        title = request.form.get('title', '').strip()
        if not title:
            flash('请输入故障描述', 'danger')
            return _publish_render(anonymous=True)

        from services.keyword_config import get_fault_keywords, get_device_keywords
        fk = get_fault_keywords()
        dk = get_device_keywords()
        fault, device = svc._guess_fault_type(title, fk, dk)
        fm = match_fault(title)
        auto_fault = fm['category'] if fm['match_type'] == 'keyword' else fault
        addr = extract_address_from_title(title)

        building = request.form.get('building', '') or addr['building']
        floor = request.form.get('floor', '') or addr['floor']
        department = request.form.get('department', '') or addr['department']
        location = request.form.get('location', '') or addr.get('location', '')

        order = WorkOrder(
            title=title,
            device_type=device,
            fault_type=auto_fault,
            fault_subcategory=fm.get('subcategory', ''),
            description=request.form.get('description', ''),
            building=building, floor=floor,
            department=department, location=location,
            person='', solution='',
            start_time=datetime.now(),
            status='pending',
            created_by='匿名',
            priority='normal', original_priority='normal',
            hospital_id=request.form.get('hospital_id', 1, type=int),
        )
        db.session.add(order)
        db.session.commit()
        flash('✅ 报修已提交，请等待工程师联系', 'success')
        return redirect(url_for('orders.anonymous_publish'))

    return _publish_render(anonymous=True)


# ==================== Excel 导入工单 ====================

@orders_bp.route('/import/template')
@login_required
def download_import_template():
    """下载工单导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工单导入模板"
    headers = ['工单名称*', '设备类型', '故障类型', '故障描述',
               '楼区', '楼层', '科室', '位置',
               '经办人', '解决方案', '开始时间', '结束时间', '状态', '紧急程度']
    ws.append(headers)
    ws.append(['示例：电脑无法开机', '电脑', '硬件', '开机黑屏',
               '1号楼', '3层', '信息科', '301室',
               '张三', '更换电源线', '2026-07-01 09:00', '2026-07-01 09:30', 'completed', 'normal'])
    for col, w in zip('ABCDEFGHIJKLMN', [30, 10, 10, 20, 10, 8, 12, 12, 10, 20, 16, 16, 12, 10]):
        ws.column_dimensions[col].width = w
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='工单导入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@orders_bp.route('/import', methods=['GET', 'POST'])
@csrf_protect
@login_required
def import_orders():
    """批量导入工单（Excel）"""
    from openpyxl import load_workbook

    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('请选择 Excel 文件', 'danger')
            return render_template('orders/import.html')
        try:
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            if not rows:
                flash('Excel 文件为空', 'danger')
                return render_template('orders/import.html')
            imported = 0
            errors = []
            now = datetime.now()
            for idx, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                title = str(row[0] or '').strip()
                if not title:
                    errors.append(f'第{idx}行：缺少工单名称')
                    continue
                try:
                    order = WorkOrder(
                        title=title, device_type=str(row[1] or '其他').strip()[:50],
                        fault_type=str(row[2] or '硬件').strip()[:50],
                        description=str(row[3] or '').strip(),
                        building=str(row[4] or '').strip()[:50],
                        floor=str(row[5] or '').strip()[:20],
                        department=str(row[6] or '').strip()[:100],
                        location=str(row[7] or '').strip()[:200],
                        person=str(row[8] or '').strip()[:50],
                        solution=str(row[9] or '').strip(),
                        status='pending', priority='normal',
                        created_by=current_user.display_name or current_user.username,
                    )
                    if row[10]:
                        try:
                            order.start_time = row[10] if isinstance(row[10], datetime) else datetime.strptime(str(row[10]).strip(), '%Y-%m-%d %H:%M')
                            order.created_at = order.start_time
                        except Exception:
                            pass
                    if row[11]:
                        try:
                            order.end_time = row[11] if isinstance(row[11], datetime) else datetime.strptime(str(row[11]).strip(), '%Y-%m-%d %H:%M')
                        except Exception:
                            pass
                    sv = str(row[12] or '').strip().lower()
                    if sv in ('pending', 'in_progress', 'completed'):
                        order.status = sv
                        if sv == 'completed' and not order.end_time:
                            order.completed_at = order.end_time or now
                        elif sv == 'in_progress':
                            order.accepted_at = now
                    pri = str(row[13] or '').strip().lower()
                    if pri in ('normal', 'urgent', 'emergency'):
                        order.priority = pri
                        order.original_priority = pri
                    if order.person and not order.accepted_at:
                        order.accepted_at = order.created_at or now
                    db.session.add(order)
                    imported += 1
                except Exception as e:
                    errors.append(f'第{idx}行：{str(e)}')
            db.session.commit()
            msg = f'✅ 成功导入 {imported} 条工单'
            if errors:
                msg += f'，{len(errors)} 条错误（见下方）'
            flash(msg, 'success' if not errors else 'warning')
            return render_template('orders/import.html', imported=imported, errors=errors[:50])
        except Exception as e:
            flash(f'导入失败：{str(e)}', 'danger')
            return render_template('orders/import.html')
    return render_template('orders/import.html')


# ==================== 工单日历 ====================

@orders_bp.route('/calendar')
@login_required
def calendar_view():
    """工单日历视图"""
    from calendar import monthrange
    from collections import defaultdict
    now = datetime.now()
    year = request.args.get('year', now.year, type=int)
    month = request.args.get('month', now.month, type=int)

    month_start = datetime(year, month, 1)
    _, days_in_month = monthrange(year, month)
    month_end = datetime(year, month, days_in_month, 23, 59, 59)

    # 组别筛选（与仪表盘同口径：管理员默认跟随 default_dashboard_team，非管理员固定本人组）
    # request.args.get('team')：None=URL无team参数(首次进入) / ''=选了全部组
    team_param = request.args.get('team')
    if team_param is None:
        if has_permission(current_user, 'order:batch'):
            _def_setting = SystemSetting.query.filter_by(key='default_dashboard_team').first()
            team = _def_setting.value if _def_setting and _def_setting.value else ''
        else:
            _person = User.query.filter_by(id=current_user.id).first()
            team = _person.team if _person and _person.team else ''
    else:
        team = team_param
    team_persons = set()
    if team:
        tp = User.query.filter(User.team == team, User.is_active == True).all()
        team_persons = {p.display_name for p in tp if p.display_name}
    teams = [t[0] for t in User.query.with_entities(User.team).filter(User.team!='', User.team!=None).distinct().order_by(User.team).all()]

    orders = WorkOrder.query.filter(
        WorkOrder.created_at >= month_start,
        WorkOrder.created_at <= month_end,
    )
    if team_persons:
        orders = orders.filter(WorkOrder.person.in_(team_persons))
    orders = orders.order_by(WorkOrder.created_at).all()

    cal_data = defaultdict(list)
    for o in orders:
        cal_data[o.created_at.day].append(o)

    weekdays = ['一', '二', '三', '四', '五', '六', '日']
    first_weekday = month_start.weekday()

    prev_month, prev_year = (month - 1, year) if month > 1 else (12, year - 1)
    next_month, next_year = (month + 1, year) if month < 12 else (1, year + 1)

    return render_template('orders/calendar.html',
                           year=year, month=month,
                           days_in_month=days_in_month,
                           first_weekday=first_weekday,
                           weekdays=weekdays,
                           cal_data=dict(cal_data),
                           prev_month=prev_month, prev_year=prev_year,
                           next_month=next_month, next_year=next_year,
                           teams=teams, team_sel=team,
                           now=now)


@orders_bp.route('/api/calendar-day')
@login_required
def calendar_day_api():
    """返回某一天的工单列表（日历弹窗JSON API）"""
    date_str = request.args.get('date', '')
    if not date_str:
        return jsonify([])
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        next_dt = dt + timedelta(days=1)
    except ValueError:
        return jsonify([])
    orders = WorkOrder.query.filter(
        WorkOrder.created_at >= dt,
        WorkOrder.created_at < next_dt,
    )
    if not has_permission(current_user, 'order:batch'):
        _person = User.query.filter_by(id=current_user.id).first()
        if _person and _person.team:
            tp = User.query.filter(
                User.team == _person.team, User.is_active == True
            ).all()
            team_names = {p.display_name for p in tp if p.display_name}
            if team_names:
                orders = orders.filter(WorkOrder.person.in_(team_names))
    orders = orders.order_by(WorkOrder.created_at).all()
    return jsonify([{
        'id': o.id,
        'title': o.title,
        'status': o.status,
        'priority': o.priority,
        'person': o.person,
        'start_time': fmt_dt(o.start_time, '%H:%M'),
        'detail_url': url_for('orders.detail', order_id=o.id),
    } for o in orders])


# ==================== 工单照片上传/删除 ====================

@orders_bp.route('/<int:order_id>/photos/upload', methods=['POST'])
@csrf_protect
@permission_required('order:edit')
def upload_photo(order_id):
    """上传工单照片"""
    order = safe_get_or_404(WorkOrder, order_id)
    files = request.files.getlist('photos')
    if not files:
        flash('请选择图片', 'warning')
        return redirect(url_for('orders.detail', order_id=order.id))

    from utils.photo import save_photo, allowed_file
    person_name = current_user.display_name or current_user.username
    count = 0
    for f in files:
        if not f.filename or not allowed_file(f.filename):
            continue
        try:
            file_data = f.read()
            rel_path, w, h, size = save_photo(file_data, f.filename)
            photo = WorkOrderPhoto(
                work_order_id=order.id, filename=f.filename,
                filepath=rel_path, file_size=size, width=w, height=h,
                uploaded_by=person_name,
            )
            db.session.add(photo)
            count += 1
        except Exception:
            pass

    db.session.commit()
    flash(f'✅ 上传成功 {count} 张图片', 'success')
    return redirect(url_for('orders.detail', order_id=order.id))


@orders_bp.route('/<int:order_id>/photos/<int:photo_id>/delete', methods=['POST'])
@csrf_protect
@permission_required('order:delete')
def delete_photo(order_id, photo_id):
    """删除工单照片"""
    from models import WorkOrderPhoto
    from utils.photo import delete_photo_file
    photo = safe_get(WorkOrderPhoto, photo_id)
    if photo and photo.work_order_id == int(order_id):
        if photo.filepath:
            delete_photo_file(photo.filepath)
        db.session.delete(photo)
        db.session.commit()
    flash('✅ 已删除', 'success')
    return redirect(url_for('orders.detail', order_id=order_id))


# ==================== 工单搜索 API ====================

@orders_bp.route('/api/search')
@login_required
def search_orders_api():
    """JSON API: 搜索工单（供知识库引用）"""
    q = request.args.get('q', '').strip()
    limit = min(int(request.args.get('limit', 10)), 20)
    if not q:
        return jsonify([])
    orders = WorkOrder.query.filter(
        db.or_(
            WorkOrder.title.ilike(f'%{q}%'),
            WorkOrder.id.ilike(f'%{q}%'),
        )
    ).order_by(WorkOrder.id.desc()).limit(limit).all()
    return jsonify([{
        'id': o.id,
        'title': o.title,
        'detail_url': url_for('orders.detail', order_id=o.id),
        'person': o.person or '',
        'building': o.building or '',
        'floor': o.floor or '',
        'status': o.status,
    } for o in orders])


# ==================== 工单讨论（聊天） ====================

WO_CHAT_UPLOAD_FOLDER = '/var/www/static/uploads/wochat'
os.makedirs(WO_CHAT_UPLOAD_FOLDER, exist_ok=True)
ALLOWED_WO_CHAT_EXT = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}


@orders_bp.route('/<int:order_id>/chat')
@login_required
def wo_chat_messages(order_id):
    """获取工单聊天消息"""
    order = safe_get_or_404(WorkOrder, order_id)
    before_id = request.args.get('before_id', type=int)
    after_id = request.args.get('after_id', type=int)
    limit = min(request.args.get('limit', 50, type=int), 100)
    query = WorkOrderChatMessage.query.filter_by(work_order_id=order.id)
    if after_id:
        query = query.filter(WorkOrderChatMessage.id > after_id)
        msgs = query.order_by(WorkOrderChatMessage.id.asc()).limit(limit).all()
    else:
        if before_id:
            query = query.filter(WorkOrderChatMessage.id < before_id)
        msgs = query.order_by(WorkOrderChatMessage.id.desc()).limit(limit).all()
        msgs.reverse()
    return jsonify([{
        'id': m.id,
        'sender_id': m.sender_id,
        'sender_name': m.sender_name,
        'content': m.content,
        'msg_type': m.msg_type,
        'created_at': m.created_at.isoformat() if m.created_at else '',
        'is_self': m.sender_id == current_user.id
    } for m in msgs])


@orders_bp.route('/<int:order_id>/chat/send', methods=['POST'])
@csrf_protect
@login_required
def wo_chat_send(order_id):
    """发送工单聊天消息"""
    order = safe_get_or_404(WorkOrder, order_id)
    data = request.get_json()
    if not data:
        return jsonify({'error': '无效请求'}), 400
    content = (data.get('content') or '').strip()
    if not content:
        return jsonify({'error': '不能发送空消息'}), 400
    msg = WorkOrderChatMessage(
        work_order_id=order.id,
        sender_id=current_user.id,
        sender_name=current_user.display_name or current_user.username,
        content=content,
        msg_type=data.get('msg_type', 'text')
    )
    db.session.add(msg)
    db.session.commit()
    return jsonify({
        'id': msg.id,
        'sender_id': msg.sender_id,
        'sender_name': msg.sender_name,
        'content': msg.content,
        'msg_type': msg.msg_type,
        'created_at': msg.created_at.isoformat() if msg.created_at else '',
        'is_self': True
    }), 201


@orders_bp.route('/<int:order_id>/chat/upload', methods=['POST'])
@csrf_protect
@login_required
def wo_chat_upload(order_id):
    """上传工单聊天文件（图片）"""
    order = safe_get_or_404(WorkOrder, order_id)
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400
    file = request.files['file']
    if file.filename == '' or not file:
        return jsonify({'error': '文件为空'}), 400
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_WO_CHAT_EXT:
        return jsonify({'error': '不支持的文件类型，仅支持图片'}), 400
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(WO_CHAT_UPLOAD_FOLDER, filename)
    file.save(filepath)
    return jsonify({
        'url': f'/static/uploads/wochat/{filename}',
        'filename': filename,
        'is_image': True
    })


# ==================== 星标切换 ====================

@orders_bp.route('/<int:order_id>/star', methods=['POST'])
@csrf_protect
@login_required
def toggle_star(order_id):
    """切换工单星标状态"""
    starred = current_user.get_pref('starred_ids', [])
    if order_id in starred:
        starred.remove(order_id)
        is_starred = False
    else:
        starred.append(order_id)
        is_starred = True
    current_user.set_pref('starred_ids', starred)
    db.session.commit()
    return jsonify({'starred': is_starred})

@orders_bp.route('/<int:order_id>/urge', methods=['POST'])
@csrf_protect
@login_required
def urge_order(order_id):
    """催单：工单紧急程度上升为紧急 + 发送企业微信通知"""
    order = safe_get_or_404(WorkOrder, order_id)
    
    # 检查催单间隔
    interval_setting = SystemSetting.query.filter_by(key='order_remind_interval').first()
    min_interval = int(interval_setting.value) if interval_setting and interval_setting.value else 30
    
    if order.last_urged_at:
        elapsed = (now() - order.last_urged_at).total_seconds() / 60
        if elapsed < min_interval:
            remain = int(min_interval - elapsed)
            return jsonify({'success': False, 'error': f'距上次催单仅{int(elapsed)}分钟，请{remain}分钟后再试'})
    
    # 升级紧急程度为「紧急」
    prev_priority = order.priority
    was_escalated = False
    if order.priority != 'emergency':
        order.priority = 'emergency'
        note = f'\n[催办升级] {now().strftime("%Y-%m-%d %H:%M")} 由 {prev_priority} 提升为 emergency（催办人：{current_user.username}）'
        order.description = (order.description or '') + note
        was_escalated = True
    
    from routes.api_mobile import send_wecom_notification
    send_wecom_notification(order, skip_time_check=True, is_urge=True)
    
    order.last_urged_at = now()
    db.session.commit()
    
    msg = '🚨 已升级为紧急 · 催办通知已发送 ✅' if was_escalated else '催办通知已发送 ✅'
    return jsonify({'success': True, 'message': msg})
