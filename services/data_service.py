"""
数据管理服务层 —— 所有基础数据 CRUD 业务逻辑收拢于此

职责：
  人员、科室、方案模板、地址、故障类型、存放位置、供应商、
  耗材、值班排班、知识库、权限、故障分类 等 CRUD
路由层（data.py）只负责 HTTP 请求/响应编排
"""
import io
import json
from datetime import datetime, date, timedelta
from calendar import monthrange

from models import (
    db, Person, Department, SolutionTemplate, WorkOrder, AddressOverride,
    User, FaultType, FaultCategory, FaultSubcategory, FaultKeyword,
    StorageLocation, Supplier, Consumable, ConsumableRecord,
    DutySchedule, DutyStaff, KnowledgeBase, Hospital,
    FaultTemplateGroup, FaultTemplateItem, PartPrice,
    SystemSetting, log_audit, get_module_permissions, save_module_permissions
)


# ===================================================================
#  1. 人员管理
# ===================================================================

def list_persons():
    """获取人员列表及对应的账号映射"""
    persons = Person.query.order_by(Person.is_active.desc(), Person.id).all()
    user_map = {}
    for p in persons:
        user = User.query.filter(
            (User.display_name == p.name) | (User.username == p.name)
        ).first()
        if user:
            user_map[p.name] = user
    return persons, user_map


def add_person(name):
    """新增人员，返回 (ok, msg)"""
    if not name:
        return False, '人员姓名不能为空'
    if Person.query.filter_by(name=name).first():
        return False, f'人员 "{name}" 已存在'
    db.session.add(Person(name=name))
    db.session.commit()
    return True, f'已添加人员 "{name}"'


def import_persons_from_orders():
    """从工单中导入人员"""
    persons_in_orders = db.session.query(WorkOrder.person).distinct().all()
    imported = 0
    for (name,) in persons_in_orders:
        if name and name.strip():
            name = name.strip()
            if not Person.query.filter_by(name=name).first():
                db.session.add(Person(name=name, is_active=True))
                imported += 1
    db.session.commit()
    return imported


def toggle_person(pid):
    """切换人员启用/停用"""
    p = Person.query.get_or_404(pid)
    p.is_active = not p.is_active
    db.session.commit()
    return p


def delete_person(pid, operator_name):
    """删除人员"""
    p = Person.query.get_or_404(pid)
    name = p.name
    db.session.delete(p)
    db.session.commit()
    log_audit('delete', 'person', operator_name,
              target_id=pid, target_desc=f'删除人员: {name}')
    return name


def edit_person_field(pid, field, value):
    """编辑人员字段（电话/组别/备注）"""
    p = Person.query.get_or_404(pid)
    if field == 'phone':
        p.phone = value
    elif field == 'team':
        p.team = value
    elif field == 'notes':
        p.notes = value
    else:
        return False, '未知字段'
    db.session.commit()
    return True, None


def person_account_info(pid):
    """获取人员账号信息（JSON）"""
    p = Person.query.get_or_404(pid)
    user = User.query.filter(
        (User.display_name == p.name) | (User.username == p.name)
    ).first()
    return {
        'has_account': bool(user),
        'username': user.username if user else '',
        'display_name': user.display_name if user else p.name,
        'is_admin': user.is_admin if user else False,
        'hospital_ids': [h.id for h in user.hospitals.all()] if user else [],
    }


def person_account_save(pid, username, password, display_name, is_admin, hospital_ids=None):
    """创建或更新人员的登录账号"""
    from models import Hospital
    from flask import g
    p = Person.query.get_or_404(pid)
    if not username:
        return False, '用户名不能为空'

    user = User.query.filter(
        (User.display_name == p.name) | (User.username == p.name)
    ).first()

    if user:
        old_username = user.username
        user.username = username
        user.display_name = display_name or p.name
        user.is_admin = is_admin
        if password:
            user.set_password(password)
        if username != old_username:
            conflict = User.query.filter(User.username == username, User.id != user.id).first()
            if conflict:
                return False, f'用户名 "{username}" 已被使用'
        # 保存所属医院（多选）
        if hospital_ids is not None:
            user.hospitals = [db.session.get(Hospital, hid) for hid in hospital_ids if hid]
        db.session.commit()
        return True, f'✅ 账号 "{old_username}" 已更新'
    else:
        conflict = User.query.filter_by(username=username).first()
        if conflict:
            return False, f'用户名 "{username}" 已被使用'
        if not password:
            return False, '新建账号必须设置密码'
        user = User(
            username=username, display_name=display_name or p.name,
            is_admin=is_admin,
        )
        user.set_password(password)
        # 保存所属医院（多选）
        if hospital_ids:
            user.hospitals = [db.session.get(Hospital, hid) for hid in hospital_ids if hid]
        db.session.add(user)
        db.session.commit()
        return True, f'✅ 账号 "{username}" 已创建'


# ===================================================================
#  2. 科室字典
# ===================================================================

def list_departments():
    """获取科室列表"""
    return Department.query.order_by(Department.sort_order, Department.name).all()


def add_department(name, building, floor, phone, operator_name):
    """新增科室"""
    if not name:
        return False, '科室名称不能为空'
    if Department.query.filter_by(name=name).first():
        return False, f'科室 "{name}" 已存在'
    dept = Department(name=name, building=building, floor=floor, phone=phone)
    db.session.add(dept)
    db.session.commit()
    log_audit('create', 'department', operator_name, target_desc=f'新增科室: {name}')
    return True, f'已添加科室 "{name}"'


def edit_department(dept_id, name, building, floor, phone):
    """编辑科室"""
    dept = db.session.get(Department, dept_id)
    if not dept:
        return False, '科室不存在'
    if name and name != dept.name:
        if Department.query.filter_by(name=name).first():
            return False, f'科室 "{name}" 已存在'
        dept.name = name
    dept.building = building
    dept.floor = floor
    dept.phone = phone
    db.session.commit()
    return True, f'已更新科室 "{dept.name}"'


def delete_department(dept_id, operator_name):
    """删除科室"""
    dept = db.session.get(Department, dept_id)
    if not dept:
        return False, '科室不存在'
    log_audit('delete', 'department', operator_name, target_desc=f'删除科室: {dept.name}')
    db.session.delete(dept)
    db.session.commit()
    return True, f'已删除科室 "{dept.name}"'


# ===================================================================
#  3. 方案模板
# ===================================================================

def list_solutions(keyword='', device_filter='', fault_filter='', page=1, per_page=20):
    """获取方案模板列表（分页）"""
    query = SolutionTemplate.query
    if keyword:
        like = f'%{keyword}%'
        query = query.filter(db.or_(
            SolutionTemplate.title.contains(keyword),
            SolutionTemplate.keywords.contains(keyword),
            SolutionTemplate.content.contains(keyword),
        ))
    if device_filter:
        query = query.filter(SolutionTemplate.device_type == device_filter)
    if fault_filter:
        query = query.filter(SolutionTemplate.fault_type == fault_filter)
    return query.order_by(SolutionTemplate.id).paginate(page=page, per_page=per_page, error_out=False)


def add_solution(title, content, keywords, device_type, fault_type, fault_subcategory):
    """新增方案模板"""
    if not title:
        return False, '方案标题不能为空'
    if SolutionTemplate.query.filter_by(title=title).first():
        return False, f'方案 "{title}" 已存在'
    s = SolutionTemplate(title=title, content=content or title,
                         keywords=keywords, device_type=device_type,
                         fault_type=fault_type, fault_subcategory=fault_subcategory)
    db.session.add(s)
    db.session.commit()
    return True, f'已添加方案 "{title}"'


def edit_solution(sid, field, value, value2=None):
    """编辑方案模板"""
    s = SolutionTemplate.query.get_or_404(sid)
    if field == 'title':
        s.title = value
    elif field == 'keywords':
        s.keywords = value
    elif field == 'device_type':
        s.device_type = value
    elif field == 'fault_type':
        s.fault_type = value
    elif field == 'fault_subcategory':
        s.fault_subcategory = value
    else:
        s.content = value2 or value
    db.session.commit()
    return True


def delete_solution(sid, operator_name):
    """删除方案模板"""
    s = SolutionTemplate.query.get_or_404(sid)
    title = s.title
    db.session.delete(s)
    db.session.commit()
    log_audit('delete', 'solution_template', operator_name,
              target_id=sid, target_desc=f'删除方案模板: {title}')
    return title


def reset_solutions():
    """从配置重置方案模板到默认值"""
    from config import SOLUTION_TEMPLATES
    count = 0
    for title, content in SOLUTION_TEMPLATES.items():
        existing = SolutionTemplate.query.filter_by(title=title).first()
        if existing:
            existing.content = content
        else:
            db.session.add(SolutionTemplate(title=title, content=content))
        count += 1
    db.session.commit()
    return count


def import_solutions_from_orders():
    """从工单中导入方案"""
    solutions = db.session.query(WorkOrder.solution).distinct().all()
    imported = 0
    for (content,) in solutions:
        if content and content.strip():
            content = content.strip()
            title = content[:30] + ('...' if len(content) > 30 else '')
            if not SolutionTemplate.query.filter_by(title=title).first():
                db.session.add(SolutionTemplate(title=title, content=content))
                imported += 1
    db.session.commit()
    return imported


# ===================================================================
#  4. 地址数据
# ===================================================================

def list_addresses(building='', keyword=''):
    """获取地址列表"""
    from services.address import get_addresses_grouped, get_all_buildings
    groups = get_addresses_grouped(building=building, keyword=keyword)
    buildings = get_all_buildings()
    all_addrs = []
    for baddr in groups.values():
        all_addrs.extend(baddr)
    current_addresses = groups.get(building, []) if building else []
    return groups, buildings, current_addresses, sum(len(v) for v in groups.values())


def edit_address(override_id, base_index, building, floor, department, location):
    """编辑地址"""
    if not all([building, floor, department, location]):
        return False, '所有字段不能为空'

    if override_id:
        o = AddressOverride.query.get(override_id)
        if o:
            o.building = building
            o.floor = floor
            o.department = department
            o.location = location
            db.session.commit()
            return True, '地址已更新'
        return False, '记录不存在'
    elif base_index is not None and base_index >= 0:
        from services.address import ADDRESS_LIST
        if base_index < len(ADDRESS_LIST):
            o = AddressOverride(base_index=base_index, building=building,
                                floor=floor, department=department, location=location)
            db.session.add(o)
            db.session.commit()
            return True, '地址已更新'
    return False, '参数错误'


def normalize_floor(floor):
    """标准化楼层格式：一楼/2楼/B1楼 → 1F/2F/B1F"""
    import re
    if not floor:
        return floor
    CN_MAP = {'一': '1', '二': '2', '三': '3', '四': '4',
              '五': '5', '六': '6', '七': '7', '八': '8', '九': '9', '十': '10'}
    for cn, num in CN_MAP.items():
        floor = floor.replace(cn, num)
    floor = re.sub(r'(\d+)楼$', r'\1F', floor)
    return floor


def add_address(building, floor, department, location):
    """新增地址"""
    if not all([building, floor, department, location]):
        return False, '所有字段不能为空'
    floor = normalize_floor(floor)
    o = AddressOverride(base_index=-1, building=building,
                        floor=floor, department=department, location=location)
    db.session.add(o)
    db.session.commit()
    return True, f'已新增地址：「{building} {floor} {department} {location}」'


def delete_address(oid):
    """软删除地址"""
    o = AddressOverride.query.get_or_404(oid)
    o.is_deleted = True
    db.session.commit()
    return o.building


def delete_base_address(base_index, building=''):
    """删除基础地址"""
    from services.address import ADDRESS_LIST
    if base_index is None or base_index < 0:
        return False, '参数错误'
    if base_index >= len(ADDRESS_LIST):
        return False, '地址索引超出范围'
    o = AddressOverride.query.filter_by(base_index=base_index).first()
    if o:
        o.is_deleted = True
    else:
        addr = ADDRESS_LIST[base_index]
        o = AddressOverride(base_index=base_index, building=addr['楼区'],
                            floor=addr['所属楼层'], department=addr['所属科室'],
                            location=addr['物理地址'], is_deleted=True)
        db.session.add(o)
    db.session.commit()
    return True, '地址已删除'


# ===================================================================
#  5. 故障类型
# ===================================================================

def list_fault_types():
    """获取故障类型列表"""
    return FaultType.query.order_by(FaultType.sort_order, FaultType.id).all()


def add_fault_type(name, keywords):
    """新增故障类型"""
    if not name:
        return False, '请输入故障类型名称'
    if FaultType.query.filter_by(name=name).first():
        return False, f'故障类型「{name}」已存在'
    max_order = db.session.query(db.func.max(FaultType.sort_order)).scalar() or 0
    ft = FaultType(name=name, keywords=keywords, sort_order=max_order + 1)
    db.session.add(ft)
    db.session.commit()
    return True, f'已新增故障类型「{name}」'


def edit_fault_type(fid, name, keywords):
    """编辑故障类型"""
    ft = FaultType.query.get_or_404(fid)
    if not name:
        return False, '请输入故障类型名称'
    existing = FaultType.query.filter(FaultType.name == name, FaultType.id != fid).first()
    if existing:
        return False, f'故障类型「{name}」已存在'
    ft.name = name
    ft.keywords = keywords
    db.session.commit()
    return True, f'已更新故障类型「{name}」'


def delete_fault_type(fid, operator_name):
    """删除故障类型"""
    ft = FaultType.query.get_or_404(fid)
    name = ft.name
    db.session.delete(ft)
    db.session.commit()
    log_audit('delete', 'fault_type', operator_name,
              target_id=fid, target_desc=f'删除故障类型: {name}')
    return name


# ===================================================================
#  6. 存放位置
# ===================================================================

def list_storage_locations():
    """获取存放位置列表"""
    return StorageLocation.query.order_by(StorageLocation.sort_order, StorageLocation.name).all()


def add_storage_location(name, building, floor, area, contact, phone, is_default):
    """新增存放位置"""
    if not name:
        return False, '位置名称不能为空'
    if StorageLocation.query.filter_by(name=name).first():
        return False, f'位置「{name}」已存在'
    if is_default:
        StorageLocation.query.update({StorageLocation.is_default: False})
    sl = StorageLocation(name=name, building=building, floor=floor,
                         area=area, contact=contact, phone=phone, is_default=is_default)
    db.session.add(sl)
    db.session.commit()
    return True, f'已添加存放位置「{name}」'


def edit_storage_location(lid, name, building, floor, area, contact, phone, is_default):
    """编辑存放位置"""
    sl = StorageLocation.query.get_or_404(lid)
    if not name:
        return False, '位置名称不能为空'
    existing = StorageLocation.query.filter(
        StorageLocation.name == name, StorageLocation.id != lid).first()
    if existing:
        return False, f'位置「{name}」已存在'
    if is_default:
        StorageLocation.query.update({StorageLocation.is_default: False})
    sl.name = name
    sl.building = building
    sl.floor = floor
    sl.area = area
    sl.contact = contact
    sl.phone = phone
    sl.is_default = is_default
    db.session.commit()
    return True, f'已更新存放位置「{name}」'


def toggle_storage_location(lid):
    """切换启用/停用"""
    sl = StorageLocation.query.get_or_404(lid)
    sl.is_active = not sl.is_active
    db.session.commit()


def delete_storage_location(lid):
    """删除存放位置"""
    sl = StorageLocation.query.get_or_404(lid)
    name = sl.name
    db.session.delete(sl)
    db.session.commit()
    return name


# ===================================================================
#  7. 供应商
# ===================================================================

def list_suppliers():
    """获取供应商列表"""
    return Supplier.query.order_by(Supplier.sort_order, Supplier.name).all()


def add_supplier(name, contact_person, phone, address, service_scope, notes, contract_end_str):
    """新增供应商"""
    if not name:
        return False, '供应商名称不能为空'
    if Supplier.query.filter_by(name=name).first():
        return False, f'供应商「{name}」已存在'
    contract_end = None
    if contract_end_str:
        contract_end = datetime.strptime(contract_end_str, '%Y-%m-%d').date()
    s = Supplier(name=name, contact_person=contact_person, phone=phone,
                 address=address, service_scope=service_scope,
                 notes=notes, contract_end=contract_end)
    db.session.add(s)
    db.session.commit()
    return True, f'已添加供应商「{name}」'


def edit_supplier(sid, name, contact_person, phone, address, service_scope, notes, contract_end_str):
    """编辑供应商"""
    s = Supplier.query.get_or_404(sid)
    if not name:
        return False, '供应商名称不能为空'
    existing = Supplier.query.filter(Supplier.name == name, Supplier.id != sid).first()
    if existing:
        return False, f'供应商「{name}」已存在'
    s.name = name
    s.contact_person = contact_person
    s.phone = phone
    s.address = address
    s.service_scope = service_scope
    s.notes = notes
    s.contract_end = None
    if contract_end_str:
        s.contract_end = datetime.strptime(contract_end_str, '%Y-%m-%d').date()
    db.session.commit()
    return True, f'已更新供应商「{name}」'


def toggle_supplier(sid):
    """切换供应商启用/停用"""
    s = Supplier.query.get_or_404(sid)
    s.is_active = not s.is_active
    db.session.commit()


def delete_supplier(sid):
    """删除供应商"""
    s = Supplier.query.get_or_404(sid)
    name = s.name
    db.session.delete(s)
    db.session.commit()
    return name


# ===================================================================
#  8. 耗材管理
# ===================================================================

def list_consumables(q=''):
    """获取耗材列表"""
    query = Consumable.query
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            Consumable.name.like(like), Consumable.spec.like(like),
            Consumable.compatible_printers.like(like)))
    return query.order_by(Consumable.quantity.asc()).all()


def add_consumable(name, spec, unit, quantity, min_quantity, location, supplier_name, compatible_printers, notes):
    """新增耗材"""
    if not name:
        return False, '耗材名称不能为空'
    c = Consumable(name=name, spec=spec, unit=unit, quantity=quantity,
                   min_quantity=min_quantity, location=location,
                   supplier_name=supplier_name, compatible_printers=compatible_printers, notes=notes)
    db.session.add(c)
    db.session.commit()
    return True, f'已添加耗材「{name}」'


def edit_consumable(cid, name, spec, unit, quantity, min_quantity, location, supplier_name, compatible_printers, notes):
    """编辑耗材"""
    c = Consumable.query.get_or_404(cid)
    c.name = name
    c.spec = spec
    c.unit = unit
    c.quantity = quantity
    c.min_quantity = min_quantity
    c.location = location
    c.supplier_name = supplier_name
    c.compatible_printers = compatible_printers
    c.notes = notes
    db.session.commit()
    return True, f'已更新耗材「{c.name}」'


def delete_consumable(cid):
    """删除耗材"""
    c = Consumable.query.get_or_404(cid)
    name = c.name
    db.session.delete(c)
    db.session.commit()
    return name


def import_consumables_from_excel(file_storage):
    """从Excel导入耗材，返回 (success, imported, skipped, errors)"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_storage, data_only=True)
    except Exception as e:
        return False, 0, 0, [f'文件读取失败: {str(e)}']

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return False, 0, 0, ['空文件']

    header = [str(c or '').strip().lower() for c in rows[0]]
    data_rows = rows[1:]

    col_map = {
        '耗材名称': 'name', '名称': 'name', 'name': 'name',
        '规格': 'spec', '规格型号': 'spec', '型号': 'spec',
        '单位': 'unit', 'unit': 'unit',
        '库存': 'quantity', '数量': 'quantity', '库存量': 'quantity', 'quantity': 'quantity',
        '预警': 'min_quantity', '最低预警': 'min_quantity', '预警量': 'min_quantity', 'min': 'min_quantity',
        '位置': 'location', '存放位置': 'location', '存放点': 'location', 'location': 'location',
        '供应商': 'supplier_name', 'supplier': 'supplier_name',
        '适配': 'compatible_printers', '适配打印机': 'compatible_printers',
        '备注': 'notes', 'note': 'notes',
    }

    imported = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(data_rows, start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_dict = {}
        for col_idx, col_name in enumerate(header):
            field = col_map.get(col_name)
            if field:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    val = str(val).strip()
                row_dict[field] = val

        name = row_dict.get('name', '')
        if not name:
            skipped += 1
            continue

        existing = Consumable.query.filter_by(name=name).first()
        if existing:
            if row_dict.get('quantity'):
                existing.quantity = int(float(row_dict['quantity']))
            for f in ('spec', 'unit', 'location', 'supplier_name', 'compatible_printers', 'notes'):
                if row_dict.get(f):
                    setattr(existing, f, row_dict[f])
            if row_dict.get('min_quantity'):
                existing.min_quantity = int(float(row_dict['min_quantity']))
            imported += 1
            continue

        try:
            c = Consumable(
                name=name, spec=row_dict.get('spec', ''),
                unit=row_dict.get('unit', '个'),
                quantity=int(float(row_dict['quantity'])) if row_dict.get('quantity') else 0,
                min_quantity=int(float(row_dict['min_quantity'])) if row_dict.get('min_quantity') else 5,
                location=row_dict.get('location', ''),
                supplier_name=row_dict.get('supplier_name', ''),
                compatible_printers=row_dict.get('compatible_printers', ''),
                notes=row_dict.get('notes', ''),
            )
            db.session.add(c)
            imported += 1
        except Exception as e:
            errors.append(f'第{idx}行({name}): {str(e)}')
            skipped += 1

    db.session.commit()
    return True, imported, skipped, errors


def consumable_inout(cid, action, qty, note, operator_name, department=''):
    """耗材出入库，返回 (ok, msg, balance)"""
    if not cid or not action or qty <= 0:
        return False, '参数错误', None

    c = Consumable.query.get_or_404(cid)
    if action == 'out' and c.quantity < qty:
        return False, f'库存不足（当前 {c.quantity}{c.unit}）', None

    c.quantity += qty if action == 'in' else -qty
    record = ConsumableRecord(consumable_id=cid, type=action, quantity=qty,
                              balance=c.quantity, operator=operator_name,
                              note=note, department=department)
    db.session.add(record)
    db.session.commit()

    action_name = '入库' if action == 'in' else '出库'
    log_audit(action, 'consumable', operator_name,
              target_id=cid, target_desc=f'耗材{action_name} {c.name} ×{qty}')
    return True, action_name, c.quantity


def batch_consumable_out(items, department, operator_name):
    """耗材一键出库，items=[{cid,qty},...]，返回 (ok, msg, out_records)"""
    if not department:
        return False, '请选择目标科室', []
    if not items:
        return False, '请选择出库物品', []

    out_records = []
    for item in items:
        cid = item.get('cid')
        qty = item.get('qty', 0)
        if not cid or qty <= 0:
            continue
        c = Consumable.query.get(cid)
        if not c:
            continue
        if c.quantity < qty:
            return False, f'「{c.name}」库存不足（当前 {c.quantity}{c.unit}）', []
        c.quantity -= qty
        record = ConsumableRecord(
            consumable_id=cid, type='out', quantity=qty,
            balance=c.quantity, operator=operator_name,
            department=department, note=f'一键出库至{department}',
        )
        db.session.add(record)
        out_records.append({'name': c.name, 'qty': qty, 'unit': c.unit})
    db.session.commit()
    log_audit('batch_out', 'consumable', operator_name,
              target_desc=f'耗材一键出库至{department}，共{len(out_records)}项')
    return True, f'已出库 {len(out_records)} 项至「{department}」', out_records


def export_consumables_template():
    """生成耗材导入模板Excel，返回 BytesIO"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '耗材导入模板'
    headers = ['耗材名称', '规格型号', '单位', '库存量', '最低预警量', '存放位置', '供应商', '适配打印机', '备注']
    hf = Font(bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='EC4899', end_color='EC4899', fill_type='solid')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = tb
    ws.column_dimensions['A'].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ===================================================================
#  9. 值班排班
# ===================================================================

def get_duty_schedule_staff():
    """获取在职人员列表"""
    return Person.query.filter_by(is_active=True).order_by(Person.sort_order, Person.id).all()


def get_duty_month_info(year, month):
    """获取月份信息：total_days, first_weekday, holidays"""
    if month < 1 or month > 12:
        month = date.today().month
    total_days = monthrange(year, month)[1]
    first_weekday = date(year, month, 1).weekday()
    holidays = set()
    try:
        from chinese_calendar import is_holiday
        for day in range(1, total_days + 1):
            d = date(year, month, day)
            if is_holiday(d):
                holidays.add(day)
    except ImportError:
        pass
    return total_days, first_weekday, holidays


def get_duty_schedules_api(year, month):
    """获取某月排班 JSON 数据"""
    records = DutySchedule.query.filter(
        db.extract('year', DutySchedule.duty_date) == year,
        db.extract('month', DutySchedule.duty_date) == month
    ).all()
    data = {}
    for r in records:
        day = r.duty_date.day
        data[f'{r.person_name}_{day}'] = r.shift
    return data


def duty_schedule_update(year, month, person, day, shift):
    """更新单个排班，返回 (ok, msg, shift)"""
    if not person or not day or not year or not month:
        return False, '参数不完整', None
    d = date(year, month, day)
    existing = DutySchedule.query.filter_by(duty_date=d, person_name=person).first()
    if shift == '×' or shift == '':
        if existing:
            db.session.delete(existing)
            db.session.commit()
        return True, None, '×'
    if existing:
        existing.shift = shift
    else:
        db.session.add(DutySchedule(duty_date=d, person_name=person, shift=shift))
    db.session.commit()
    return True, None, shift


def duty_schedule_batch(action, year, month, **kwargs):
    """批量排班操作，返回 (ok, msg)"""
    if action == 'fill_row':
        person = kwargs.get('person', '')
        shift = kwargs.get('shift', '')
        if not person or not shift:
            return False, '参数不完整'
        total_days = monthrange(year, month)[1]
        for day in range(1, total_days + 1):
            d = date(year, month, day)
            existing = DutySchedule.query.filter_by(duty_date=d, person_name=person).first()
            if shift == '×':
                if existing:
                    db.session.delete(existing)
            else:
                if existing:
                    existing.shift = shift
                else:
                    db.session.add(DutySchedule(duty_date=d, person_name=person, shift=shift))
        db.session.commit()
        return True, f'已填充 {person} {total_days} 天'

    elif action == 'fill_workdays':
        person = kwargs.get('person', '')
        shift = kwargs.get('shift', '')
        if not person or not shift:
            return False, '参数不完整'
        total_days = monthrange(year, month)[1]
        count = 0
        skipped = 0
        for day in range(1, total_days + 1):
            wd = date(year, month, day).weekday()
            if wd >= 5:
                continue
            d = date(year, month, day)
            try:
                from chinese_calendar import is_holiday
                if is_holiday(d):
                    skipped += 1
                    continue
            except ImportError:
                pass
            existing = DutySchedule.query.filter_by(duty_date=d, person_name=person).first()
            if shift == '×':
                if existing:
                    db.session.delete(existing)
                    count += 1
            else:
                if existing:
                    existing.shift = shift
                else:
                    db.session.add(DutySchedule(duty_date=d, person_name=person, shift=shift))
                count += 1
        db.session.commit()
        msg = f'已填充 {person} {count} 个工作日'
        if skipped:
            msg += f'（跳过 {skipped} 个法定节假日）'
        return True, msg

    elif action == 'clear':
        DutySchedule.query.filter(
            db.extract('year', DutySchedule.duty_date) == year,
            db.extract('month', DutySchedule.duty_date) == month
        ).delete()
        db.session.commit()
        return True, '已清空该月排班'

    elif action == 'copy':
        prev_month = month - 1
        prev_year = year
        if prev_month == 0:
            prev_month = 12
            prev_year = year - 1
        prev_records = DutySchedule.query.filter(
            db.extract('year', DutySchedule.duty_date) == prev_year,
            db.extract('month', DutySchedule.duty_date) == prev_month
        ).all()
        cur_days = monthrange(year, month)[1]
        for r in prev_records:
            old_day = r.duty_date.day
            new_day = min(old_day, cur_days)
            d = date(year, month, new_day)
            if not DutySchedule.query.filter_by(duty_date=d, person_name=r.person_name).first():
                db.session.add(DutySchedule(duty_date=d, person_name=r.person_name, shift=r.shift))
        db.session.commit()
        return True, f'已从 {prev_year}年{prev_month}月 复制排班'

    return False, '未知操作'


def duty_schedule_import_excel(file_storage, year, month):
    """导入 Excel 排班表"""
    import openpyxl
    if not year or not month:
        return False, '请选择年月'

    name = getattr(file_storage, 'filename', '')
    if not name or not name.endswith(('.xlsx', '.xls')):
        return False, '仅支持 .xlsx/.xls 文件'

    try:
        wb = openpyxl.load_workbook(file_storage, data_only=True)
    except Exception as e:
        return False, f'文件读取失败: {str(e)}'

    ws = wb.active
    skip_names = {'×', '夜班', '病假', '事假', '年假', '24H', '支援', '日班', '日期', '星期', ''}
    schedules = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name_val = row[0]
        if not name_val or not str(name_val).strip():
            continue
        name_val = str(name_val).strip()
        if name_val in skip_names:
            continue
        for day_idx in range(1, len(row)):
            shift_val = row[day_idx]
            if shift_val is None:
                continue
            shift = str(shift_val).strip()
            if not shift:
                continue
            day = day_idx
            if day < 1 or day > 31:
                continue
            try:
                d = date(year, month, day)
            except ValueError:
                continue
            schedules.append({'person': name_val, 'date': d, 'shift': shift})

    if not schedules:
        return False, '未解析到有效排班数据，请检查Excel格式'

    DutySchedule.query.filter(
        db.extract('year', DutySchedule.duty_date) == year,
        db.extract('month', DutySchedule.duty_date) == month
    ).delete()

    for s in schedules:
        db.session.add(DutySchedule(duty_date=s['date'], person_name=s['person'], shift=s['shift']))

    db.session.commit()
    person_names = sorted(set(s['person'] for s in schedules))
    return True, f'导入成功！共 {len(schedules)} 条排班，{len(person_names)} 人（{"、".join(person_names)}）'


# ---- 排班人员管理 ----

def list_duty_staff():
    """获取排班人员列表"""
    return DutyStaff.query.order_by(DutyStaff.sort_order).all()


def add_duty_staff(name):
    """新增排班人员"""
    if not name:
        return False, '名称不能为空'
    if DutyStaff.query.filter_by(name=name).first():
        return False, f'人员「{name}」已存在'
    max_order = db.session.query(db.func.max(DutyStaff.sort_order)).scalar() or 0
    db.session.add(DutyStaff(name=name, sort_order=max_order + 1))
    db.session.commit()
    return True, f'已添加人员「{name}」'


def toggle_duty_staff(sid):
    """启用/禁用排班人员"""
    s = DutyStaff.query.get_or_404(sid)
    s.is_active = not s.is_active
    db.session.commit()
    return s.is_active


def delete_duty_staff(sid):
    """删除排班人员"""
    s = DutyStaff.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()


# ===================================================================
#  10. 知识库
# ===================================================================

def list_knowledge(category=''):
    """获取知识库文章列表"""
    query = KnowledgeBase.query
    if category:
        query = query.filter(KnowledgeBase.category == category)
    articles = query.order_by(KnowledgeBase.is_pinned.desc(), KnowledgeBase.created_at.desc()).all()
    cats = db.session.query(KnowledgeBase.category).distinct().all()
    categories = [c[0] for c in cats if c[0]]
    return articles, categories


def add_knowledge(title, category, content, is_pinned):
    """新增知识库文章"""
    if not title:
        return False, '标题不能为空'
    a = KnowledgeBase(title=title, category=category, content=content, is_pinned=is_pinned)
    db.session.add(a)
    db.session.commit()
    return True, f'已添加文章「{title}」'


def edit_knowledge(kid, title, category, content, is_pinned):
    """编辑知识库文章"""
    a = KnowledgeBase.query.get_or_404(kid)
    a.title = title
    a.category = category
    a.content = content
    a.is_pinned = is_pinned
    db.session.commit()
    return True, f'已更新文章「{a.title}」'


def get_knowledge_api(kid):
    """获取单篇文章 JSON"""
    a = KnowledgeBase.query.get_or_404(kid)
    return {'id': a.id, 'title': a.title, 'category': a.category,
            'content': a.content, 'is_pinned': a.is_pinned}


def delete_knowledge(kid):
    """删除知识库文章"""
    a = KnowledgeBase.query.get_or_404(kid)
    title = a.title
    db.session.delete(a)
    db.session.commit()
    return title


# ===================================================================
#  11. 权限管理
# ===================================================================

def get_permissions_page_data():
    """获取权限管理页面数据"""
    users = User.query.order_by(User.is_admin.desc(), User.username).all()
    module_perms = get_module_permissions(refresh=True)
    from models import ALL_MODULE_NAMES
    persons = Person.query.filter_by(is_active=True).order_by(Person.sort_order, Person.name).all()
    # 构建角色组用户分组
    users_by_group = {}
    for u in users:
        if u.is_admin:
            gname = '管理员'
        else:
            gname = u.group or '普通用户'
        if gname not in users_by_group:
            users_by_group[gname] = {'group_name': gname, 'users': [], 'count': 0}
        users_by_group[gname]['users'].append(u)
        users_by_group[gname]['count'] += 1
    role_groups = module_perms.get('groups', {})
    return users, module_perms, ALL_MODULE_NAMES, persons, users_by_group, role_groups


def sync_users_from_persons(operator_name):
    """从人员同步创建用户"""
    from models import Hospital
    persons = Person.query.filter_by(is_active=True).all()
    created = 0
    for p in persons:
        existing = User.query.filter_by(display_name=p.name).first()
        if not existing:
            pinyin = p.name
            if not User.query.filter_by(username=pinyin).first():
                user = User(username=pinyin, display_name=p.name, phone=p.phone,
                            group=p.group or 'IT', is_admin=False,
                            hospital_id=p.hospital_id)
                user.set_password('123456')
                # 同步所属医院
                if p.hospital_id:
                    hospital = db.session.get(Hospital, p.hospital_id)
                    if hospital:
                        user.hospitals = [hospital]
                db.session.add(user)
                created += 1
    if created:
        db.session.commit()
        log_audit('sync_users', 'system', operator_name,
                  detail=f'从人员同步创建了 {created} 个用户账号')
        return created, f'已从人员同步创建 {created} 个用户账号（默认密码: 123456）'
    else:
        return 0, '所有人员已有对应账号，无需同步'


def save_permissions(data):
    """保存角色组权限"""
    save_module_permissions(data)
    return True, '权限配置已保存'


def toggle_admin(uid, current_uid, operator_name):
    """切换用户管理员权限，返回 (ok, msg, action_desc)"""
    if uid == current_uid:
        return False, '不能撤销自己的管理员权限', None
    user = User.query.get_or_404(uid)
    user.is_admin = not user.is_admin
    action_desc = f'{"赋予" if user.is_admin else "撤销"} {user.display_name or user.username} 管理员权限'
    log_audit('toggle_admin', 'user', operator_name,
              target_id=user.id, target_desc=action_desc)
    db.session.commit()
    return True, f'已{action_desc}', action_desc


def set_user_group(uid, group, operator_name):
    """设置用户角色组"""
    user = User.query.get_or_404(uid)
    user.group = group
    db.session.commit()
    log_audit('set_user_group', 'user', operator_name,
              target_id=uid, detail=f'将用户 {user.display_name or user.username} 设为 [{group}] 组')
    return True, f'已将 {user.display_name or user.username} 设为 [{group}] 组'


def add_permission_group(name):
    """添加角色组"""
    if not name:
        return False, '请输入角色组名称'
    perms = get_module_permissions(refresh=True)
    groups = perms.get('groups', {})
    if name in groups:
        return False, f'角色组「{name}」已存在'
    default_user = groups.get('普通用户', {})
    groups[name] = dict(default_user)
    save_module_permissions(perms)
    return True, f'已添加角色组「{name}」'


def delete_permission_group(name):
    """删除角色组"""
    if name in ('管理员', '普通用户'):
        return False, '不能删除默认角色组'
    perms = get_module_permissions(refresh=True)
    groups = perms.get('groups', {})
    if name not in groups:
        return False, f'角色组「{name}」不存在'
    del groups[name]
    for u in User.query.filter_by(group=name).all():
        u.group = '普通用户'
    db.session.commit()
    save_module_permissions(perms)
    return True, f'已删除角色组「{name}」，相关用户已重置为「普通用户」'


def add_permission_module(module, category):
    """添加权限模块"""
    if not module or not category:
        return False, '请填写模块名称和分组'
    perms = get_module_permissions(refresh=True)
    for cat, mods in perms.get('categories', []):
        if module in mods:
            return False, f'模块「{module}」已存在'
    categories = perms.get('categories', [])
    found = False
    for cat, mods in categories:
        if cat == category:
            mods.append(module)
            found = True
            break
    if not found:
        categories.append([category, [module]])
    for gname, g in perms.get('groups', {}).items():
        g[module] = (gname == '管理员')
    perms['categories'] = categories
    save_module_permissions(perms)
    return True, f'已添加模块「{module}」到 [{category}] 分组'


def delete_permission_module(module):
    """删除权限模块"""
    if not module:
        return False, '请指定要删除的模块'
    perms = get_module_permissions(refresh=True)
    categories = perms.get('categories', [])
    for cat in categories:
        if module in cat[1]:
            cat[1].remove(module)
            break
    for g in perms.get('groups', {}).values():
        g.pop(module, None)
    perms['categories'] = categories
    save_module_permissions(perms)
    return True, f'已删除模块「{module}」'


# ===================================================================
#  12. 医院管理
# ===================================================================

def add_hospital(name, code, address, phone):
    """新增医院"""
    if not name or not code:
        return False, '医院名称和编码不能为空'
    if Hospital.query.filter_by(code=code).first():
        return False, f'编码 "{code}" 已存在'
    h = Hospital(name=name, code=code, address=address, phone=phone)
    db.session.add(h)
    db.session.commit()
    return True, f'已添加医院 "{name}"'


def edit_hospital(hid, name, code, address, phone):
    """编辑医院"""
    h = db.session.get(Hospital, hid)
    if not h:
        return False, '医院不存在'
    h.name = name or h.name
    if code and code != h.code:
        if Hospital.query.filter_by(code=code).first():
            return False, f'编码 "{code}" 已存在'
        h.code = code
    h.address = address or h.address
    h.phone = phone or h.phone
    db.session.commit()
    return True, f'已更新医院 "{h.name}"'


def toggle_hospital(hid):
    """切换医院启用/禁用"""
    h = db.session.get(Hospital, hid)
    if h:
        h.is_active = not h.is_active
        db.session.commit()
        return True, f'医院 "{h.name}" 已{"启用" if h.is_active else "禁用"}'
    return False, '医院不存在'


# ===================================================================
#  13. 零件价格管理
# ===================================================================

def list_parts(q='', cat='', supplier=''):
    """获取零件价格列表"""
    query = PartPrice.query
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            PartPrice.product_name.ilike(like), PartPrice.brand.ilike(like),
            PartPrice.model_no.ilike(like), PartPrice.supplier.ilike(like),
            PartPrice.spec.ilike(like), PartPrice.remark.ilike(like)))
    if cat:
        query = query.filter(PartPrice.category == cat)
    if supplier:
        query = query.filter(PartPrice.supplier.ilike(f'%{supplier}%'))
    parts = query.order_by(PartPrice.category, PartPrice.product_name).all()
    categories = [r[0] for r in db.session.query(PartPrice.category).distinct().order_by(PartPrice.category).all() if r[0]]
    suppliers = [r[0] for r in db.session.query(PartPrice.supplier).distinct().order_by(PartPrice.supplier).all() if r[0]]
    return parts, categories, suppliers


def add_part(product_name, unit, unit_price, category, spec, brand, model_no, supplier, remark, operator_name):
    """新增零件"""
    if not product_name:
        return False, '请输入零件名称'
    p = PartPrice(product_name=product_name, unit=unit or '个',
                  unit_price=unit_price, category=category or '电脑配件',
                  spec=spec, brand=brand, model_no=model_no,
                  supplier=supplier, remark=remark)
    db.session.add(p)
    db.session.commit()
    log_audit('create', 'part_price', operator_name, target_desc=f'新增零件: {product_name}')
    return True, f'已添加零件 "{product_name}"'


def delete_part(part_id, operator_name):
    """删除零件"""
    p = PartPrice.query.get_or_404(part_id)
    name = p.product_name
    db.session.delete(p)
    db.session.commit()
    log_audit('delete', 'part_price', operator_name, target_id=part_id, target_desc=f'删除零件: {name}')
    return name


# ===================================================================
#  14. 耗材出入库记录
# ===================================================================

def list_consumable_records(q='', action='', department='', page=1, page_size=50):
    """获取耗材出入库记录"""
    query = ConsumableRecord.query.join(Consumable, ConsumableRecord.consumable_id == Consumable.id)
    if q:
        like = f'%{q}%'
        query = query.filter(Consumable.name.like(like))
    if action in ('in', 'out'):
        query = query.filter(ConsumableRecord.type == action)
    if department:
        query = query.filter(ConsumableRecord.department == department)
    query = query.order_by(ConsumableRecord.created_at.desc())
    pagination = query.paginate(page=page, per_page=page_size, error_out=False)
    records = pagination.items
    total = query.count()
    return records, pagination, total


def get_consumable_departments():
    """获取所有耗材出库科室列表"""
    rows = db.session.query(ConsumableRecord.department)\
        .filter(ConsumableRecord.type == 'out', ConsumableRecord.department != '')\
        .distinct().order_by(ConsumableRecord.department).all()
    return [r[0] for r in rows]


def delete_consumable_record(rid):
    """删除出入库记录"""
    r = ConsumableRecord.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return True


# ===================================================================
#  15. 故障模板组
# ===================================================================

def list_fault_template_groups():
    """获取故障模板组列表"""
    groups = FaultTemplateGroup.query.order_by(FaultTemplateGroup.id).all()
    import re
    team_setting = SystemSetting.query.filter_by(key='person_teams').first()
    all_teams = []
    if team_setting and team_setting.value:
        all_teams = [t.strip() for t in re.split(r'[,，]', team_setting.value) if t.strip()]
    return groups, all_teams


def add_fault_template_group(name, teams, operator_name):
    """新增故障模板组"""
    if not name:
        return False, '模板组名称不能为空'
    if FaultTemplateGroup.query.filter_by(name=name).first():
        return False, f'模板组 "{name}" 已存在'
    g = FaultTemplateGroup(name=name, teams=','.join(teams))
    db.session.add(g)
    db.session.commit()
    log_audit('create', 'fault_template_group', operator_name, target_desc=f'创建故障模板组: {name}')
    return True, f'已创建模板组 "{name}"'


def edit_fault_template_group(gid, field, value):
    """编辑故障模板组"""
    g = FaultTemplateGroup.query.get_or_404(gid)
    if field == 'name':
        g.name = value
    elif field == 'teams':
        g.teams = value
    db.session.commit()
    return True


def delete_fault_template_group(gid, operator_name):
    """删除故障模板组"""
    g = FaultTemplateGroup.query.get_or_404(gid)
    name = g.name
    FaultTemplateItem.query.filter_by(group_id=gid).delete()
    db.session.delete(g)
    db.session.commit()
    log_audit('delete', 'fault_template_group', operator_name,
              target_id=gid, target_desc=f'删除故障模板组: {name}')
    return name


def add_fault_template_item(gid, fault_type, display_name, default_count):
    """新增故障模板项"""
    g = FaultTemplateGroup.query.get_or_404(gid)
    item = FaultTemplateItem(group_id=g.id, fault_type=fault_type or '硬件',
                             display_name=display_name or fault_type or '硬件',
                             default_count=max(1, default_count or 1),
                             sort_order=g.items.count() + 1)
    db.session.add(item)
    db.session.commit()
    return True, f'已添加故障项 "{display_name}"'


def edit_fault_template_item(gid, iid, fault_type, display_name, default_count):
    """编辑故障模板项"""
    item = FaultTemplateItem.query.get_or_404(iid)
    item.fault_type = fault_type or item.fault_type
    item.display_name = display_name or item.display_name
    item.default_count = max(1, int(default_count or item.default_count))
    db.session.commit()
    return True


def delete_fault_template_item(gid, iid):
    """删除故障模板项"""
    item = FaultTemplateItem.query.get_or_404(iid)
    db.session.delete(item)
    db.session.commit()
    return True


# ===================================================================
#  16. 故障二级分类
# ===================================================================

def list_fault_categories():
    """获取故障二级分类列表"""
    return FaultCategory.query.order_by(FaultCategory.sort_order).all()


def add_fault_subcategory(cat_id, name):
    """新增子分类"""
    if not cat_id or not name:
        return False, '参数错误'
    cat = FaultCategory.query.get(cat_id)
    if not cat:
        return False, '分类不存在'
    max_order = db.session.query(db.func.max(FaultSubcategory.sort_order)).filter_by(category_id=cat_id).scalar() or 0
    sub = FaultSubcategory(category_id=cat_id, name=name, sort_order=max_order + 1)
    db.session.add(sub)
    db.session.commit()
    return True, f'已添加子分类「{name}」'


def delete_fault_subcategory(sub_id):
    """删除子分类及其关键词"""
    sub = FaultSubcategory.query.get_or_404(sub_id)
    FaultKeyword.query.filter_by(subcategory_id=sub.id).delete()
    db.session.delete(sub)
    db.session.commit()
    return True


def add_fault_keywords(sub_id, keywords_raw):
    """批量添加关键词"""
    if not sub_id or not keywords_raw:
        return False, '参数错误'
    sub = FaultSubcategory.query.get(sub_id)
    if not sub:
        return False, '子分类不存在'
    kw_list = [k.strip() for k in keywords_raw.replace('\\n', ',').split(',') if k.strip()]
    max_order = db.session.query(db.func.max(FaultKeyword.sort_order)).filter_by(subcategory_id=sub_id).scalar() or 0
    for kw in kw_list:
        max_order += 1
        k = FaultKeyword(subcategory_id=sub_id, keyword=kw, sort_order=max_order)
        db.session.add(k)
    db.session.commit()
    return True, f'已添加 {len(kw_list)} 个关键词'


def delete_fault_keyword(kw_id):
    """删除关键词"""
    kw = FaultKeyword.query.get_or_404(kw_id)
    db.session.delete(kw)
    db.session.commit()
    return True
