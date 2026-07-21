"""资产台账管理：列表/导入/导出/批量操作/操作日志"""
import io
import json
from datetime import datetime, date
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, g, abort
from flask_login import login_required, current_user
from routes.auth import admin_required
from models import db, Asset, AssetLog, log_audit

asset_bp = Blueprint('asset', __name__, url_prefix='/asset')


# ===================== 台账列表 =====================

@asset_bp.route('/')
@login_required
def calendar_redirect():
    """重定向到列表页"""
    return redirect(url_for('asset.list_view'))


@asset_bp.route('/calendar')
@login_required
def calendar():
    """保修日历：按剩余天数排序展示所有设备/软件"""
    # --- 获取筛选参数：状态筛选和分类筛选 ---
    cat_filter = request.args.get('cat', 'all')
    status_filter = request.args.get('status', 'all')
    # --- 构建基础查询 ---
    query = Asset.query

    # --- 按分类筛选：硬件或软件 ---
    if cat_filter == 'hardware':
        query = query.filter(Asset.category == 'hardware')
    elif cat_filter == 'software':
        query = query.filter(Asset.category == 'software')

    # --- 按保修状态筛选 ---
    if status_filter == 'expiring':
        # --- 即将到期：保修结束在30天内 ---
        today = date.today()
        in_30 = date(today.year + (today.month + 1) // 12, (today.month % 12) + 1, 1)
        from sqlalchemy import and_
        query = query.filter(Asset.warranty_end.isnot(None))\
                     .filter(Asset.warranty_end >= today)\
                     .filter(Asset.warranty_end <= in_30)
    elif status_filter == 'expired':
        # --- 已过期：无保修或已过保修期 ---
        query = query.filter(db.or_(
            Asset.warranty_end.is_(None),
            Asset.warranty_end < date.today()
        ))
    elif status_filter == 'valid':
        # --- 保修中：有未过期的保修 ---
        query = query.filter(Asset.warranty_end.isnot(None))\
                     .filter(Asset.warranty_end >= date.today())

    # --- 按保修到期时间升序排列，空值置后，返回模板 ---
    assets = query.order_by(Asset.warranty_end.asc().nullslast()).all()
    return render_template('asset/calendar.html', assets=assets, now=datetime.now(),
                           status_filter=status_filter, cat_filter=cat_filter)


@asset_bp.route('/list')
@login_required
def list_view():
    """台账列表：搜索/筛选/分页"""
    # --- 获取分页和筛选参数 ---
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')
    cat_filter = request.args.get('cat', '')
    dept_filter = request.args.get('department', '')

    # --- 构建基础查询 ---
    query = Asset.query

    # --- 多字段模糊搜索 ---
    if search:
        safe_search = search.replace('%', '\\%').replace('_', '\\_')
        like = f'%{safe_search}%'
        query = query.filter(
            db.or_(
                Asset.asset_no.ilike(like),
                Asset.brand.ilike(like),
                Asset.model_no.ilike(like),
                Asset.sn.ilike(like),
                Asset.department.ilike(like),
                Asset.device_type.ilike(like),
                Asset.ip_address.ilike(like),
                Asset.cpu.ilike(like),
                Asset.financial_code.ilike(like),
                Asset.notes.ilike(like),
            )
        )
    # --- 按状态筛选 ---
    if status_filter:
        query = query.filter(Asset.status == status_filter)
    # --- 按分类筛选 ---
    if cat_filter:
        query = query.filter(Asset.category == cat_filter)
    # --- 按科室筛选 ---
    if dept_filter:
        safe_dept = dept_filter.replace('%', '\\%').replace('_', '\\_')
        query = query.filter(Asset.department.ilike(f'%{safe_dept}%'))

    # --- 分页查询，按更新时间降序排列 ---
    pagination = query.order_by(Asset.updated_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    # --- 获取所有科室列表供前端筛选下拉框使用 ---
    departments = [r[0] for r in db.session.query(Asset.department).distinct().all() if r[0]]
    return render_template('asset/list.html', pagination=pagination,
                           assets=pagination.items, departments=sorted(departments),
                           search=search, status_filter=status_filter,
                           cat_filter=cat_filter, dept_filter=dept_filter,
                           now=datetime.now())


# ===================== 详情/添加/编辑/删除（继承已有逻辑） =====================

@asset_bp.route('/detail/<int:asset_id>')
@login_required
def detail(asset_id):
    asset = Asset.query.filter(Asset.id == asset_id).filter(Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
    if not asset:
        abort(404)
    # --- 关联工单：按相同科室 + 设备类型/故障类型筛选 ---
    query = WorkOrder.query.filter(
        WorkOrder.department == asset.department,
        WorkOrder.created_at >= datetime.now().replace(year=datetime.now().year - 1)
    )
    if asset.device_type == 'PR':
        # --- 打印机 → 匹配故障类型为"打印机" ---
        query = query.filter(WorkOrder.fault_type == '打印机')
    elif asset.device_type in ('PC', 'NB', '一体机'):
        # --- 硬件 → 匹配 device_type ---
        query = query.filter(WorkOrder.device_type == asset.device_type)
    elif asset.device_type and asset.device_type not in ('','其他'):
        # --- 其他类型 ---
        query = query.filter(WorkOrder.device_type == asset.device_type)
    # --- 取最近20条关联工单 ---
    orders = query.order_by(WorkOrder.created_at.desc()).limit(20).all()
    return render_template('asset/detail.html', asset=asset, orders=orders,
                           now=datetime.now())


@asset_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    # --- 权限校验：仅管理员可新增资产 ---
    if not current_user.is_admin:
        flash('仅管理员可操作', 'danger')
        return redirect(url_for('asset.calendar'))
    if request.method == 'POST':
        try:
            # --- 从表单数据构建新资产对象 ---
            asset = Asset(
                category=request.form.get('category', 'hardware'),
                asset_no=request.form.get('asset_no', ''),
                device_type=request.form.get('device_type', 'PC'),
                brand=request.form.get('brand', ''),
                model_no=request.form.get('model_no', ''),
                sn=request.form.get('sn', ''),
                license_key=request.form.get('license_key', ''),
                license_seats=_parse_int(request.form.get('license_seats')),
                vendor=request.form.get('vendor', ''),
                department=request.form.get('department', ''),
                building=request.form.get('building', ''),
                floor=request.form.get('floor', ''),
                location=request.form.get('location', ''),
                purchase_date=_parse_date(request.form.get('purchase_date')),
                price=_parse_float(request.form.get('price')),
                warranty_start=_parse_date(request.form.get('warranty_start')),
                warranty_end=_parse_date(request.form.get('warranty_end')),
                status=request.form.get('status', 'in_use'),
                notes=request.form.get('notes', ''),
                ip_address=request.form.get('ip_address', ''),
                mac_address=request.form.get('mac_address', ''),
                cpu=request.form.get('cpu', ''),
                memory=request.form.get('memory', ''),
                disk_size=request.form.get('disk_size', ''),
                operating_system=request.form.get('operating_system', ''),
                financial_code=request.form.get('financial_code', ''),
                financial_name=request.form.get('financial_name', ''),
                hospital_id=getattr(g, 'hospital_id', 1),
            )
            # --- 写入数据库并记录操作日志 ---
            db.session.add(asset)
            db.session.commit()
            log_audit('create', 'asset', current_user.display_name,
                      target_id=asset.id, target_desc=f'新增 {asset.asset_no}')
            flash('添加成功', 'success')
            return redirect(url_for('asset.list_view'))
        except Exception as e:
            # --- 失败时回滚并提示错误 ---
            db.session.rollback()
            flash(f'添加失败: {str(e)}', 'danger')
    # --- GET 请求时返回空表单 ---
    return render_template('asset/form.html', asset=None, now=datetime.now())


@asset_bp.route('/edit/<int:asset_id>', methods=['GET', 'POST'])
@login_required
def edit(asset_id):
    # --- 权限校验：仅管理员可编辑资产 ---
    if not current_user.is_admin:
        flash('仅管理员可操作', 'danger')
        return redirect(url_for('asset.calendar'))
    # --- 获取待编辑的资产记录 ---
    asset = Asset.query.filter(Asset.id == asset_id).filter(Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
    if not asset:
        abort(404)
    if request.method == 'POST':
        try:
            # --- 保存修改前的旧值用于变更追溯 ---
            old = _asset_to_dict(asset)
            # --- 逐字段更新表单数据 ---
            asset.asset_no = request.form.get('asset_no', '')
            asset.category = request.form.get('category', 'hardware')
            asset.device_type = request.form.get('device_type', 'PC')
            asset.brand = request.form.get('brand', '')
            asset.model_no = request.form.get('model_no', '')
            asset.sn = request.form.get('sn', '')
            asset.license_key = request.form.get('license_key', '')
            asset.license_seats = _parse_int(request.form.get('license_seats'))
            asset.vendor = request.form.get('vendor', '')
            asset.department = request.form.get('department', '')
            asset.building = request.form.get('building', '')
            asset.floor = request.form.get('floor', '')
            asset.location = request.form.get('location', '')
            asset.purchase_date = _parse_date(request.form.get('purchase_date'))
            asset.price = _parse_float(request.form.get('price'))
            asset.warranty_start = _parse_date(request.form.get('warranty_start'))
            asset.warranty_end = _parse_date(request.form.get('warranty_end'))
            asset.status = request.form.get('status', 'in_use')
            asset.notes = request.form.get('notes', '')
            asset.ip_address = request.form.get('ip_address', '')
            asset.mac_address = request.form.get('mac_address', '')
            asset.cpu = request.form.get('cpu', '')
            asset.memory = request.form.get('memory', '')
            asset.disk_size = request.form.get('disk_size', '')
            asset.operating_system = request.form.get('operating_system', '')
            asset.financial_code = request.form.get('financial_code', '')
            asset.financial_name = request.form.get('financial_name', '')
            # --- 提交变更并记录新旧差异及审计日志 ---
            db.session.commit()
            new = _asset_to_dict(asset)
            _log_change(asset.id, 'edit', old, new, current_user.display_name)
            log_audit('update', 'asset', current_user.display_name,
                      target_id=asset.id, target_desc=f'修改 {asset.asset_no}')
            flash('修改成功', 'success')
            return redirect(url_for('asset.list_view'))
        except Exception as e:
            # --- 修改失败时回滚事务 ---
            db.session.rollback()
            flash(f'修改失败: {str(e)}', 'danger')
    # --- GET 请求渲染编辑表单 ---
    return render_template('asset/form.html', asset=asset, now=datetime.now())


@asset_bp.route('/delete/<int:asset_id>', methods=['POST'])
@login_required
def delete(asset_id):
    # --- 权限校验：仅管理员可删除资产 ---
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    # --- 查找资产并记录删除审计日志 ---
    asset = Asset.query.filter(Asset.id == asset_id).filter(Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
    if not asset:
        abort(404)
    log_audit('delete', 'asset', current_user.display_name,
              target_id=asset.id, target_desc=f'删除资产 {asset.asset_no}')
    # --- 执行删除操作 ---
    db.session.delete(asset)
    db.session.commit()
    return jsonify({'ok': True})


# ===================== 导入 Excel =====================

@asset_bp.route('/import', methods=['POST'])
@login_required
def import_excel():
    """从Excel导入资产（盘点信息参考格式）"""
    # --- 权限校验：仅管理员可导入 ---
    if not current_user.is_admin:
        flash('仅管理员可操作', 'danger')
        return redirect(url_for('asset.list_view'))
    # --- 检查 openpyxl 库是否可用 ---
    try:
        import openpyxl
    except ImportError:
        flash('服务端缺少 openpyxl 库，请联系管理员安装', 'danger')
        return redirect(url_for('asset.list_view'))

    # --- 检查上传文件是否存在 ---
    file = request.files.get('file')
    if not file or not file.filename:
        flash('请选择文件', 'danger')
        return redirect(url_for('asset.list_view'))

    try:
        # --- 打开 Excel 工作簿，读取所有行数据 ---
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash('空文件', 'danger')
            return redirect(url_for('asset.list_view'))

        # --- 解析表头和数据行 ---
        header = [str(c or '').strip() for c in rows[0]]
        data_rows = rows[1:]

        # --- 列名映射：中文列名 -> 模型字段名 ---
        col_map = {
            '资产名称': 'device_type', '设备类型': 'device_type',
            '资产编码': 'asset_no', '资产编号': 'asset_no',
            '品牌': 'brand', '品牌/型号': 'brand',
            '型号': 'model_no',
            'CPU': 'cpu', 'cpu': 'cpu',
            '内存': 'memory',
            '硬盘': 'disk_size', '磁盘': 'disk_size',
            'IP': 'ip_address', 'ip': 'ip_address', 'IP地址': 'ip_address', 'ip地址': 'ip_address',
            'MAC': 'mac_address', 'mac': 'mac_address', 'MAC地址': 'mac_address',
            '操作系统': 'operating_system', '系统': 'operating_system', 'OS': 'operating_system',
            '物理存放点': 'location', '存放位置': 'location', '位置': 'location',
            '科室': 'department', '所属科室': 'department', '使用科室': 'department',
            '杀毒': 'antivirus_placeholder', '桌管': 'desktop_mgmt_placeholder',
            '序列号': 'sn', 'SN': 'sn', 'sn': 'sn',
            '资产状态': 'status', '状态': 'status',
            '备注': 'notes',
            '财务编号': 'financial_code', '财务代码': 'financial_code',
            '资产名称(财务)': 'financial_name', '财务名称': 'financial_name',
            '楼栋': 'building', '楼层': 'floor',
            '购入日期': 'purchase_date', '采购日期': 'purchase_date',
            '价格': 'price', '金额': 'price',
            '保修开始': 'warranty_start', '保修起始': 'warranty_start',
            '保修到期': 'warranty_end', '保修结束': 'warranty_end',
        }

        imported = 0
        skipped = 0
        errors = []

        # --- 逐行解析导入数据 ---
        for idx, row in enumerate(data_rows, start=2):
            # --- 跳过完全空行 ---
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            # --- 将行数据按列名映射转换为字典 ---
            row_dict = {}
            for col_idx, col_name in enumerate(header):
                field = col_map.get(col_name)
                if field:
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is not None:
                        val = str(val).strip()
                        if val.lower() == 'none' or val == '':
                            val = None
                    row_dict[field] = val

            # --- 校验资产编码必填 ---
            asset_no = row_dict.get('asset_no')
            if not asset_no:
                errors.append(f'第{idx}行: 缺少资产编码，跳过')
                skipped += 1
                continue

            # --- 检查是否已存在，已存在则更新 ---
            existing = Asset.query.filter_by(asset_no=asset_no).first()
            if existing:
                old = _asset_to_dict(existing)
                for k, v in row_dict.items():
                    if k == 'asset_no':
                        continue
                    if v is not None:
                        setattr(existing, k, v)
                _log_change(existing.id, 'import', old, _asset_to_dict(existing), current_user.display_name)
                imported += 1
                continue

            # --- 不存在则创建新资产 ---
            try:
                asset = Asset(
                    asset_no=asset_no,
                    device_type=row_dict.get('device_type', 'PC'),
                    brand=row_dict.get('brand', ''),
                    model_no=row_dict.get('model_no', ''),
                    sn=row_dict.get('sn', ''),
                    department=row_dict.get('department', ''),
                    building=row_dict.get('building', ''),
                    floor=row_dict.get('floor', ''),
                    location=row_dict.get('location', ''),
                    ip_address=row_dict.get('ip_address', ''),
                    mac_address=row_dict.get('mac_address', ''),
                    cpu=row_dict.get('cpu', ''),
                    memory=row_dict.get('memory', ''),
                    disk_size=row_dict.get('disk_size', ''),
                    operating_system=row_dict.get('operating_system', ''),
                    financial_code=row_dict.get('financial_code', ''),
                    financial_name=row_dict.get('financial_name', ''),
                    status=row_dict.get('status', 'in_use'),
                    notes=row_dict.get('notes', ''),
                    purchase_date=_parse_date(row_dict.get('purchase_date')),
                    price=_parse_float(row_dict.get('price')),
                    warranty_start=_parse_date(row_dict.get('warranty_start')),
                    warranty_end=_parse_date(row_dict.get('warranty_end')),
                    hospital_id=getattr(g, 'hospital_id', 1),
                )
                db.session.add(asset)
                db.session.flush()
                _log_change(asset.id, 'import', {}, _asset_to_dict(asset), current_user.display_name)
                imported += 1
            except Exception as e:
                # --- 单行导入失败时记录错误并继续 ---
                db.session.rollback()
                errors.append(f'第{idx}行({asset_no}): {str(e)}')
                skipped += 1
                continue

        # --- 提交所有变更并汇总导入结果 ---
        db.session.commit()
        msg = f'导入完成: 成功 {imported} 条'
        if skipped:
            msg += f', 跳过 {skipped} 条'
        if errors:
            msg += f'<br>错误详情:<br>' + '<br>'.join(errors[:10])
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'导入失败: {str(e)}', 'danger')

    return redirect(url_for('asset.list_view'))


# ===================== 导出 Excel =====================

@asset_bp.route('/export')
@login_required
def export_excel():
    """导出资产台账为 Excel"""
    # --- 检查 openpyxl 库是否可用 ---
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        flash('服务端缺少 openpyxl 库，请联系管理员安装', 'danger')
        return redirect(url_for('asset.list_view'))

    # --- 获取筛选条件下的数据 ---
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    cat_filter = request.args.get('cat', '')
    dept_filter = request.args.get('department', '')

    # --- 构建查询并应用筛选条件 ---
    query = Asset.query
    if search:
        safe_search = search.replace('%', '\\%').replace('_', '\\_')
        like = f'%{safe_search}%'
        query = query.filter(db.or_(
            Asset.asset_no.ilike(like), Asset.brand.ilike(like),
            Asset.model_no.ilike(like), Asset.sn.ilike(like),
            Asset.department.ilike(like), Asset.device_type.ilike(like),
        ))
    if status_filter:
        query = query.filter(Asset.status == status_filter)
    if cat_filter:
        query = query.filter(Asset.category == cat_filter)
    if dept_filter:
        safe_dept = dept_filter.replace('%', '\\%').replace('_', '\\_')
        query = query.filter(Asset.department.ilike(f'%{safe_dept}%'))

    # --- 获取所有资产数据，按更新时间降序排列 ---
    assets = query.order_by(Asset.updated_at.desc()).all()

    # --- 创建 Excel 工作簿和工作表 ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '资产台账'

    # --- 写入表头并设置样式 ---
    headers = [
        '资产编码', '设备类型', '品牌', '型号', '序列号', 'CPU', '内存', '硬盘',
        '操作系统', 'IP地址', 'MAC地址', '科室', '楼栋', '楼层', '存放位置',
        '状态', '购入日期', '价格', '保修开始', '保修到期',
        '财务编号', '财务名称', '备注',
    ]
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # --- 状态值中文映射 ---
    status_map = {'in_use': '使用中', 'idle': '闲置', 'scrapped': '已报废', 'repair': '维修中'}

    # --- 逐行写入资产数据 ---
    for row_idx, a in enumerate(assets, 2):
        vals = [
            a.asset_no, a.device_type, a.brand, a.model_no, a.sn,
            a.cpu, a.memory, a.disk_size, a.operating_system,
            a.ip_address, a.mac_address,
            a.department, a.building, a.floor, a.location,
            status_map.get(a.status, a.status),
            a.purchase_date.strftime('%Y-%m-%d') if a.purchase_date else '',
            a.price, a.warranty_start.strftime('%Y-%m-%d') if a.warranty_start else '',
            a.warranty_end.strftime('%Y-%m-%d') if a.warranty_end else '',
            a.financial_code, a.financial_name, a.notes or '',
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # --- 自适应列宽 ---
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # --- 输出到字节流并返回下载 ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'资产台账_{date.today().strftime("%Y%m%d")}.xlsx',
    )


@asset_bp.route('/export-template')
@login_required
def export_template():
    """下载资产导入表头模板"""
    # --- 检查 openpyxl 库是否可用 ---
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        flash('服务端缺少 openpyxl 库', 'danger')
        return redirect(url_for('asset.list_view'))
    # --- 创建工作簿并写入中文表头 ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '资产导入模板'
    headers = ['资产编码', '设备类型', '品牌', '型号', '序列号', 'CPU', '内存', '硬盘',
               '操作系统', 'IP地址', 'MAC地址', '科室', '楼栋', '楼层', '存放位置',
               '状态', '购入日期', '价格', '保修开始', '保修到期', '财务编号', '财务名称', '备注']
    hf = Font(bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center', vertical='center'); c.border = tb
    ws.column_dimensions['A'].width = 18
    for col in ws.columns:
        col[0].border = tb
    # --- 输出模板文件供用户下载 ---
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='资产导入模板.xlsx')


# ===================== 批量操作 =====================

@asset_bp.route('/batch-edit', methods=['POST'])
@login_required
def batch_edit():
    """批量修改：对选中资产统一修改指定字段"""
    # --- 权限校验：仅管理员可批量修改 ---
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    try:
        # --- 获取选中资产 ID 列表和待修改字段 ---
        ids = request.form.getlist('ids[]')
        if not ids:
            return jsonify({'ok': False, 'msg': '未选择资产'}), 400
        ids = [int(i) for i in ids]
        fields = json.loads(request.form.get('fields', '{}'))
        if not fields:
            return jsonify({'ok': False, 'msg': '未指定修改字段'}), 400

        # --- 可批量修改的字段白名单 ---
        allowed = {'department', 'building', 'floor', 'location', 'status', 'category',
                   'device_type', 'brand', 'model_no', 'notes', 'vendor',
                   'ip_address', 'mac_address', 'cpu', 'memory', 'disk_size',
                   'operating_system', 'financial_code', 'financial_name'}
        # --- 过滤出白名单内的字段 ---
        update_data = {k: v for k, v in fields.items() if k in allowed and v is not None}

        count = 0
        # --- 遍历选中资产，逐条应用修改 ---
        for asset_id in ids:
            asset = Asset.query.filter(Asset.id == asset_id, Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
            if not asset:
                continue
            old = _asset_to_dict(asset)
            changed = False
            for k, v in update_data.items():
                if str(getattr(asset, k, '') or '') != str(v or ''):
                    setattr(asset, k, v)
                    changed = True
            if changed:
                # --- 记录每条资产的变更日志 ---
                _log_change(asset.id, 'batch_edit', old, _asset_to_dict(asset), current_user.display_name)
                count += 1
        db.session.commit()
        return jsonify({'ok': True, 'count': count, 'msg': f'已修改 {count} 条资产'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@asset_bp.route('/batch-transfer', methods=['POST'])
@login_required
def batch_transfer():
    """批量调拨：变更资产所属科室"""
    # --- 权限校验：仅管理员可调拨 ---
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    try:
        # --- 获取选中资产和目标科室 ---
        ids = request.form.getlist('ids[]')
        new_department = request.form.get('department', '').strip()
        if not ids:
            return jsonify({'ok': False, 'msg': '未选择资产'}), 400
        if not new_department:
            return jsonify({'ok': False, 'msg': '未指定目标科室'}), 400

        count = 0
        # --- 遍历资产，变更科室并记录日志 ---
        for asset_id in ids:
            asset = Asset.query.filter(Asset.id == asset_id, Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
            if not asset:
                continue
            # --- 跳过科室相同的资产 ---
            if asset.department == new_department:
                continue
            old = _asset_to_dict(asset)
            asset.department = new_department
            _log_change(asset.id, 'transfer', old, _asset_to_dict(asset), current_user.display_name)
            count += 1
        db.session.commit()
        return jsonify({'ok': True, 'count': count, 'msg': f'已调拨 {count} 条资产至 {new_department}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@asset_bp.route('/batch-recover', methods=['POST'])
@login_required
def batch_recover():
    """资产回收：批量改为报废状态"""
    # --- 权限校验：仅管理员可回收 ---
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    try:
        # --- 获取选中资产 ID 列表 ---
        ids = request.form.getlist('ids[]')
        if not ids:
            return jsonify({'ok': False, 'msg': '未选择资产'}), 400

        count = 0
        # --- 遍历资产，将状态改为已报废 ---
        for asset_id in ids:
            asset = Asset.query.filter(Asset.id == asset_id, Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
            if not asset:
                continue
            # --- 跳过已报废资产 ---
            if asset.status == 'scrapped':
                continue
            old = _asset_to_dict(asset)
            asset.status = 'scrapped'
            _log_change(asset.id, 'recover', old, _asset_to_dict(asset), current_user.display_name)
            count += 1
        db.session.commit()
        return jsonify({'ok': True, 'count': count, 'msg': f'已回收 {count} 条资产'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@asset_bp.route('/batch-relocate', methods=['POST'])
@login_required
def batch_relocate():
    """资产移位：批量变更楼栋/楼层/存放位置"""
    # --- 权限校验：仅管理员可移位 ---
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    try:
        # --- 获取选中资产和新位置信息 ---
        ids = request.form.getlist('ids[]')
        new_building = request.form.get('building', '').strip()
        new_floor = request.form.get('floor', '').strip()
        new_location = request.form.get('location', '').strip()

        if not ids:
            return jsonify({'ok': False, 'msg': '未选择资产'}), 400
        if not any([new_building, new_floor, new_location]):
            return jsonify({'ok': False, 'msg': '请至少填写一项变更内容'}), 400

        count = 0
        # --- 遍历资产，逐一更新位置字段 ---
        for asset_id in ids:
            asset = Asset.query.filter(Asset.id == asset_id, Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
            if not asset:
                continue
            old = _asset_to_dict(asset)
            changed = False
            if new_building and asset.building != new_building:
                asset.building = new_building
                changed = True
            if new_floor and asset.floor != new_floor:
                asset.floor = new_floor
                changed = True
            if new_location and asset.location != new_location:
                asset.location = new_location
                changed = True
            if changed:
                _log_change(asset.id, 'relocate', old, _asset_to_dict(asset), current_user.display_name)
                count += 1
        db.session.commit()
        return jsonify({'ok': True, 'count': count, 'msg': f'已移位 {count} 条资产'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500


# ===================== 位置互换 =====================

@asset_bp.route('/swap-location', methods=['POST'])
@login_required
def swap_location():
    """位置互换：两个资产互换位置信息"""
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    try:
        ids = request.form.getlist('ids[]')
        if len(ids) != 2:
            return jsonify({'ok': False, 'msg': '请选择恰好2个资产进行位置互换'}), 400

        a1 = Asset.query.get(ids[0])
        a2 = Asset.query.get(ids[1])
        if not a1 or not a2:
            return jsonify({'ok': False, 'msg': '资产不存在'}), 400

        old1 = _asset_to_dict(a1)
        old2 = _asset_to_dict(a2)

        # 互换位置相关字段
        a1.building, a2.building = a2.building, a1.building
        a1.floor, a2.floor = a2.floor, a1.floor
        a1.location, a2.location = a2.location, a1.location
        a1.department, a2.department = a2.department, a1.department

        _log_change(a1.id, 'swap', old1, _asset_to_dict(a1), current_user.display_name)
        _log_change(a2.id, 'swap', old2, _asset_to_dict(a2), current_user.display_name)
        db.session.commit()
        return jsonify({'ok': True, 'msg': f'已互换「{a1.asset_no}」与「{a2.asset_no}」的位置'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500


# ===================== 资产借用 =====================

@asset_bp.route('/batch-borrow', methods=['POST'])
@login_required
def batch_borrow():
    """资产借用：标记资产为借出状态"""
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    try:
        ids = request.form.getlist('ids[]')
        borrower = request.form.get('borrowed_to', '').strip()
        if not ids:
            return jsonify({'ok': False, 'msg': '未选择资产'}), 400
        if not borrower:
            return jsonify({'ok': False, 'msg': '请输入借用人'}), 400

        from datetime import date
        bdate_str = request.form.get('borrow_date', '').strip()
        rdate_str = request.form.get('return_date', '').strip()
        note = request.form.get('borrow_note', '').strip()

        bdate = date.fromisoformat(bdate_str) if bdate_str else date.today()
        rdate = date.fromisoformat(rdate_str) if rdate_str else None

        count = 0
        for aid in ids:
            asset = Asset.query.filter(Asset.id == aid, Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
            if not asset or asset.borrow_status == 'borrowed':
                continue
            old = _asset_to_dict(asset)
            asset.borrow_status = 'borrowed'
            asset.borrowed_to = borrower
            asset.borrow_date = bdate
            asset.return_date = rdate
            asset.borrow_note = note
            _log_change(asset.id, 'borrow', old, _asset_to_dict(asset), current_user.display_name)
            count += 1
        db.session.commit()
        return jsonify({'ok': True, 'count': count, 'msg': f'已借出 {count} 条资产给 {borrower}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@asset_bp.route('/batch-return', methods=['POST'])
@login_required
def batch_return():
    """资产归还：取消借用状态"""
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    try:
        ids = request.form.getlist('ids[]')
        if not ids:
            return jsonify({'ok': False, 'msg': '未选择资产'}), 400

        count = 0
        for aid in ids:
            asset = Asset.query.filter(Asset.id == aid, Asset.hospital_id == getattr(g, 'hospital_id', None)).first()
            if not asset or asset.borrow_status != 'borrowed':
                continue
            old = _asset_to_dict(asset)
            asset.borrow_status = ''
            asset.borrowed_to = ''
            asset.borrow_date = None
            asset.return_date = None
            asset.borrow_note = ''
            _log_change(asset.id, 'return', old, _asset_to_dict(asset), current_user.display_name)
            count += 1
        db.session.commit()
        return jsonify({'ok': True, 'count': count, 'msg': f'已归还 {count} 条资产'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500


# ===================== 库存概览 =====================

@asset_bp.route('/inventory-stats')
@login_required
def inventory_stats():
    """资产库存概览JSON"""
    try:
        total = Asset.query.count()
        stats = {
            'total': total,
            'in_use': Asset.query.filter_by(status='in_use').count(),
            'spare': Asset.query.filter_by(status='spare').count(),
            'scrapped': Asset.query.filter_by(status='scrapped').count(),
            'lost': Asset.query.filter_by(status='lost').count(),
            'borrowed': Asset.query.filter(Asset.borrow_status == 'borrowed').count(),
            'categories': {},
            'departments': {},
        }
        # 按分类统计
        for row in db.session.query(Asset.category, db.func.count(Asset.id)).group_by(Asset.category).all():
            if row[0]:
                stats['categories'][row[0]] = row[1]
        # 按科室统计
        for row in db.session.query(Asset.department, db.func.count(Asset.id)).group_by(Asset.department).all():
            if row[0]:
                stats['departments'][row[0]] = row[1]
        return jsonify({'ok': True, 'stats': stats})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500


# ===================== 操作日志 =====================

@asset_bp.route('/logs')
@login_required
def logs():
    """操作日志列表"""
    # --- 获取分页和筛选参数 ---
    page = request.args.get('page', 1, type=int)
    asset_id = request.args.get('asset_id', type=int)
    action = request.args.get('action', '')

    # --- 构建日志查询，可选按资产或操作类型筛选 ---
    query = AssetLog.query
    if asset_id:
        query = query.filter(AssetLog.asset_id == asset_id)
    if action:
        query = query.filter(AssetLog.action == action)

    # --- 按时间降序分页返回 ---
    pagination = query.order_by(AssetLog.created_at.desc()).paginate(
        page=page, per_page=30, error_out=False
    )
    import json as _json
    # 不修改模型实例（会触发 autoflush），创建独立解析列表
    parsed_logs = []
    for log in pagination.items:
        old_val = {}
        new_val = {}
        if isinstance(log.old_value, str) and log.old_value:
            try:
                old_val = _json.loads(log.old_value)
            except:
                old_val = {}
        if isinstance(log.new_value, str) and log.new_value:
            try:
                new_val = _json.loads(log.new_value)
            except:
                new_val = {}
        parsed_logs.append({
            'id': log.id,
            'created_at': log.created_at,
            'asset_id': log.asset_id,
            'action': log.action,
            'operator': log.operator,
            'old_value': old_val,
            'new_value': new_val,
        })
    return render_template('asset/logs.html', pagination=pagination,
                           logs=parsed_logs, asset_id=asset_id, action=action)


@asset_bp.route('/logs/<int:log_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_log(log_id):
    """删除操作日志"""
    log = AssetLog.query.get_or_404(log_id)
    db.session.delete(log)
    db.session.commit()
    flash('已删除操作日志', 'success')
    return redirect(url_for('asset.logs'))


@asset_bp.route('/recent-logs')
@login_required
def recent_logs():
    """返回最近操作记录 JSON"""
    try:
        action = request.args.get('action', '')
        query = AssetLog.query.order_by(AssetLog.created_at.desc()).limit(50)
        if action:
            query = AssetLog.query.filter(AssetLog.action == action).order_by(AssetLog.created_at.desc()).limit(50)
        logs = []
        for log in query.all():
            asset = db.session.get(Asset, log.asset_id)
            logs.append({
                'id': log.id,
                'asset_id': log.asset_id,
                'asset_no': asset.asset_no if asset else '已删除',
                'action': log.action,
                'operator': log.operator,
                'created_at': log.created_at.strftime('%m-%d %H:%M') if log.created_at else '',
            })
        return jsonify({'ok': True, 'logs': logs})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500


@asset_bp.route('/api/stats')
@login_required
def api_stats():
    # --- 统计资产总数 ---
    total = Asset.query.count()
    # --- 统计即将到期、已过期、保修有效数量 ---
    expiring = sum(1 for a in Asset.query.all() if a.warranty_status == 'expiring')
    expired = sum(1 for a in Asset.query.all() if a.warranty_status == 'expired')
    valid = sum(1 for a in Asset.query.all() if a.warranty_status == 'valid')
    return jsonify({'total': total, 'expiring': expiring, 'expired': expired, 'valid': valid})


@asset_bp.route('/api/grouped')
@login_required
def api_grouped():
    """按楼区或科室分组的资产数据 JSON"""
    group_by = request.args.get('group_by', 'building')
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')
    cat_filter = request.args.get('cat', '')

    query = Asset.query
    if search:
        safe_search = search.replace('%', '\\%').replace('_', '\\_')
        like = f'%{safe_search}%'
        query = query.filter(db.or_(
            Asset.asset_no.ilike(like), Asset.brand.ilike(like),
            Asset.model_no.ilike(like), Asset.sn.ilike(like),
            Asset.department.ilike(like), Asset.device_type.ilike(like),
        ))
    if status_filter:
        query = query.filter(Asset.status == status_filter)
    if cat_filter:
        query = query.filter(Asset.category == cat_filter)

    assets = query.order_by(Asset.updated_at.desc()).all()

    groups = {}
    for a in assets:
        key = getattr(a, group_by, None) or '未知'
        if key not in groups:
            groups[key] = {'name': key, 'count': 0, 'assets': []}
        groups[key]['count'] += 1
        groups[key]['assets'].append({
            'id': a.id,
            'asset_no': a.asset_no,
            'device_type': a.device_type,
            'brand': a.brand,
            'model_no': a.model_no,
            'sn': a.sn,
            'department': a.department,
            'building': a.building,
            'floor': a.floor,
            'location': a.location,
            'status': a.status,
            'ip_address': a.ip_address,
        })

    sorted_groups = sorted(groups.values(), key=lambda x: x['name'])
    return jsonify({'ok': True, 'groups': sorted_groups, 'group_by': group_by, 'total': len(assets)})


# ===================== 地址池视图 API =====================

@asset_bp.route('/api/pool-tree')
@login_required
def api_pool_tree():
    """返回楼区→楼层层级树+各层资产数"""
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')

    query = Asset.query
    if search:
        safe_search = search.replace('%', '\\%').replace('_', '\\_')
        like = f'%{safe_search}%'
        query = query.filter(db.or_(
            Asset.asset_no.ilike(like), Asset.brand.ilike(like),
            Asset.model_no.ilike(like), Asset.sn.ilike(like),
            Asset.department.ilike(like), Asset.ip_address.ilike(like),
        ))
    if status_filter:
        query = query.filter(Asset.status == status_filter)

    from sqlalchemy import func
    rows = query.with_entities(Asset.building, Asset.floor, func.count(Asset.id)).\
        filter(Asset.building.isnot(None), Asset.building != '').\
        group_by(Asset.building, Asset.floor).\
        order_by(Asset.building, Asset.floor).all()

    tree = {}
    for building, floor, cnt in rows:
        if building not in tree:
            tree[building] = {'name': building, 'total': 0, 'floors': {}}
        f = floor or '未知'
        tree[building]['floors'][f] = tree[building]['floors'].get(f, 0) + cnt
        tree[building]['total'] += cnt

    result = []
    for bname, bdata in sorted(tree.items()):
        floors_list = [{'name': f, 'count': c}
                       for f, c in sorted(bdata['floors'].items(),
                                          key=lambda x: (len(x[0]), x[0]))]
        result.append({
            'name': bname,
            'total': bdata['total'],
            'floors': floors_list,
        })

    no_addr = query.filter(db.or_(Asset.building.is_(None), Asset.building == '')).count()
    return jsonify({'ok': True, 'tree': result, 'no_address': no_addr, 'total': len(result)})


@asset_bp.route('/api/pool-assets')
@login_required
def api_pool_assets():
    """按楼区+楼层+设备类型获取资产列表"""
    building = request.args.get('building', '').strip()
    floor = request.args.get('floor', '').strip()
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')
    device_type = request.args.get('device_type', '').strip()

    query = Asset.query
    if building:
        query = query.filter(Asset.building == building)
    if floor:
        query = query.filter(Asset.floor == floor)
    if device_type:
        query = query.filter(Asset.device_type == device_type)

    if search:
        safe_search = search.replace('%', '\\%').replace('_', '\\_')
        like = f'%{safe_search}%'
        query = query.filter(db.or_(
            Asset.asset_no.ilike(like), Asset.brand.ilike(like),
            Asset.model_no.ilike(like), Asset.sn.ilike(like),
            Asset.department.ilike(like), Asset.ip_address.ilike(like),
        ))
    if status_filter:
        query = query.filter(Asset.status == status_filter)

    assets = query.order_by(Asset.asset_no).all()
    items = []
    for a in assets:
        items.append({
            'id': a.id,
            'asset_no': a.asset_no,
            'device_type': a.device_type,
            'brand': a.brand,
            'model_no': a.model_no,
            'sn': a.sn,
            'department': a.department,
            'location': a.location,
            'status': a.status,
            'ip_address': a.ip_address,
            'category': a.category,
        })
    return jsonify({'ok': True, 'assets': items, 'count': len(items)})


# ===================== 辅助函数 =====================

def _asset_to_dict(asset):
    """将 Asset 实例转为可 JSON 序列化的 dict"""
    d = {}
    # --- 遍历所有列，将日期时间格式化为 ISO 字符串 ---
    for col in asset.__table__.columns:
        val = getattr(asset, col.name)
        if isinstance(val, (datetime, date)):
            val = val.isoformat()
        d[col.name] = val
    return d


def _log_change(asset_id, action, old, new, operator):
    """记录资产变更日志"""
    import json
    # --- 构建变更日志记录，包含操作前后的差异 ---
    log = AssetLog(
        asset_id=asset_id, action=action,
        old_value=json.dumps(old, ensure_ascii=False) if old else None,
        new_value=json.dumps(new, ensure_ascii=False) if new else None,
        operator=operator
    )
    db.session.add(log)


def _parse_date(val):
    # --- 空值返回 None ---
    if not val:
        return None
    try:
        # --- 解析 YYYY-MM-DD 格式日期 ---
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _parse_float(val):
    # --- 空值返回 None ---
    if not val:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_int(val):
    # --- 空值返回 None ---
    if not val:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


# 确保 WorkOrder 在 detail 中可用
from models import WorkOrder
