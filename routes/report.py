"""月度报表导出"""
import io
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, send_file, jsonify
from flask_login import login_required, current_user
from sqlalchemy import func, case
from models import AuditLog, db, WorkOrder, ConsumableRecord, StockRecord, Consumable, SparePart, RepairOrder, Asset, AssetLog, FormTemplate
from models import log_audit

report_bp = Blueprint('report', __name__, url_prefix='/report')

# ======================== 数据查询函数 ========================

def _get_report_data(year, month):
    """获取完整报表数据。month=0 表示全年"""
    is_year = (month == 0)
    if is_year:
        first = datetime(year, 1, 1)
        last = datetime(year + 1, 1, 1)
    else:
        first = datetime(year, month, 1)
        if month == 12:
            last = datetime(year + 1, 1, 1)
        else:
            last = datetime(year, month + 1, 1)

    # --- Sheet 1: 月度概览 ---
    total = WorkOrder.query.filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
    ).count()

    completed = WorkOrder.query.filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.status == 'completed',
    ).count()

    pending = WorkOrder.query.filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.status == 'pending',
    ).count()

    in_progress = WorkOrder.query.filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.status == 'in_progress',
    ).count()

    # 平均响应时长（接单）- 小时
    avg_response = WorkOrder.query.with_entities(
        func.avg(
            func.julianday(WorkOrder.accepted_at) -
            func.julianday(WorkOrder.created_at)
        ) * 24
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.accepted_at.isnot(None),
    ).scalar() or 0

    # 平均完成时长 - 小时
    avg_complete = WorkOrder.query.with_entities(
        func.avg(
            func.julianday(WorkOrder.completed_at) -
            func.julianday(WorkOrder.created_at)
        ) * 24
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.completed_at.isnot(None),
    ).scalar() or 0

    overview = {
        'year': year, 'month': month,
        'label': f'{year}年全年' if is_year else f'{year}年{month}月',
        'total': total,
        'completed': completed,
        'in_progress': in_progress,
        'pending': pending,
        'complete_rate': round(completed / total * 100, 1) if total else 0,
        'avg_response': round(avg_response, 1),
        'avg_complete': round(avg_complete, 1),
    }

    # --- Sheet 2: 各周统计（年模式=各月统计） ---
    weekly = []
    if is_year:
        for m in range(1, 13):
            ms = datetime(year, m, 1)
            me = datetime(year + 1, 1, 1) if m == 12 else datetime(year, m + 1, 1)
            wk_total = WorkOrder.query.filter(
                WorkOrder.created_at >= ms,
                WorkOrder.created_at < me,
                WorkOrder.person != 'admin',
            ).count()
            wk_done = WorkOrder.query.filter(
                WorkOrder.created_at >= ms,
                WorkOrder.created_at < me,
                WorkOrder.person != 'admin',
                WorkOrder.status == 'completed',
            ).count()
            weekly.append({
                'week': f'{m}月',
                'total': wk_total,
                'completed': wk_done,
            })
    else:
        for week in range(1, 6):
            ws_start = first + timedelta(weeks=week - 1)
            ws_end = min(first + timedelta(weeks=week), last)
            if ws_start >= last:
                break
            wk_total = WorkOrder.query.filter(
                WorkOrder.created_at >= ws_start,
                WorkOrder.created_at < ws_end,
                WorkOrder.person != 'admin',
            ).count()
            wk_done = WorkOrder.query.filter(
                WorkOrder.created_at >= ws_start,
                WorkOrder.created_at < ws_end,
                WorkOrder.person != 'admin',
                WorkOrder.status == 'completed',
            ).count()
            weekly.append({
                'week': f'第{week}周',
                'total': wk_total,
                'completed': wk_done,
            })

    # --- Sheet 3: 人员排行 ---
    person_rows = WorkOrder.query.with_entities(
        WorkOrder.person, func.count(WorkOrder.id)
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != '',
        WorkOrder.person != 'admin',
        WorkOrder.person.isnot(None),
    ).group_by(WorkOrder.person).order_by(
        func.count(WorkOrder.id).desc()
    ).all()

    person_stats = []
    for i, (name, cnt) in enumerate(person_rows, 1):
        person_stats.append({
            'rank': i, 'name': name, 'count': cnt,
            'pct': round(cnt / total * 100, 1) if total else 0,
        })

    # --- Sheet 4: 故障类型分布 ---
    fault_rows = WorkOrder.query.with_entities(
        WorkOrder.fault_type, func.count(WorkOrder.id)
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.fault_type.isnot(None),
        WorkOrder.fault_type != '',
    ).group_by(WorkOrder.fault_type).order_by(
        func.count(WorkOrder.id).desc()
    ).all()

    fault_stats = []
    for i, (ft, cnt) in enumerate(fault_rows, 1):
        fault_stats.append({
            'rank': i, 'type': ft, 'count': cnt,
            'pct': round(cnt / total * 100, 1) if total else 0,
        })

    # --- Sheet 5: 科室排行 ---
    dept_rows = WorkOrder.query.with_entities(
        WorkOrder.department, func.count(WorkOrder.id)
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.department.isnot(None),
        WorkOrder.department != '',
    ).group_by(WorkOrder.department).order_by(
        func.count(WorkOrder.id).desc()
    ).limit(20).all()

    dept_stats = []
    for i, (dept, cnt) in enumerate(dept_rows, 1):
        dept_stats.append({
            'rank': i, 'department': dept, 'count': cnt,
            'pct': round(cnt / total * 100, 1) if total else 0,
        })

    # --- Sheet 6: 响应趋势（近14天） ---
    now = datetime.now()
    trend = []
    for i in range(13, -1, -1):
        day = now - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)

        day_orders = WorkOrder.query.with_entities(
            WorkOrder.created_at, WorkOrder.completed_at
        ).filter(
            WorkOrder.created_at >= day_start,
            WorkOrder.created_at < day_end,
            WorkOrder.person != 'admin',
        ).all()

        cnt = len(day_orders)
        durations = []
        for o in day_orders:
            if o.completed_at:
                minutes = (o.completed_at - o.created_at).total_seconds() / 60
                durations.append(min(minutes, 60))

        avg_val = round(sum(durations) / len(durations), 1) if durations else None
        trend.append({
            'day': day.strftime('%m/%d'),
            'count': cnt,
            'avg_minutes': avg_val,
        })

    # --- Sheet 7: 楼区分布 ---
    bld_rows = WorkOrder.query.with_entities(
        WorkOrder.building, func.count(WorkOrder.id)
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.building.isnot(None),
        WorkOrder.building != '',
    ).group_by(WorkOrder.building).order_by(
        func.count(WorkOrder.id).desc()
    ).all()

    building_stats = []
    for i, (bld, cnt) in enumerate(bld_rows, 1):
        building_stats.append({
            'rank': i, 'building': bld, 'count': cnt,
            'pct': round(cnt / total * 100, 1) if total else 0,
        })

    # --- Sheet 8: SLA 时效分析（人均响应/解决/超时） ---
    all_orders = WorkOrder.query.filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
    ).all()
    person_sla = {}
    for o in all_orders:
        p = o.person
        if not p or p == 'admin':
            continue
        if p not in person_sla:
            person_sla[p] = {'total': 0, 'resp_sum': 0.0, 'resp_n': 0,
                             'resol_sum': 0.0, 'resol_n': 0, 'overdue': 0}
        person_sla[p]['total'] += 1
        if o.accepted_at and o.created_at:
            mins = abs((o.accepted_at - o.created_at).total_seconds() / 60)
            person_sla[p]['resp_sum'] += mins
            person_sla[p]['resp_n'] += 1
        if o.completed_at and o.created_at:
            mins = abs((o.completed_at - o.created_at).total_seconds() / 60)
            person_sla[p]['resol_sum'] += mins
            person_sla[p]['resol_n'] += 1
        # 是否超时（简单：完成时间 > 阈值）
        if o.status == 'completed' and o.completed_at and o.created_at:
            cost_h = (o.completed_at - o.created_at).total_seconds() / 3600
            th = {'emergency': 2, 'urgent': 8, 'normal': 24}.get(o.priority, 24)
            if cost_h > th:
                person_sla[p]['overdue'] += 1
    sla_analysis = []
    for p, d in sorted(person_sla.items(),
                       key=lambda x: x[1].get('resp_sum', 0) / max(x[1].get('resp_n', 1), 1)):
        avg_resp = round(d['resp_sum'] / d['resp_n'], 1) if d['resp_n'] else None
        avg_resol = round(d['resol_sum'] / d['resol_n'], 1) if d['resol_n'] else None
        if avg_resp is None:
            label = '暂无数据'
        elif avg_resp <= 30:
            label = '响应极速'
        elif avg_resp <= 50:
            label = '响应快速'
        elif avg_resp <= 100:
            label = '响应较快'
        elif avg_resp <= 150:
            label = '响应较慢'
        else:
            label = '响应一般'
        sla_analysis.append({
            'name': p, 'total': d['total'],
            'avg_resp': avg_resp if avg_resp else 0,
            'avg_resol': avg_resol if avg_resol else 0,
            'overdue': d['overdue'],
            'sla_label': label,
        })
    # 按响应时长排序
    sla_analysis.sort(key=lambda x: x['avg_resp'])

    # --- Sheet 9: 紧急程度分布 ---
    priority_rows = WorkOrder.query.with_entities(
        WorkOrder.priority, func.count(WorkOrder.id),
        func.sum(case((WorkOrder.status == 'completed', 1), else_=0))
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
    ).group_by(WorkOrder.priority).all()
    priority_map = {'normal': '普通', 'urgent': '加急', 'emergency': '紧急'}
    priority_stats = []
    for pr, cnt, done in priority_rows:
        label = priority_map.get(pr, pr)
        # 平均响应
        avg_r = WorkOrder.query.with_entities(func.avg(
            func.julianday(WorkOrder.accepted_at) - func.julianday(WorkOrder.created_at)
        ) * 24).filter(
            WorkOrder.created_at >= first,
            WorkOrder.created_at < last,
            WorkOrder.person != 'admin',
            WorkOrder.priority == pr,
            WorkOrder.accepted_at.isnot(None),
        ).scalar() or 0
        priority_stats.append({
            'priority': label, 'count': cnt,
            'completed': done or 0,
            'rate': round((done or 0) / cnt * 100, 1) if cnt else 0,
            'avg_response': round(avg_r, 1),
        })

    # --- Sheet 10: 处理效率排行（按平均解决时长） ---
    efficiency_rank = []
    for p, d in sorted(person_sla.items(),
                       key=lambda x: x[1].get('resol_sum', 0) / max(x[1].get('resol_n', 1), 1)):
        avg_resol = round(d['resol_sum'] / d['resol_n'], 1) if d['resol_n'] else None
        if avg_resol is None:
            continue
        # 速度标签
        if avg_resol <= 60:
            label = '极速处理', '#10b981'
        elif avg_resol <= 180:
            label = '高效处理', '#059669'
        elif avg_resol <= 480:
            label = '正常处理', '#f59e0b'
        elif avg_resol <= 1440:
            label = '较慢处理', '#f97316'
        else:
            label = '低效处理', '#ef4444'
        efficiency_rank.append({
            'name': p, 'count': d['total'],
            'avg_resol': avg_resol,
            'hours': round(avg_resol / 60, 1),
            'label': label[0],
            'color': label[1],
        })

    # --- Sheet 11: 时段分布 ---
    hour_buckets = {h: 0 for h in range(24)}
    dow_buckets = {d: 0 for d in ['周一','周二','周三','周四','周五','周六','周日']}
    for o in all_orders:
        h = o.created_at.hour
        hour_buckets[h] = hour_buckets.get(h, 0) + 1
        wd = o.created_at.weekday()  # 0=Mon
        dow_label = ['周一','周二','周三','周四','周五','周六','周日'][wd]
        dow_buckets[dow_label] = dow_buckets.get(dow_label, 0) + 1
    hour_dist = [{'hour': f'{h:02d}:00', 'count': hour_buckets[h]} for h in range(24)]
    dow_dist = [{'day': d, 'count': dow_buckets[d]} for d in ['周一','周二','周三','周四','周五','周六','周日']]

    # --- Sheet 12: 耗材出入库工作报表 ---
    consumable_records = []
    rows = db.session.query(
        ConsumableRecord, Consumable.name, Consumable.spec
    ).join(Consumable, ConsumableRecord.consumable_id == Consumable.id
    ).filter(
        ConsumableRecord.created_at >= first,
        ConsumableRecord.created_at < last,
    ).order_by(ConsumableRecord.created_at.desc()).all()
    for r, cname, cspec in rows:
        consumable_records.append({
            'time': r.created_at.strftime('%Y-%m-%d %H:%M'),
            'name': cname,
            'spec': cspec,
            'type': '入库' if r.type == 'in' else '出库',
            'quantity': r.quantity,
            'balance': r.balance,
            'operator': r.operator,
            'note': r.note,
        })
    # 统计
    con_in = sum(1 for r in rows if r.ConsumableRecord.type == 'in')
    con_out = sum(1 for r in rows if r.ConsumableRecord.type == 'out')
    con_qty_in = sum(r.ConsumableRecord.quantity for r in rows if r.ConsumableRecord.type == 'in')
    con_qty_out = sum(r.ConsumableRecord.quantity for r in rows if r.ConsumableRecord.type == 'out')

    # --- Sheet 10: 备件出入库工作报表 ---
    spare_records = []
    srows = db.session.query(
        StockRecord, SparePart.name, SparePart.model_no
    ).join(SparePart, StockRecord.part_id == SparePart.id
    ).filter(
        StockRecord.created_at >= first,
        StockRecord.created_at < last,
    ).order_by(StockRecord.created_at.desc()).all()
    for r, sname, smodel in srows:
        spare_records.append({
            'time': r.created_at.strftime('%Y-%m-%d %H:%M'),
            'name': sname,
            'spec': smodel,
            'type': '入库' if r.type == 'in' else '出库',
            'quantity': r.quantity,
            'balance': r.balance,
            'operator': r.operator,
            'note': r.note,
        })
    sp_in = sum(1 for r in srows if r.StockRecord.type == 'in')
    sp_out = sum(1 for r in srows if r.StockRecord.type == 'out')
    sp_qty_in = sum(r.StockRecord.quantity for r in srows if r.StockRecord.type == 'in')
    sp_qty_out = sum(r.StockRecord.quantity for r in srows if r.StockRecord.type == 'out')

    # --- Sheet 13: 维修单统计报告 ---
    repair_all = RepairOrder.query.filter(
        RepairOrder.created_at >= first,
        RepairOrder.created_at < last,
    ).all()
    r_total = len(repair_all)
    r_draft = sum(1 for r in repair_all if r.status == 'draft')
    r_pending = sum(1 for r in repair_all if r.status == 'pending')
    r_approved = sum(1 for r in repair_all if r.status == 'approved')
    r_rejected = sum(1 for r in repair_all if r.status == 'rejected')
    r_approval_rate = round(r_approved / (r_approved + r_rejected) * 100, 1) if (r_approved + r_rejected) else 0

    # 维修单模板分布 + 审批时效
    r_template_rows = db.session.query(
        FormTemplate.name, func.count(RepairOrder.id)
    ).join(FormTemplate, RepairOrder.template_id == FormTemplate.id
    ).filter(
        RepairOrder.created_at >= first,
        RepairOrder.created_at < last,
    ).group_by(FormTemplate.name).order_by(
        func.count(RepairOrder.id).desc()
    ).limit(8).all()
    r_template_stats = [{'name': name or '未知模板', 'count': cnt} for name, cnt in r_template_rows]

    # 审批时效统计
    r_approval_hours = []
    for r in repair_all:
        if r.approved_at and r.created_at:
            h = (r.approved_at - r.created_at).total_seconds() / 3600
            r_approval_hours.append(h)
    r_avg_approval = round(sum(r_approval_hours) / len(r_approval_hours), 1) if r_approval_hours else 0
    r_quick = sum(1 for h in r_approval_hours if h <= 1)       # ≤1h
    r_normal = sum(1 for h in r_approval_hours if 1 < h <= 24)  # 1-24h
    r_slow = sum(1 for h in r_approval_hours if h > 24)         # >24h
    r_approval_dist = [
        {'label': '≤1小时', 'count': r_quick, 'color': '#10b981'},
        {'label': '1~24小时', 'count': r_normal, 'color': '#f59e0b'},
        {'label': '>24小时', 'count': r_slow, 'color': '#ef4444'},
    ]

    # 维修单月度趋势（按周）
    r_trend = []
    for week in range(1, 6):
        ws = first + timedelta(weeks=week - 1)
        we = min(first + timedelta(weeks=week), last)
        if ws >= last:
            break
        wc = sum(1 for r in repair_all if ws <= r.created_at < we)
        r_trend.append({'week': f'第{week}周', 'count': wc})

    repair_stats = {
        'total': r_total, 'draft': r_draft, 'pending': r_pending,
        'approved': r_approved, 'rejected': r_rejected,
        'approval_rate': r_approval_rate,
        'template_stats': r_template_stats,
        'trend': r_trend,
        'avg_approval': r_avg_approval,
        'approval_dist': r_approval_dist,
    }

    # --- Sheet 14: 资产更替报告 ---
    # 资产状态分布
    status_rows = db.session.query(
        Asset.status, func.count(Asset.id)
    ).group_by(Asset.status).all()
    status_map = {'in_use': '使用中', 'idle': '闲置', 'scrapped': '已报废', 'repair': '维修中', 'lost': '已丢失'}
    asset_status = [{'status': status_map.get(s, s), 'count': c, 'key': s} for s, c in status_rows]

    # 资产分类分布
    cat_rows = db.session.query(
        Asset.category, func.count(Asset.id)
    ).group_by(Asset.category).all()
    cat_map = {'hardware': '硬件设备', 'software': '软件授权'}
    asset_category = [{'category': cat_map.get(c, c), 'count': cnt} for c, cnt in cat_rows]

    # 设备类型分布（TOP 8）
    dev_rows = db.session.query(
        Asset.device_type, func.count(Asset.id)
    ).group_by(Asset.device_type).order_by(func.count(Asset.id).desc()).limit(8).all()
    asset_device_type = [{'type': dt, 'count': c} for dt, c in dev_rows]

    # 本月资产变更日志
    asset_logs = AssetLog.query.filter(
        AssetLog.created_at >= first,
        AssetLog.created_at < last,
    ).order_by(AssetLog.created_at.desc()).limit(20).all()
    action_map = {'import': '导入', 'edit': '编辑', 'transfer': '转移', 'relocate': '调拨', 'recover': '回收'}
    asset_log_list = [{
        'time': log.created_at.strftime('%m/%d %H:%M'),
        'action': action_map.get(log.action, log.action),
        'asset_no': log.asset.asset_no if log.asset else '-',
        'operator': log.operator,
    } for log in asset_logs]

    # 即将过保设备（30天内）
    from datetime import date
    today = date.today()
    expiring = Asset.query.filter(
        Asset.warranty_end >= today,
        Asset.warranty_end <= (today.replace(day=28) + timedelta(days=30) if today.day <= 28 else today.replace(month=today.month+1, day=1) + timedelta(days=30)),
    ).order_by(Asset.warranty_end.asc()).limit(10).all()
    # Simple approach: re-filter in Python
    expiring = Asset.query.filter(
        Asset.warranty_end.isnot(None),
        Asset.warranty_end >= today,
    ).all()
    expiring = [a for a in expiring if a.warranty_days_left is not None and a.warranty_days_left <= 30]
    expiring.sort(key=lambda a: a.warranty_days_left)
    expiring = expiring[:10]
    expiring_list = [{
        'asset_no': a.asset_no, 'device_type': a.device_type,
        'department': a.department, 'days_left': a.warranty_days_left,
    } for a in expiring]

    # 资产楼区分布
    bld_asset_rows = db.session.query(
        Asset.building, func.count(Asset.id)
    ).filter(
        Asset.building.isnot(None), Asset.building != '',
    ).group_by(Asset.building).order_by(func.count(Asset.id).desc()).all()
    asset_building = [{'building': b, 'count': c} for b, c in bld_asset_rows]

    # ----- 全年逐月故障类型分布（全年图表） -----
    # 取最近一整年（或选择年份的全年）数据
    chart_year_start = datetime(year, 1, 1)
    chart_year_end = datetime(year + 1, 1, 1)
    year_rows = db.session.query(
        func.strftime('%m', WorkOrder.created_at).label('mon'),
        WorkOrder.fault_type,
        func.count(WorkOrder.id).label('cnt'),
    ).filter(
        WorkOrder.created_at >= chart_year_start,
        WorkOrder.created_at < chart_year_end,
        WorkOrder.person != 'admin',
        WorkOrder.fault_type.isnot(None),
        WorkOrder.fault_type != '',
    ).group_by(
        'mon',
        WorkOrder.fault_type,
    ).order_by('mon', db.desc('cnt')).all()

    # 收集所有月份和故障类型
    all_types = list(dict.fromkeys(r.fault_type for r in year_rows))
    top_types = all_types[:8]  # 最多8种类型，其余归为"其他"
    chart_data = []
    for r in year_rows:
        if r.fault_type in top_types:
            chart_data.append(r)
        elif '其他' not in top_types:
            chart_data.append(type('row', (), {'mon': r.mon, 'fault_type': '其他', 'cnt': r.cnt})())
            top_types.append('其他')
        else:
            # 归入"其他"
            for cd in chart_data:
                if cd.mon == r.mon and cd.fault_type == '其他':
                    cd.cnt += r.cnt
                    break
            else:
                chart_data.append(type('row', (), {'mon': r.mon, 'fault_type': '其他', 'cnt': r.cnt})())

    # 构建图表数据集
    month_labels = [f'{m}月' for m in range(1, 13)]
    colors = ['#4f46e5','#0ea5e9','#10b981','#f59e0b','#ef4444','#8b5cf6','#ec4899','#14b8a6','#94a3b8']
    chart_types = [t for t in top_types if t != '其他'] + (['其他'] if any(c.fault_type == '其他' for c in chart_data) else [])
    datasets = []
    for i, ft in enumerate(chart_types):
        month_counts = {str(m).zfill(2): 0 for m in range(1, 13)}
        for r in chart_data:
            if r.fault_type == ft:
                month_counts[r.mon] = (month_counts.get(r.mon, 0) + r.cnt)
        datasets.append({
            'label': ft,
            'data': [month_counts[str(m).zfill(2)] for m in range(1, 13)],
            'backgroundColor': colors[i % len(colors)],
        })

    yearly_fault_chart = {
        'labels': month_labels,
        'datasets': datasets,
    }

    # ==================== 新增 10 项报表数据 ====================

    # 1. 📊 每人每月处理趋势（全年各月每人处理数）
    person_monthly = []
    if is_year or True:  # 全年模式下各月每人
        pm_rows = db.session.query(
            func.strftime('%m', WorkOrder.created_at).label('mon'),
            WorkOrder.person,
            func.count(WorkOrder.id).label('cnt'),
        ).filter(
            WorkOrder.created_at >= chart_year_start,
            WorkOrder.created_at < chart_year_end,
            WorkOrder.person != 'admin',
            WorkOrder.person != '',
            WorkOrder.person.isnot(None),
        ).group_by('mon', WorkOrder.person).order_by('mon').all()
        pm_persons = list(dict.fromkeys(r.person for r in pm_rows))
        pm_chart = {}
        for r in pm_rows:
            if r.person not in pm_chart:
                pm_chart[r.person] = {str(m).zfill(2): 0 for m in range(1, 13)}
            pm_chart[r.person][r.mon] = r.cnt
        person_monthly = {
            'labels': [f'{m}月' for m in range(1, 13)],
            'datasets': [{
                'label': p,
                'data': [pm_chart[p][str(m).zfill(2)] for m in range(1, 13)],
                'backgroundColor': colors[i % len(colors)],
            } for i, p in enumerate(pm_persons[:10])],
        }

    # 2. ⏱ 响应时长逐日趋势（本月每天平均响应分钟数）
    daily_response = []
    if not is_year:
        days_in_month = (last - first).days
        for d in range(days_in_month):
            ds = first + timedelta(days=d)
            de = ds + timedelta(days=1)
            row = WorkOrder.query.with_entities(
                func.avg(
                    (func.julianday(WorkOrder.accepted_at) - func.julianday(WorkOrder.created_at)) * 24 * 60
                )
            ).filter(
                WorkOrder.created_at >= ds,
                WorkOrder.created_at < de,
                WorkOrder.person != 'admin',
                WorkOrder.accepted_at.isnot(None),
            ).scalar()
            daily_response.append({
                'day': ds.strftime('%m/%d'),
                'avg_minutes': round(row, 1) if row else 0,
            })

    # 3. ✅ 首次响应达标率（30分钟内接单占比）
    resp_ok = WorkOrder.query.filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.accepted_at.isnot(None),
        (WorkOrder.accepted_at - WorkOrder.created_at) <= timedelta(minutes=30),
    ).count()
    resp_total = WorkOrder.query.filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.accepted_at.isnot(None),
    ).count()
    resp_compliance_rate = round(resp_ok / resp_total * 100, 1) if resp_total else 0
    # 按人统计达标率
    person_compliance = []
    for p_entry in person_stats:
        pname = p_entry['name']
        pok = WorkOrder.query.filter(
            WorkOrder.created_at >= first,
            WorkOrder.created_at < last,
            WorkOrder.person == pname,
            WorkOrder.accepted_at.isnot(None),
            (WorkOrder.accepted_at - WorkOrder.created_at) <= timedelta(minutes=30),
        ).count()
        ptotal = WorkOrder.query.filter(
            WorkOrder.created_at >= first,
            WorkOrder.created_at < last,
            WorkOrder.person == pname,
            WorkOrder.accepted_at.isnot(None),
        ).count()
        person_compliance.append({
            'name': pname,
            'count': ptotal,
            'ok': pok,
            'rate': round(pok / ptotal * 100, 1) if ptotal else 0,
        })

    # 4. ⚠️ 超时工单排行榜（处理时间最长的 TOP 10）
    longest_orders = WorkOrder.query.with_entities(
        WorkOrder.id, WorkOrder.title, WorkOrder.person, WorkOrder.created_at,
        WorkOrder.completed_at, WorkOrder.status,
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.completed_at.isnot(None),
    ).order_by(WorkOrder.completed_at.desc()).limit(50).all()
    # Python 侧计算耗时并排序
    overdue_raw = []
    for o in longest_orders:
        if o.completed_at and o.created_at:
            hours = (o.completed_at - o.created_at).total_seconds() / 3600
            overdue_raw.append((o, hours))
    overdue_raw.sort(key=lambda x: -x[1])
    overdue_list = []
    for o, hours in overdue_raw[:10]:
        overdue_list.append({
            'id': o.id,
            'title': o.title or '无标题',
            'person': o.person,
            'created_at': o.created_at.strftime('%m-%d %H:%M'),
            'hours': round(hours, 1),
            'status': o.status,
        })

    # 5. 🔁 重复报修分析（同一位置+同类型≥2次）
    repeat_query = db.session.query(
        WorkOrder.building, WorkOrder.floor, WorkOrder.location,
        WorkOrder.fault_type, WorkOrder.fault_subcategory,
        func.count(WorkOrder.id).label('cnt'),
        func.max(WorkOrder.created_at).label('last_time'),
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.building != '',
        WorkOrder.fault_type != '',
    ).group_by(
        WorkOrder.building, WorkOrder.floor, WorkOrder.location,
        WorkOrder.fault_type, WorkOrder.fault_subcategory,
    ).having(func.count(WorkOrder.id) >= 2).order_by(
        db.desc('cnt')
    ).limit(15).all()
    repeat_analysis = []
    for r in repeat_query:
        addr = '/'.join(filter(None, [r.building, r.floor, r.location]))
        repeat_analysis.append({
            'address': addr,
            'fault_type': r.fault_type,
            'subcategory': r.fault_subcategory or '',
            'count': r.cnt,
            'last_time': r.last_time.strftime('%m-%d') if r.last_time else '',
        })

    # 6. 🖥️ 设备故障排行
    device_fault_rows = WorkOrder.query.with_entities(
        WorkOrder.device_type, func.count(WorkOrder.id)
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.device_type.isnot(None),
        WorkOrder.device_type != '',
    ).group_by(WorkOrder.device_type).order_by(
        func.count(WorkOrder.id).desc()
    ).limit(12).all()
    device_fault_stats = [
        {'device_type': dt, 'count': cnt}
        for dt, cnt in device_fault_rows
    ]

    # 7. 🌙 非工作时间工单
    after_hours = WorkOrder.query.filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
    ).all()
    night_count = 0
    weekend_count = 0
    for o in after_hours:
        h = o.created_at.hour
        wd = o.created_at.weekday()
        if h < 8 or h >= 18:
            night_count += 1
        if wd >= 5:
            weekend_count += 1
    total_after_hours = len(after_hours)
    night_pct = round(night_count / total_after_hours * 100, 1) if total_after_hours else 0
    weekend_pct = round(weekend_count / total_after_hours * 100, 1) if total_after_hours else 0
    work_time_count = total_after_hours - night_count - weekend_count
    # 非工作时间按小时分布
    hour_after = {h: 0 for h in range(24)}
    for o in after_hours:
        hour_after[o.created_at.hour] += 1
    after_hours_dist = [{'hour': f'{h:02d}:00', 'count': hour_after[h]} for h in range(24)]

    # 8. 📈 工单状态耗时（各阶段平均时长）
    stage_times = {}
    # 创建→接单
    t1 = db.session.query(
        func.avg((WorkOrder.accepted_at - WorkOrder.created_at) * 24 * 60)
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.accepted_at.isnot(None),
    ).scalar() or 0
    stage_times['to_accept'] = round(float(t1), 1) if t1 else 0
    # 接单→完成
    t2 = db.session.query(
        func.avg((WorkOrder.completed_at - WorkOrder.accepted_at) * 24 * 60)
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
        WorkOrder.accepted_at.isnot(None),
        WorkOrder.completed_at.isnot(None),
    ).scalar() or 0
    stage_times['accept_to_complete'] = round(float(t2), 1) if t2 else 0
    # 创建→完成（总耗时）
    stage_times['total'] = round(stage_times['to_accept'] + stage_times['accept_to_complete'], 1)

    # 9. 🏆 人员工作量对比（已存在 person_stats，前端复用）

    # 10. 📋 本月工单明细清单
    order_list = WorkOrder.query.with_entities(
        WorkOrder.id, WorkOrder.title, WorkOrder.person, WorkOrder.status,
        WorkOrder.priority, WorkOrder.fault_type, WorkOrder.department, WorkOrder.created_at,
    ).filter(
        WorkOrder.created_at >= first,
        WorkOrder.created_at < last,
        WorkOrder.person != 'admin',
    ).order_by(WorkOrder.created_at.desc()).limit(50).all()
    priority_label = {'normal': '普通', 'urgent': '加急', 'emergency': '紧急'}
    status_label = {'pending': '待处理', 'in_progress': '处理中', 'completed': '已完成'}
    recent_orders = [{
        'id': o.id,
        'title': o.title or '无标题',
        'person': o.person,
        'status': status_label.get(o.status, o.status),
        'priority': priority_label.get(o.priority, o.priority),
        'fault_type': o.fault_type,
        'department': o.department,
        'created_at': o.created_at.strftime('%m-%d %H:%M') if o.created_at else '',
    } for o in order_list]

    stock_outbound = _get_stock_outbound_data(first, last)

    return {
        'overview': overview,
        'weekly': weekly,
        'person_stats': person_stats,
        'fault_stats': fault_stats,
        'dept_stats': dept_stats,
        'trend': trend,
        'building_stats': building_stats,
        'sla_analysis': sla_analysis,
        'priority_stats': priority_stats,
        'efficiency_rank': efficiency_rank,
        'hour_dist': hour_dist,
        'dow_dist': dow_dist,
        'consumable_records': consumable_records,
        'con_in': con_in, 'con_out': con_out,
        'con_qty_in': con_qty_in, 'con_qty_out': con_qty_out,
        'spare_records': spare_records,
        'sp_in': sp_in, 'sp_out': sp_out,
        'sp_qty_in': sp_qty_in, 'sp_qty_out': sp_qty_out,
        'repair_stats': repair_stats,
        'asset_status': asset_status,
        'asset_category': asset_category,
        'asset_device_type': asset_device_type,
        'asset_log_list': asset_log_list,
        'expiring_list': expiring_list,
        'asset_building': asset_building,
        'yearly_fault_chart': yearly_fault_chart,
        # 新增 10 项
        'person_monthly': person_monthly,
        'daily_response': daily_response,
        'resp_compliance_rate': resp_compliance_rate,
        'resp_ok': resp_ok, 'resp_total': resp_total,
        'person_compliance': person_compliance,
        'overdue_list': overdue_list,
        'repeat_analysis': repeat_analysis,
        'device_fault_stats': device_fault_stats,
        'night_count': night_count, 'weekend_count': weekend_count,
        'night_pct': night_pct, 'weekend_pct': weekend_pct,
        'work_time_count': work_time_count,
        'after_hours_dist': after_hours_dist,
        'stage_times': stage_times,
        'recent_orders': recent_orders,
        'stock_outbound': stock_outbound,
    }


def _get_stock_outbound_data(first, last):
    """获取月度备件出库详情"""
    records = StockRecord.query.filter(
        StockRecord.type == 'out',
        StockRecord.created_at >= first,
        StockRecord.created_at < last,
    ).order_by(StockRecord.created_at.desc()).all()
    total_items = sum(r.quantity for r in records)
    dept_stats = {}
    for r in records:
        dept = r.department or '未指定'
        if dept not in dept_stats:
            dept_stats[dept] = {'count': 0, 'items': 0}
        dept_stats[dept]['count'] += 1
        dept_stats[dept]['items'] += r.quantity
    dept_list = [{'department': k, 'count': v['count'], 'items': v['items']}
                 for k, v in sorted(dept_stats.items(), key=lambda x: -x[1]['items'])]
    return {
        'total_records': len(records),
        'total_items': total_items,
        'dept_stats': dept_list,
        'details': [{
            'time': r.created_at.strftime('%m-%d %H:%M'),
            'part': r.part.name if r.part else '(已删除)',
            'qty': r.quantity,
            'unit': r.part.unit if r.part else '',
            'department': r.department or '未指定',
            'operator': r.operator,
        } for r in records],
    }


# ======================== 路由 ========================

@report_bp.route('/')
@login_required
def report_page():
    """报表页面（在线预览）"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    if month < 0 or month > 12:
        month = datetime.now().month
    data = _get_report_data(year, month)
    return render_template('report/index.html',
                           data=data, now=datetime.now(),
                           selected_year=year, selected_month=month)


@report_bp.route('/data')
@login_required
def report_data():
    """JSON 报表数据接口"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    if month < 0 or month > 12:
        return jsonify({'error': '月份参数错误'}), 400
    data = _get_report_data(year, month)
    return jsonify(data)


@report_bp.route('/download')
@login_required
def download_report():
    """生成并下载月度报表 Excel（7个工作表）"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    if month < 0 or month > 12:
        return '月份参数错误', 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return '服务器未安装 openpyxl', 500

    data = _get_report_data(year, month)
    ov = data['overview']
    is_year = (month == 0)

    wb = Workbook()

    # 样式
    title_font = Font(name='微软雅黑', size=16, bold=True, color='1e293b')
    subtitle_font = Font(name='微软雅黑', size=14, bold=True, color='1e293b')
    header_fill = PatternFill('solid', fgColor='4f46e5')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='ffffff')
    body_font = Font(name='微软雅黑', size=11)
    bold_font = Font(name='微软雅黑', size=11, bold=True)
    note_font = Font(name='微软雅黑', size=10, color='94a3b8')
    thin = Side(style='thin', color='d0d5dd')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def style_header(ws, row, cols):
        for i in range(1, cols + 1):
            c = ws.cell(row=row, column=i)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')

    def apply_border(ws, max_row, max_col):
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    cell.border = border

    # ===== Sheet 1: 月度概览 =====
    ws1 = wb.active
    ws1.title = ov['label'].replace('年', '-').replace('月', '') + '报表'

    ws1.merge_cells('A1:D1')
    ws1['A1'].value = f'{ov["label"]} IT 运维工单报表'
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.row_dimensions[1].height = 36

    style_header(ws1, 3, 3)
    for i, h in enumerate(['指标', '数值', '说明'], 1):
        ws1.cell(row=3, column=i, value=h)
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 30

    rows_data = [
        ('工单总数', ov['total'], ''),
        ('已完成', ov['completed'], f"完成率 {ov['complete_rate']}%"),
        ('处理中', ov['in_progress'], ''),
        ('待接单', ov['pending'], ''),
        ('平均响应时长', f"{ov['avg_response']} 小时", '从创建到接单'),
        ('平均完成时长', f"{ov['avg_complete']} 小时", '从创建到完成'),
    ]
    for i, (label, val, note) in enumerate(rows_data, 4):
        ws1.cell(row=i, column=1, value=label).font = body_font
        ws1.cell(row=i, column=2, value=val).font = bold_font
        ws1.cell(row=i, column=3, value=note).font = note_font

    # ===== Sheet 2: 各周/各月统计 =====
    ws2 = wb.create_sheet('各月统计' if is_year else '各周统计')
    ws2.merge_cells('A1:C1')
    ws2['A1'].value = f'{ov["label"]} 工单趋势'
    ws2['A1'].font = subtitle_font

    style_header(ws2, 3, 3)
    for i, h in enumerate(['周次', '工单数', '完成数'], 1):
        ws2.cell(row=3, column=i, value=h)
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14

    for i, w in enumerate(data['weekly'], 4):
        ws2.cell(row=i, column=1, value=w['week']).font = body_font
        ws2.cell(row=i, column=2, value=w['total']).font = body_font
        ws2.cell(row=i, column=3, value=w['completed']).font = body_font

    # ===== Sheet 3: 人员排行 =====
    ws3 = wb.create_sheet('人员排行')
    ws3.merge_cells('A1:D1')
    ws3['A1'].value = f'{year}年{month}月 维护人员处理排行'
    ws3['A1'].font = subtitle_font

    style_header(ws3, 3, 4)
    for i, h in enumerate(['排名', '姓名', '完成数', '占比'], 1):
        ws3.cell(row=3, column=i, value=h)
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 14

    for i, p in enumerate(data['person_stats'], 4):
        ws3.cell(row=i, column=1, value=p['rank']).font = body_font
        ws3.cell(row=i, column=2, value=p['name']).font = body_font
        ws3.cell(row=i, column=3, value=p['count']).font = body_font
        ws3.cell(row=i, column=4, value=f"{p['pct']}%").font = body_font

    # ===== Sheet 4: 故障类型 =====
    ws4 = wb.create_sheet('故障类型')
    ws4.merge_cells('A1:D1')
    ws4['A1'].value = f'{year}年{month}月 故障类型分布'
    ws4['A1'].font = subtitle_font

    style_header(ws4, 3, 4)
    for i, h in enumerate(['排名', '故障类型', '数量', '占比'], 1):
        ws4.cell(row=3, column=i, value=h)
    ws4.column_dimensions['A'].width = 10
    ws4.column_dimensions['B'].width = 22
    ws4.column_dimensions['C'].width = 14
    ws4.column_dimensions['D'].width = 14

    for i, f in enumerate(data['fault_stats'], 4):
        ws4.cell(row=i, column=1, value=f['rank']).font = body_font
        ws4.cell(row=i, column=2, value=f['type']).font = body_font
        ws4.cell(row=i, column=3, value=f['count']).font = body_font
        ws4.cell(row=i, column=4, value=f"{f['pct']}%").font = body_font

    # ===== Sheet 5: 科室排行 =====
    ws5 = wb.create_sheet('科室排行')
    ws5.merge_cells('A1:D1')
    ws5['A1'].value = f'{year}年{month}月 报修科室排行'
    ws5['A1'].font = subtitle_font

    style_header(ws5, 3, 4)
    for i, h in enumerate(['排名', '科室', '工单数', '占比'], 1):
        ws5.cell(row=3, column=i, value=h)
    ws5.column_dimensions['A'].width = 10
    ws5.column_dimensions['B'].width = 26
    ws5.column_dimensions['C'].width = 14
    ws5.column_dimensions['D'].width = 14

    for i, d in enumerate(data['dept_stats'], 4):
        ws5.cell(row=i, column=1, value=d['rank']).font = body_font
        ws5.cell(row=i, column=2, value=d['department']).font = body_font
        ws5.cell(row=i, column=3, value=d['count']).font = body_font
        ws5.cell(row=i, column=4, value=f"{d['pct']}%").font = body_font

    # ===== Sheet 6: 响应趋势 =====
    ws6 = wb.create_sheet('响应趋势')
    ws6.merge_cells('A1:C1')
    ws6['A1'].value = '近14天响应时长趋势'
    ws6['A1'].font = subtitle_font

    style_header(ws6, 3, 3)
    for i, h in enumerate(['日期', '工单数', '平均响应(分钟)'], 1):
        ws6.cell(row=3, column=i, value=h)
    ws6.column_dimensions['A'].width = 14
    ws6.column_dimensions['B'].width = 14
    ws6.column_dimensions['C'].width = 20

    for i, t in enumerate(data['trend'], 4):
        ws6.cell(row=i, column=1, value=t['day']).font = body_font
        ws6.cell(row=i, column=2, value=t['count']).font = body_font
        ws6.cell(row=i, column=3, value=t['avg_minutes'] if t['avg_minutes'] else '-').font = body_font

    # ===== Sheet 7: 楼区分布 =====
    ws7 = wb.create_sheet('楼区分布')
    ws7.merge_cells('A1:D1')
    ws7['A1'].value = f'{year}年{month}月 楼区工单分布'
    ws7['A1'].font = subtitle_font

    style_header(ws7, 3, 4)
    for i, h in enumerate(['排名', '楼区', '工单数', '占比'], 1):
        ws7.cell(row=3, column=i, value=h)
    ws7.column_dimensions['A'].width = 10
    ws7.column_dimensions['B'].width = 22
    ws7.column_dimensions['C'].width = 14
    ws7.column_dimensions['D'].width = 14

    for i, b in enumerate(data['building_stats'], 4):
        ws7.cell(row=i, column=1, value=b['rank']).font = body_font
        ws7.cell(row=i, column=2, value=b['building']).font = body_font
        ws7.cell(row=i, column=3, value=b['count']).font = body_font
        ws7.cell(row=i, column=4, value=f"{b['pct']}%").font = body_font

    # ===== Sheet 8: 备件出库详情 =====
    stock_data = data.get('stock_outbound', {})
    ws8 = wb.create_sheet('备件出库')
    ws8.merge_cells('A1:F1')
    ws8['A1'].value = f"{year}年{month}月 备件出库明细"
    ws8['A1'].font = subtitle_font

    style_header(ws8, 3, 6)
    for i, h in enumerate(['时间', '物品名称', '数量', '目标科室', '经办人', '备注'], 1):
        ws8.cell(row=3, column=i, value=h)
    ws8.column_dimensions['A'].width = 14
    ws8.column_dimensions['B'].width = 22
    ws8.column_dimensions['C'].width = 10
    ws8.column_dimensions['D'].width = 20
    ws8.column_dimensions['E'].width = 14
    ws8.column_dimensions['F'].width = 20

    for i, d in enumerate(stock_data.get('details', []), 4):
        ws8.cell(row=i, column=1, value=d['time']).font = body_font
        ws8.cell(row=i, column=2, value=d['part']).font = body_font
        ws8.cell(row=i, column=3, value=f"{d['qty']} {d['unit']}").font = bold_font
        ws8.cell(row=i, column=4, value=d['department']).font = body_font
        ws8.cell(row=i, column=5, value=d['operator']).font = body_font
        ws8.cell(row=i, column=6, value='').font = body_font

    # 出库汇总（按科室）
    summary_start = len(stock_data.get('details', [])) + 5
    ws8.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=6)
    ws8.cell(row=summary_start, column=1, value='按科室汇总').font = subtitle_font
    sr = summary_start + 1
    style_header(ws8, sr, 3)
    for i, h in enumerate(['科室', '出库次数', '出库数量'], 1):
        ws8.cell(row=sr, column=i, value=h)
    for i, ds in enumerate(stock_data.get('dept_stats', []), sr + 1):
        ws8.cell(row=i, column=1, value=ds['department']).font = body_font
        ws8.cell(row=i, column=2, value=ds['count']).font = body_font
        ws8.cell(row=i, column=3, value=ds['items']).font = bold_font

    # 统一加边框
    for ws_ in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8]:
        apply_border(ws_, ws_.max_row, ws_.max_column)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    log_audit('export', 'report', current_user.display_name or current_user.username,
              target_desc=f'导出{ov["label"]}报表')
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{ov["label"]}IT运维报表.xlsx',
    )
