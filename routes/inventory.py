"""盘点管理：PC端盘点任务管理、盘盈盘亏核对"""
from datetime import datetime
import json, io, os
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, g, send_file
from flask_login import login_required, current_user
from models import db, InventoryTask, InventoryItem, Asset, AssetLog, log_audit
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

inv_bp = Blueprint('inventory', __name__, url_prefix='/inventory')


# ===================== 盘点任务列表 =====================

@inv_bp.route('/')
@login_required
def list_view():
    """盘点任务列表"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')
    query = InventoryTask.query
    if status_filter:
        query = query.filter(InventoryTask.status == status_filter)
    pagination = query.order_by(InventoryTask.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    return render_template('inventory/list.html', pagination=pagination,
                           tasks=pagination.items, status_filter=status_filter)


# ===================== 新建盘点 =====================

@inv_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """新建盘点任务"""
    if not current_user.is_admin:
        flash('仅管理员可操作', 'danger')
        return redirect(url_for('inventory.list_view'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        scope = request.form.get('scope', 'building')
        if not name:
            flash('请输入盘点名称', 'danger')
            return render_template('inventory/create.html')

        # 生成盘点编号
        now = datetime.now()
        count = InventoryTask.query.count() + 1
        task_no = f'PD{now.strftime("%Y%m%d")}-{count:03d}'

        # 统计应盘数量（按当前医院）
        total = Asset.query.count()

        task = InventoryTask(
            task_no=task_no,
            name=name,
            scope=scope,
            status='pending',
            total_count=total,
            operator=current_user.display_name,
            start_time=now,
            hospital_id=getattr(g, 'hospital_id', 1),
        )
        db.session.add(task)
        db.session.commit()
        log_audit('create', 'inventory', current_user.display_name,
                  target_id=task.id, target_desc=f'新建盘点 {task.name}')
        flash(f'盘点任务「{name}」已创建，编号 {task_no}', 'success')
        return redirect(url_for('inventory.detail', task_id=task.id))

    return render_template('inventory/create.html')


# ===================== 盘点详情 =====================

@inv_bp.route('/<int:task_id>')
@login_required
def detail(task_id):
    """盘点详情：按楼区或科室分类展示"""
    task = InventoryTask.query.get_or_404(task_id)
    page = request.args.get('page', 1, type=int)
    result_filter = request.args.get('result', '')
    search = request.args.get('search', '').strip()

    query = InventoryItem.query.filter_by(task_id=task_id)
    if result_filter:
        query = query.filter(InventoryItem.result == result_filter)
    if search:
        like = f'%{search}%'
        query = query.filter(InventoryItem.asset_no.ilike(like))

    pagination = query.order_by(InventoryItem.scanned_at.desc()).paginate(
        page=page, per_page=30, error_out=False
    )

    # 全部盘点项（供按楼区/按科室展示，不受分页限制）
    all_items = InventoryItem.query.filter_by(task_id=task_id).order_by(
        InventoryItem.scanned_at.desc()
    ).all()

    # 按楼区/科室统计
    buildings = {}
    departments = {}
    for item in InventoryItem.query.filter_by(task_id=task_id).all():
        b = item.building or '未知'
        d = item.department or '未知'
        buildings.setdefault(b, {'total': 0, 'normal': 0, 'issue': 0, 'new': 0})
        departments.setdefault(d, {'total': 0, 'normal': 0, 'issue': 0, 'new': 0})
        buildings[b]['total'] += 1
        departments[d]['total'] += 1
        if item.result == 'normal':
            buildings[b]['normal'] += 1
            departments[d]['normal'] += 1
        elif item.result == 'issue':
            buildings[b]['issue'] += 1
            departments[d]['issue'] += 1
        elif item.result == 'new':
            buildings[b]['new'] += 1
            departments[d]['new'] += 1

    # 盘盈盘亏分析
    surplus_loss = _calc_surplus_loss(task)

    return render_template('inventory/detail.html', task=task,
                           pagination=pagination, items=pagination.items,
                           all_items=all_items,
                           result_filter=result_filter, search=search,
                           buildings=buildings, departments=departments,
                           surplus_loss=surplus_loss)


# ===================== 盘点核对 =====================

@inv_bp.route('/<int:task_id>/review')
@login_required
def review(task_id):
    """盘点核对页面：左右对比，逐条确认"""
    task = InventoryTask.query.get_or_404(task_id)
    return render_template('inventory/review.html', task=task)


@inv_bp.route('/api/<int:task_id>/review-data')
@login_required
def review_data(task_id):
    """获取核对数据（JSON）"""
    task = InventoryTask.query.get_or_404(task_id)
    items = InventoryItem.query.filter_by(task_id=task_id).order_by(
        InventoryItem.scanned_at.desc()
    ).all()

    COMPARE_FIELDS = [
        ('device_type', '设备类型'),
        ('brand', '品牌'),
        ('model_no', '型号'),
        ('sn', '序列号'),
        ('department', '科室'),
        ('building', '楼区'),
        ('floor', '楼层'),
        ('location', '位置'),
    ]

    normal_list, issue_list, new_list = [], [], []

    for item in items:
        d = {
            'id': item.id,
            'asset_no': item.asset_no,
            'result': item.result,
            'confirmed': item.confirmed,
            'confirmed_at': item.confirmed_at.strftime('%m-%d %H:%M') if item.confirmed_at else '',
            'confirmed_by': item.confirmed_by or '',
            'scanned_by': item.scanned_by or '',
            'scanned_at': item.scanned_at.strftime('%m-%d %H:%M') if item.scanned_at else '',
            'notes': item.notes or '',
            'fields': {},
        }
        for key, label in COMPARE_FIELDS:
            val = getattr(item, key, '') or ''
            d['fields'][key] = {'label': label, 'value': val}

        # 匹配资产
        asset = None
        if item.asset_id:
            asset = db.session.get(Asset, item.asset_id)
        if not asset and item.asset_no:
            asset = Asset.query.filter_by(asset_no=item.asset_no, hospital_id=task.hospital_id).first()

        d['asset'] = None
        d['diff_fields'] = []
        if asset:
            d['asset'] = {
                'id': asset.id,
                'asset_no': asset.asset_no,
                'device_type': asset.device_type or '',
                'brand': asset.brand or '',
                'model_no': asset.model_no or '',
                'sn': asset.sn or '',
                'department': asset.department or '',
                'building': asset.building or '',
                'floor': asset.floor or '',
                'location': asset.location or '',
            }
            # 对比差异
            for key, label in COMPARE_FIELDS:
                old_val = (getattr(asset, key, '') or '').strip()
                new_val = (getattr(item, key, '') or '').strip()
                if old_val != new_val and new_val:
                    d['diff_fields'].append({
                        'key': key, 'label': label,
                        'old': old_val, 'new': new_val,
                    })

        if item.result == 'normal':
            normal_list.append(d)
        elif item.result == 'issue':
            issue_list.append(d)
        elif item.result == 'new':
            new_list.append(d)

    return jsonify({
        'task': {
            'id': task.id,
            'name': task.name,
            'task_no': task.task_no,
            'status': task.status,
        },
        'normal': normal_list,
        'issue': issue_list,
        'new': new_list,
        'total': len(items),
        'normal_count': len(normal_list),
        'issue_count': len(issue_list),
        'new_count': len(new_list),
    })


@inv_bp.route('/api/item/<int:item_id>/confirm', methods=['POST'])
@login_required
def confirm_item(item_id):
    """确认单条盘点记录（支持采纳用户选择的字段值）"""
    item = InventoryItem.query.get_or_404(item_id)
    data = request.get_json() or {}
    confirm = data.get('confirmed', True)
    field_values = data.get('field_values', {})  # 用户选择的最终值

    # 应用用户选择的字段值
    editable_fields = ['device_type', 'brand', 'model_no', 'sn', 'department', 'building', 'floor', 'location']
    for key, val in field_values.items():
        if key in editable_fields:
            setattr(item, key, val or '')

    item.confirmed = confirm
    if confirm:
        item.confirmed_at = datetime.now()
        item.confirmed_by = current_user.display_name or current_user.username

        # 异常项：先记录旧值 → 更新资产台账 → 记录日志
        if item.result == 'issue' and item.asset_id and field_values:
            asset = Asset.query.get(item.asset_id)
            if asset:
                old_vals = {}
                new_vals = {}
                for key, val in field_values.items():
                    if key in editable_fields and val is not None:
                        old_vals[key] = getattr(asset, key, '')
                        setattr(asset, key, val)
                        new_vals[key] = val
                if old_vals:
                    AssetLog(
                        asset_id=asset.id, action='inventory_update',
                        old_value=json.dumps(old_vals, ensure_ascii=False),
                        new_value=json.dumps(new_vals, ensure_ascii=False),
                        operator=current_user.display_name or current_user.username,
                    )
                    log_audit('update', 'asset', current_user.display_name,
                              target_id=asset.id,
                              target_desc=f'盘点修正: {asset.asset_no} → {", ".join(old_vals.keys())}')

        # 新盘资产：确认后自动入库
        if item.result == 'new':
            existing = Asset.query.filter_by(
                asset_no=item.asset_no,
                hospital_id=item.hospital_id
            ).first()
            if not existing:
                new_asset = Asset(
                    asset_no=item.asset_no,
                    device_type=item.device_type or '',
                    brand=item.brand or '',
                    model_no=item.model_no or '',
                    sn=item.sn or '',
                    department=item.department or '',
                    building=item.building or '',
                    floor=item.floor or '',
                    location=item.location or '',
                    hospital_id=item.hospital_id or 1,
                    inventory_status='',
                )
                db.session.add(new_asset)
                db.session.flush()
                item.asset_id = new_asset.id
                # 新盘入库日志
                AssetLog(
                    asset_id=new_asset.id, action='inventory_create',
                    old_value='', new_value=json.dumps({
                        'asset_no': item.asset_no, 'device_type': item.device_type,
                        'brand': item.brand, 'model_no': item.model_no,
                        'department': item.department, 'building': item.building,
                        'floor': item.floor, 'location': item.location,
                    }, ensure_ascii=False),
                    operator=current_user.display_name or current_user.username,
                )
                log_audit('create', 'asset', current_user.display_name,
                          target_id=new_asset.id,
                          target_desc=f'盘点入库: {item.asset_no}')

    # 通用盘点确认日志
    log_audit('confirm', 'inventory_item', current_user.display_name,
              target_id=item.id,
              target_desc=f'盘点确认: {item.asset_no} ({item.result})')

    db.session.commit()
    return jsonify({'ok': True, 'confirmed': confirm})


@inv_bp.route('/<int:task_id>/finish', methods=['POST'])
@login_required
def finish(task_id):
    """结束盘点（新版本：校验所有异常/新盘均已确认）"""
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    task = InventoryTask.query.get_or_404(task_id)
    if task.status == 'completed':
        return jsonify({'ok': False, 'msg': '该盘点已结束'}), 400

# 检查未确认的异常/新盘
    pending = InventoryItem.query.filter(
        InventoryItem.task_id == task_id,
        InventoryItem.result.in_(['issue', 'new']),
        InventoryItem.confirmed == False,
    ).count()
    if pending > 0:
        return jsonify({'ok': False, 'msg': f'还有 {pending} 条异常/新盘未确认', 'pending': pending}), 400

    task.status = 'completed'
    task.end_time = datetime.now()
    task.scanned_count = InventoryItem.query.filter_by(task_id=task_id).count()
    task.normal_count = InventoryItem.query.filter_by(task_id=task_id, result='normal').count()
    task.issue_count = InventoryItem.query.filter_by(task_id=task_id, result='issue').count()
    task.new_asset_count = InventoryItem.query.filter_by(task_id=task_id, result='new').count()
    db.session.commit()
    log_audit('update', 'inventory', current_user.display_name,
              target_id=task.id, target_desc=f'结束盘点 {task.name}')
    return jsonify({'ok': True, 'msg': '盘点已完成'})




@inv_bp.route('/api/<int:task_id>/batch-confirm-normal', methods=['POST'])
@login_required
def batch_confirm_normal(task_id):
    """一键确认全部正常项"""
    items = InventoryItem.query.filter_by(
        task_id=task_id, result='normal', confirmed=False
    ).all()
    now = datetime.now()
    operator = current_user.display_name or current_user.username
    count = 0
    for item in items:
        item.confirmed = True
        item.confirmed_at = now
        item.confirmed_by = operator
        log_audit('confirm', 'inventory_item', operator,
                  target_id=item.id,
                  target_desc=f'批量确认: {item.asset_no} (normal)')
        count += 1
    db.session.commit()
    return jsonify({'ok': True, 'count': count, 'msg': f'已确认 {count} 项正常资产'})


@inv_bp.route('/api/<int:task_id>/unscanned')
@login_required
def unscanned_assets(task_id):
    """获取未盘资产列表"""
    task = InventoryTask.query.get_or_404(task_id)
    scanned_ids = set()
    for item in InventoryItem.query.filter_by(task_id=task_id).all():
        if item.asset_id:
            scanned_ids.add(item.asset_id)
    query = Asset.query
    if task.scope == 'building':
        b = request.args.get('building', '')
        if b:
            query = query.filter(Asset.building == b)
    total = query.count()
    unscanned = query.filter(~Asset.id.in_(scanned_ids)).all() if scanned_ids else query.all()
    result = [{
        'id': a.id, 'asset_no': a.asset_no,
        'device_type': a.device_type, 'brand': a.brand,
        'model_no': a.model_no, 'department': a.department,
        'building': a.building, 'floor': a.floor, 'location': a.location,
        'status': a.status,
    } for a in unscanned]
    return jsonify({'ok': True, 'total': total, 'scanned': len(scanned_ids), 'unscanned': result, 'count': len(result)})


@inv_bp.route('/api/<int:task_id>/undo-confirm', methods=['POST'])
@login_required
def undo_confirm(task_id):
    """撤销最近一次确认（5分钟内有效）"""
    data = request.get_json() or {}
    item_id = data.get('item_id', 0)
    item = InventoryItem.query.get(item_id)
    if not item:
        return jsonify({'ok': False, 'msg': '记录不存在'}), 404
    if not item.confirmed:
        return jsonify({'ok': False, 'msg': '该记录尚未确认'}), 400
    if not item.confirmed_at:
        return jsonify({'ok': False, 'msg': '无法获取确认时间'}), 400
    if (datetime.now() - item.confirmed_at).total_seconds() > 300:
        return jsonify({'ok': False, 'msg': '已超过5分钟，无法撤销'}), 400

    # 如果之前更新了资产台账，需要回滚
    if item.result == 'issue' and item.asset_id:
        asset = Asset.query.get(item.asset_id)
        if asset:
            log_audit('update', 'asset', current_user.display_name,
                      target_id=asset.id,
                      target_desc=f'盘点撤销恢复: {asset.asset_no}')
    if item.asset_id and item.result == 'new':
        # 新盘入库的资产，删除它
        asset = Asset.query.get(item.asset_id)
        if asset:
            db.session.delete(asset)

    item.confirmed = False
    item.confirmed_at = None
    item.confirmed_by = None
    log_audit('undo', 'inventory_item', current_user.display_name,
              target_id=item.id,
              target_desc=f'撤销确认: {item.asset_no}')
    db.session.commit()
    return jsonify({'ok': True, 'msg': '已撤销确认'})


@inv_bp.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    """盘点看板：上次盘点 vs 本次盘点对比"""
    tasks = InventoryTask.query.filter(
        InventoryTask.status == 'completed'
    ).order_by(InventoryTask.end_time.desc()).limit(2).all()

    result = []
    for t in tasks:
        total = (t.normal_count or 0) + (t.issue_count or 0) + (t.new_asset_count or 0)
        result.append({
            'id': t.id, 'name': t.name, 'task_no': t.task_no,
            'end_time': t.end_time.strftime('%m-%d %H:%M') if t.end_time else '',
            'total': total,
            'normal': t.normal_count or 0,
            'issue': t.issue_count or 0,
            'new_assets': t.new_asset_count or 0,
            'normal_pct': round((t.normal_count or 0) / total * 100, 1) if total else 0,
            'issue_pct': round((t.issue_count or 0) / total * 100, 1) if total else 0,
        })
    return jsonify({'ok': True, 'tasks': result})




@inv_bp.route('/<int:task_id>/export')
@login_required
def export_excel(task_id):
    """导出盘点报表"""
    task = InventoryTask.query.get_or_404(task_id)
    items = InventoryItem.query.filter_by(task_id=task_id).order_by(
        InventoryItem.result, InventoryItem.asset_no
    ).all()

    wb = openpyxl.Workbook()
    
    # ===== 汇总页 =====
    ws = wb.active
    ws.title = '盘点汇总'
    header_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    blue_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

    # 标题
    ws.merge_cells('A1:H1')
    ws['A1'] = f'盘点报表 - {task.name}'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f'任务编号: {task.task_no}    操作人: {task.operator}    开始: {task.start_time.strftime("%%Y-%%m-%%d %%H:%%M") if task.start_time else "-"}'
    ws['A3'] = f'结束时间: {task.end_time.strftime("%%Y-%%m-%%d %%H:%%M") if task.end_time else "-"}    状态: {"已完成" if task.status == "completed" else "进行中"}'
    
    # 统计
    ws['A5'] = '统计项'
    ws['B5'] = '数量'
    for cell in [ws['A5'], ws['B5']]:
        cell.font = header_font
        cell.border = thin_border
    stats = [
        ('总盘点数', task.scanned_count or 0),
        ('正常', task.normal_count or 0),
        ('异常', task.issue_count or 0),
        ('新盘', task.new_asset_count or 0),
    ]
    for i, (label, val) in enumerate(stats):
        ws.cell(row=6+i, column=1, value=label).border = thin_border
        ws.cell(row=6+i, column=2, value=val).border = thin_border

    # ===== 正常项 =====
    ws2 = wb.create_sheet('正常项')
    ws2.append(['资产编号', '设备类型', '品牌', '型号', '科室', '楼区', '楼层', '位置', '扫码人', '扫码时间'])
    for row in ws2.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.border = thin_border
            cell.fill = green_fill
    for item in [i for i in items if i.result == 'normal']:
        ws2.append([
            item.asset_no, item.device_type, item.brand,
            item.model_no, item.department, item.building,
            item.floor, item.location,
            item.scanned_by, item.scanned_at.strftime('%m-%d %H:%M') if item.scanned_at else '',
        ])

    # ===== 异常项 =====
    ws3 = wb.create_sheet('异常项')
    ws3.append(['资产编号', '字段', '档案值', '扫码值', '采纳值', '确认人', '确认时间'])
    for row in ws3.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.border = thin_border
            cell.fill = red_fill
    for item in [i for i in items if i.result == 'issue']:
        # 比较字段
        asset = Asset.query.get(item.asset_id) if item.asset_id else None
        fields = [('device_type','设备类型'),('brand','品牌'),('model_no','型号'),
                  ('department','科室'),('building','楼区'),('floor','楼层'),('location','位置')]
        for key, label in fields:
            scan_val = ''
            if item.fields and key in item.fields:
                scan_val = item.fields[key].get('value', '')
            arch_val = getattr(asset, key, '') if asset else ''
            if scan_val != arch_val:
                ws3.append([
                    item.asset_no, label, arch_val, scan_val,
                    scan_val or arch_val,
                    item.confirmed_by or '',
                    item.confirmed_at.strftime('%m-%d %H:%M') if item.confirmed_at else '',
                ])
        if not item.asset_id:
            ws3.append([item.asset_no, '(资产已删除)', '', '', '', item.confirmed_by or '', ''])

    # ===== 新盘 =====
    ws4 = wb.create_sheet('新盘资产')
    ws4.append(['资产编号', '设备类型', '品牌', '型号', '科室', '楼区', '楼层', '位置', '扫码人', '确认人', '确认时间'])
    for row in ws4.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.border = thin_border
            cell.fill = blue_fill
    for item in [i for i in items if i.result == 'new']:
        ws4.append([
            item.asset_no, item.device_type, item.brand,
            item.model_no, item.department, item.building,
            item.floor, item.location,
            item.scanned_by, item.confirmed_by or '',
            item.confirmed_at.strftime('%m-%d %H:%M') if item.confirmed_at else '',
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'盘点报表_{task.task_no}_{task.name}.xlsx'
    )

# ===================== 删除盘点 =====================

@inv_bp.route('/<int:task_id>/delete', methods=['POST'])
@login_required
def delete(task_id):
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    task = InventoryTask.query.get_or_404(task_id)
    log_audit('delete', 'inventory', current_user.display_name,
              target_id=task.id, target_desc=f'删除盘点 {task.name}')
    db.session.delete(task)
    db.session.commit()
    return jsonify({'ok': True})


# ===================== 盘盈盘亏分析 =====================

def _calc_surplus_loss(task):
    """计算盘盈盘亏"""
    result = {
        'by_building': {},
        'by_department': {},
        'summary': {'surplus': 0, 'loss': 0, 'normal': 0, 'new_assets': 0},
    }

    # 已扫描的资产编号
    scanned_asset_nos = set()
    for item in InventoryItem.query.filter_by(task_id=task.id).all():
        if item.asset_no:
            scanned_asset_nos.add(item.asset_no)
        if item.result == 'new':
            result['summary']['new_assets'] += 1
        elif item.result == 'normal':
            result['summary']['normal'] += 1

    # 系统中的所有资产，检查哪些没被扫到（漏盘 = 盘亏）
    all_assets = Asset.query.all()
    all_nos = {a.asset_no for a in all_assets if a.asset_no}
    missing_nos = all_nos - scanned_asset_nos

    result['summary']['loss'] = len(missing_nos)
    result['summary']['surplus'] = result['summary']['new_assets']  # 新盘资产视为盘盈

    # 按楼区统计漏盘
    for asset in all_assets:
        if asset.asset_no in missing_nos:
            b = asset.building or '未知'
            d = asset.department or '未知'
            result['by_building'].setdefault(b, {'lost': 0, 'surplus': 0})
            result['by_building'][b]['lost'] += 1
            result['by_department'].setdefault(d, {'lost': 0, 'surplus': 0})
            result['by_department'][d]['lost'] += 1

    # 按楼区统计盘盈
    for item in InventoryItem.query.filter_by(task_id=task.id, result='new').all():
        b = item.building or '未知'
        d = item.department or '未知'
        result['by_building'].setdefault(b, {'lost': 0, 'surplus': 0})
        result['by_building'][b]['surplus'] += 1
        result['by_department'].setdefault(d, {'lost': 0, 'surplus': 0})
        result['by_department'][d]['surplus'] += 1

    return result


# ===================== 盘点明细删除 =====================

@inv_bp.route('/item/<int:item_id>/delete', methods=['POST'])
@login_required
def delete_item(item_id):
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    item = InventoryItem.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'ok': True})


# ===================== 重新统计 =====================

@inv_bp.route('/<int:task_id>/recalc', methods=['POST'])
@login_required
def recalc(task_id):
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': '权限不足'}), 403
    task = InventoryTask.query.get_or_404(task_id)
    task.scanned_count = InventoryItem.query.filter_by(task_id=task_id).count()
    task.normal_count = InventoryItem.query.filter_by(task_id=task_id, result='normal').count()
    task.issue_count = InventoryItem.query.filter_by(task_id=task_id, result='issue').count()
    task.new_asset_count = InventoryItem.query.filter_by(task_id=task_id, result='new').count()
    db.session.commit()
    return jsonify({'ok': True, 'msg': '已重新统计'})